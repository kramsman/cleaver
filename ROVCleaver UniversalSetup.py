""" ROVCleaver UniCode is a complete redo under the hood!  Taken from working V16.3
    Setup file drastically changed - variable definitions and documentation are contained in setup - no hardcoded rows!
    ROV. structure replace with dictionary for ease in adding and removing entries
    Allows moving and sorting rows.
    Some repetitious code - like pivot tables 1-5 - are looped.
"""
# 12/6/23 Add ability to split by a group variable in addition to county for multi scripts.  Try git branch with
# group_split


# Temporarily save curl text to enter in terminal
# curl https://raw.githubusercontent.com/kramsman/ROVCleaver/master/ROVCleaver%20UniversalSetup.py?token=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 --output /Users/Denise/Downloads/dest.py

# copy private to ROVCleaver_Prod w $HOME
# curl https://raw.githubusercontent.com/kramsman/ROVCleaver/master/ROVCleaver%20UniversalSetup.py?token=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 --output $HOME/Library/CloudStorage/Dropbox/Postcard Files/PythonProgs/ROVCleaver_Prod/x.py

# ROVCleaver_prod = https://www.dropbox.com/scl/fo/1pp2mg69019h2dijxvbke/h?rlkey=l6cs53fnqw4c2iqp6crpfjl24&dl=0
# works on public
# curl https://raw.githubusercontent.com/kramsman/ROVCleaver/master/ROVCleaver%20UniversalSetup.py --output /Users/Denise/Downloads/dest.py

# wget https://raw.githubusercontent.com/kramsman/ROVCleaver/master/ROVCleaver%20UniversalSetup.py?token=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
# wget --directory-prefix=/var/cache/foobar/ [...]

# copied from terminal
#  wget -P /Users/Denise/Downloads https://raw.githubusercontent.com/kramsman/ROVCleaver/master/ROVCleaver%20UniversalSetup.py?token=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
#  wget https://raw.githubusercontent.com/kramsman/ROVCleaver/master/ROVCleaver%20UniversalSetup.py?token=GHSAT0AAAAAACLDGV6B4MYVBVQD2DP7LE5OZLQUEHA -P /Users/Denise/Downloads -O x.py
#  wget https://raw.githubusercontent.com/kramsman/ROVCleaver/master/ROVCleaver%20UniversalSetup.py?token=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
# ------
#  to remove unused functions used vulture.  In Pycharm terminal: vulture 'xxx.py'

# TODO: put filetype in to select file and specify xlsx for Setup.
# TODO: change all tkinter & pymsgbox to simplegui

# FIXME: list of imported code (first_code, etc) doesn't print on console if error found (raises first). Maybe use
#  try/except and print file (not compile) to screen if error?
# TODO redo formatcopy using pathlib / remove os.path.join
# TODO can we skip padding variable lists?  zip will work on shortest one
# FIXME make sure Format merge works - which fields missing are ok?
# TODO put .py code created from first, middle, last sheets in root dir (with setup) rather than exe dir (ROVCleaver)
#   so different runs of Cleaver don't collide.  Problem is compiled object goes to exe so is not found when moved.
# TODO calc max_pll_group and pass to first, middle, last.  code remove if < max.

import ast
import collections
import datetime
import importlib
import inspect
import itertools
import math  # for ceil function
import os
import pathlib
import re
import shutil  # to copy file from other dirs
import webbrowser
from datetime import datetime
from itertools import islice  # to skip 1st row of iterated spreadsheet
from pathlib import *
from tkinter import Tk  # from tkinter import Tk for Python 3.x
from typing import Union

# from tkinter.filedialog import askopenfilename

import numpy as np
import pandas as pd
import pymsgbox
from openpyxl import load_workbook
from openpyxl.styles import Font
import PySimpleGUI as sg
import sys
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.utils.cell import column_index_from_string
import json
from loguru import logger

log_level = "DEBUG"  # used for log file; screen set to INFO. TRACE, DEBUG, INFO, WARNING, ERROR

# INITIAL_CAMPAIGN_DIR = os.path.expanduser(r"/Users/Denise/Dropbox/Postcard Files/InputFiles/Campaigns")
INITIAL_CAMPAIGN_DIR = pathlib.Path("~/Dropbox/Postcard Files/TestInputFiles/TestCampaigns/").expanduser()
MAIN_ZIP_FILE = 'zip-codes-database-DELUXE-BUSINESS.csv'
MULTI_COUNTY_ZIP_FILE = 'zip-codes-database-MULTI-COUNTY.csv'
ZIP_TO_COUNTY_LIST_FILE = 'Zip_To_County_List_dict.py'  # file where the numeric zip to county list is stored (ie
# 1011: ['hampden', 'hampshire'])
PROP_CONCENTRATION = 50
ZIP_CONCENTRATION = 10

ROV_SETUP = {}
logger.info(f"{id(ROV_SETUP)=}")

REMOVE_XL_FROM_SETUP = False  # remove or keep the temporary 'wor' variables prefixed with 'XL_'

FIELD_DEF_COL_ALPHA = 'E'  # field_keep is assumed one column left
FIELD_DEF_COL_NUMERIC = column_index_from_string(coordinate_from_string(FIELD_DEF_COL_ALPHA + '1')[0]) - 1  # -1 to
# zero index

DATA_STARTS_COL_ALPHA = 'G'  # field_keep is assumed one column left
DATA_STARTS_COL_NUMERIC = column_index_from_string(coordinate_from_string(DATA_STARTS_COL_ALPHA + '1')[0]) - 1  # -1 to


def pad_list(my_list, to_len, pad_val=""):
    """ pad list with an element to a given length """
    list_len = len(my_list)
    elem_needed = to_len - list_len
    padded_list = my_list + [pad_val] * elem_needed
    return padded_list


def exit_yes_no(msg, title=None, display_exiting=False):
    """ makes this choice to continue one line"""
    if not title:
        title = "Exit?"
    choice = pymsgbox.confirm(msg, title, ['Yes', 'No'])
    if choice == "No":
        if display_exiting:
            pymsgbox.alert("Exiting", "Alert")
        logger.debug("here'")
        exit()


def exit_yes(msg: str, title: str = None, *, errmsg: str = None) -> None:
    """ exits program after giving user a popup window and raising an error. """
    msg = (msg + "\n\n\nExiting." +
           f"\n\nCalled from {calling_func(level=3)}"
           f"\nCalled from {calling_func(level=2)}"
           f"\nCalled from {calling_func(level=1)}"
           )
    if not errmsg:
        errmsg = msg.replace("\n", " ")  # dont fill the console with linefeeds
    if not title:
        title = "** Exiting Program **"
    logger.debug("here")
    pymsgbox.alert(msg, title)
    raise Exception(errmsg)


def is_number(s: str) -> bool:
    """  Used as check, particularly before trying to set zip to numeric for lookup.
    expects param to be a string to trap all types of data.   """
    if s == np.NAN:  # this is needed- np.nan are int which are numbers
        return False
    try:
        float(s)
        return True
    except ValueError:
        return False


# def clean_path(original_path):
#     """ Return a path cleaned: adds ~ to start, returns path object using os.path """
#     new_path = original_path
#     if new_path[1] != '~':
#         new_path = "~" + new_path
#     if original_path[-1] != '/':
#         new_path = new_path + '/'
#     new_path = os.path.expanduser(new_path)
#     return new_path


# def max_used_col(ws, rw):
#     """ Returns the column number (1 indexed) maximum non-none column in the input row of a sheet. """
#     mxcol = 0
#     for cell in reversed(ws[rw]):
#         if cell.value is not None:
#             mxcol = cell.col_idx
#             break
#     return mxcol


# def max_used_col_in_row(row):
#     """ Returns the column number (1 indexed) maximum non-none column in the input row of a sheet. """
#     for index, cell in enumerate(reversed(row)):
#         if cell is not None:
#             break
#     # return cell.col_idx
#     return cell.index


def row_to_list(row):
    """ Returns the column number (1 indexed) maximum non-none column in the input row of a sheet. """
    for index, cell in enumerate(reversed(row)):
        if cell.value is not None:
            break
    row_list = [val.value for val in islice(row, 0, cell.col_idx)]
    return row_list


def range_to_list(ws, start_row, end_row, start_col, end_col):
    """ converts range of cells to a list of values
    V2.0 Worksheet to list. if one col, multi lists in list rather than one list of elements.  Best?
    """
    # tried using openpyl's cells = setup['A18': 'E19'] but parsing letter cols would be ugly so iterate rows.
    # openpyxl does not offer cells((r,c),(r,c)) using numerics
    final_list = []
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col,
                            max_col=end_col):
        row_list = []
        for cell in row:
            if cell.value is not None:
                row_list.append(str(cell.value).lower().strip())
            else:
                row_list.append("")
        final_list.append(row_list)
    if start_row == end_row:  # produce a one dimensional list instead of a one element list in a list
        final_list = final_list[0]
    return final_list


def clean_field(fld, case_convert='lower'):
    """
    returns a string in lower, strip, no space, no -, no ., no '
    can be used with dataframe like IP['clean2'] = IP['B'].apply(clean_field, convert_case='keep')
    1/28/23 added optional parameter convert_case defaulting to lower, as was done before, but allowing 'upper' or
    'keep'.
    """
    return_fld = str(fld).strip().replace(" ", "").replace("'", "").replace(".", "").replace("-", "")
    if case_convert == 'lower':
        return_fld = return_fld.lower()
    elif case_convert == 'upper':
        return_fld = return_fld.upper()
    elif case_convert == 'keep':
        pass
    else:
        exit_yes(f"wrong value fed to clean_field parameter case_convert, '{case_convert}' - exiting")
    return return_fld


def autosize_xls_cols(ws):
    """ BEKs routine that works on the wks rather than df.  Datetime format set to width of 10. """
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                if cell.data_type == 'd':
                    date_width = 10
                else:
                    date_width = len(str(cell.value))
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), date_width))

    for col, value in dims.items():
        ws.column_dimensions[col].width = value + 1


def replace_boolean_column_vals(df, field_list):
    """ Check field in list for literal TRUE or FALSE because it gets converted to boolean and kills the program"""
    for field in field_list:
        if field in df.columns:
            df.loc[(df[field].str.upper() == "TRUE") | (df[field].str.upper() == "FALSE"), field] = ''


def bad_file_exit(file, msg=None):
    """ checks for file existence and exits if not found"""
    if msg is None:
        msg = f"File:\n\n'{file}'\n\ndoes not exist."
    if not file.expanduser().exists():
        logger.debug("here")
        exit_yes(msg)


def bad_path_exit(path, msg=None):
    """ checks for directory existence and exits if not found"""
    if msg is None:
        msg = f"Directory:\n\n'{path}'\n\ndoes not exist."
    # if not Path(os.path.expanduser(path)).exists():  # need expanduser for ~; only os works (not pathlib)
    if not path.expanduser().exists():
        # pymsgbox.alert(msg, "** Exiting via bad_path_exit **")
        # # FIXME: close TKINTER window here.  https://stackoverflow.com/questions/8009176/function-to-close-the-window-in-tkinter
        # exit()
        logger.debug("here")
        exit_yes(msg)


def bad_path_create(path, msg=None):
    """ checks for directory existence and creates if not found"""
    if msg is None:
        msg = ("Directory:\n\n" + str(path) + "\n\ndoes not exist.  Creating." +
               "\n\nCalled from " + calling_func(level=2))
    if not os.path.isdir(path):
        logger.debug("here")
        pymsgbox.alert(msg, "Adding Directory via bad_path_create")
        os.makedirs(path)


def calling_func(level=0):
    """ returns the various levels of calling function.  0 is current, 1 is caller of current, etc """
    try:
        func = f"'{inspect.stack()[level][3]}', line #: {inspect.stack()[level][2]}"
    except Exception:
        logger.debug("here")
        func = f"** error ** inspect level too deep: {str(level)} called from {inspect.stack()[level][3]}"
    return func


def identify_duplicates(df, key, dupe_id_field):
    """
    creates a field in input df identifying duplicates as: X:not dupe;F:first,L:Last;D:other dupe;O:Other
    Parameters
    ----------
    df : input dataframe
    key : field checked as duplicate
    dupe_id_field : field in df filled with identifiers above
    """
    logger.info("Identifying duplicates")

    any_dupe_bool = df.duplicated([key], keep=False)  # all dupes as true/false
    # 'First' returns true for non-dupes as well as 'real' firsts, so must 'and' with all dupes
    first_bool = ~df.duplicated([key], keep='first') & any_dupe_bool  # first AND dupe as true
    last_bool = ~df.duplicated([key], keep='last') & any_dupe_bool  # last AND dupe as true

    # set different values to first and last so we can identify them as well as other dupes later
    first_numeric = np.where(first_bool, 2, 0)
    last_numeric = np.where(last_bool, 1, 0)
    any_dupe_numeric = np.where(any_dupe_bool, 1, 0)
    dupe_numeric = any_dupe_numeric + first_numeric + last_numeric  # sets each to different value
    dupe_alpha = [{0: 'X', 1: 'D', 2: 'L', 3: 'F', 4: 'O'}[element] for element in dupe_numeric]  #
    # note that setting non-dupe, 0, to blank above removes them from list so can not be merged into df
    if len(dupe_alpha) != len(df):
        logger.info(f"*** assigning duplicate identifier error:  "
              f"Orig dataframe has {len(df)} rows but merged list has {len(dupe_alpha)}")
        raise Exception
    df[dupe_id_field] = dupe_alpha

    return


# def split_tuples(tuple_list):
#     """
#     Accepts a list of tuples, ex for sort fields and ascending/descending, checks to make sure all list elements are
#     tuples, then splits them into list of tuples keys and list of tuple values
#     Parameters
#     ----------
#     tuple_list : list of tuples containing pairs of key/values
#
#     Returns
#     -------
#     key_list: list of tuples keys
#     val_list: list of tuple values
#     """
#     if tuple_list is None or tuple_list.strip() == '':
#         return None, None
#     else:
#         tuple_list = eval(tuple_list)
#         tuple_list = list(tuple_list)
#         # check for list items that are not tuples - user error
#         tuple_check_list = [True if isinstance(val, tuple) else False
#                                  for val in tuple_list]
#         if False in tuple_check_list:
#             exit_yes('List of tuples contained a non-tuple', 'Error in Tuple List')
#             raise ValueError
#         key_list = [my_tuple[0] for my_tuple in tuple_list]
#         val_list = [my_tuple[1] for my_tuple in tuple_list]
#
#         return key_list, val_list


def merge_into_format_file(orig_df, update_file, cur_path):
    # TODO:  THIS NEEDS TO BE RE-WRITTEN AND CHECKED.  Copied and some refactoring but no logi or working checked.
    """ updates fields in format file from update file"""
    logger.info("Merging into format file")

    update_file_w_path = os.path.join(cur_path, update_file)
    # NOTE: update file is expected to have headers row 1 unlike raw data which may have title lines at top.

    update_df = read_file_to_df(update_file_w_path, **{'sheet_name': 0, 'nrows': ROV_SETUP['rows_to_read_limit'],
                                                       'keep_default_na': False})

    update_df.columns = [x.strip().lower() for x in list(update_df.columns)]
    # rename df fields as lower() because col matching is case-sensitive

    if set(ROV_SETUP['update_field_list']).issubset(set(update_df.columns)):  # update field list all contained in df fields
        update_df = update_df[update_df.columns[update_df.columns.isin(ROV_SETUP['update_field_list'])]]
    else:  # some specified update fields not on df
        extraelems = set(ROV_SETUP['update_field_list']) - set(update_df.columns)
        exit_yes((f"Following field(s) in update list not in input.\n\n"
                  f"{os.linesep.join(extraelems)}"
                  ))

    # add the key fields specified in setup to both files
    # update_df['update_key'] = address + city
    update_df['update_key'] = eval(
        ROV_SETUP['update_file_key_formula'])  # TODO: try sorting by keys to speed up dropping duplicates
    orig_df['master_key'] = eval(ROV_SETUP['orig_file_key_formula'])

    # BEK2/22 hardcode field input as variable updateField to 'updateField'
    # update_df.rename(columns={updateField : 'update_field'}, inplace=True)  # why specify field name to create in setup then rename to hardcoded 'update_field'?

    # update_df.columns += '_updt' # add suffix to all columns names to keep separate from master. Prefix is df.columns = 'prefix_' + df.columns
    update_df.columns = [str(col) + '_updt' if col not in ['update_key'] else 'update_key' for col in
                         update_df.columns]  # add suffix to all columns names except 'update_key' to keep separate from master.

    # create df after dropping  duplicates based on key
    updt_w_duplicates_removed = update_df.drop_duplicates(subset=['update_key'])
    update_count_before_dups_removed = len(update_df)
    update_count_after_dups_removed = len(updt_w_duplicates_removed)
    num_dups_in_updt = update_count_before_dups_removed - update_count_after_dups_removed

    # Output a df of duplicate key values
    logger.debug("list of counts of updt duplicate key values in ", update_file)
    # dups = update_df.duplicated(subset=['update_key'], keep='first')  # not sure what this shows.  needed?
    dups_in_update = update_df[update_df.duplicated(subset=['update_key'],
                                                    keep=False)]  # Keep = false will show all dups; only first is kept above
    dups_in_update = dups_in_update.sort_values("update_key")
    logger.debug("list of updt duplicate key values (showing the first) in ", update_file)
    logger.debug(dups_in_update)  # does this work for df or do we need str() or something to avoid object?
    # write all duplicates to a file so we can take a look if desired
    file_of_dups = ROV_SETUP['format_path'] / "Duplicates" / \
                   ("UPDATE DUPLICATES " + str(PurePath(update_file).stem) + ".xlsx")
    dups_in_update.to_excel(file_of_dups, index=False)

    # must check for dups because key might not be unique (eg using truncated name)
    if num_dups_in_updt > 0:  # calculate above by comparing before and after de-dup
        exit_yes_no("Continue?  Dups in update will be removed.\n\n\nUpdate file\n" +
                    orig_df + "\n contains " + str(num_dups_in_updt) + " duplicates.",
                    "CHECK FOR DUPLICATE KEYS IN UPDATE",
                    display_exiting=False)

        logger.debug('Count before de-dup in update file', len(update_df))
        logger.debug('Count after de-dup in update file', len(updt_w_duplicates_removed))
        update_df = updt_w_duplicates_removed
        # print('New update_df without dupes - should be old updt_w_duplicates_removed ',update_df.shape[0])
        logger.debug('New update_df without dupes - should be old updt_w_duplicates_removed ', len(update_df))

    # update_df = update_df[['update_key', 'update_field']]  # keep only the key to prevent dup fields from being renamed in df with suffix (_x)

    master_w_dupkey_removed = orig_df.drop_duplicates(subset=['master_key'])  # new df with all duplicate master_key

    num_of_master_dups = len(orig_df) - len(master_w_dupkey_removed)

    # prompt if wanting to update if dupe on master present
    if num_of_master_dups > 0:
        exit_yes_no("Continue with dups?\n\n\nMaster file\n" + orig_df + "\ncontains " +
                    str(num_of_master_dups) + " duplicates.",
                    "CHECK FOR DUPLICATE KEYS IN MASTER",
                    display_exiting=False)

        # write dupes in master to file for reviewing
        file_of_dups = ROV_SETUP['format_path'] / "Duplicates" / \
                       ("MASTER DUPLICATES " + str(PurePath(orig_df).stem) + ".xlsx")
        # writer = pd.ExcelWriter(file_of_dups)
        orig_dups = orig_df[orig_df.duplicated(subset=['master_key'], keep=False)]
        orig_dups = orig_dups.sort_values("master_key")
        orig_dups.to_excel(file_of_dups, index=False)

    #### Do the actual update!
    # In perfect world could use ", validate='1:1'" in merge to ensure integrity, but input data has dups,
    # so code was added to point out how many and allow continuance
    # Get counts before and after merge
    hold_orig_address_count = len(orig_df)
    orig_df = pd.merge(orig_df, update_df, how="left", left_on=['master_key'], right_on=['update_key'])
    after_address_count = len(orig_df)
    logger.debug('Orig address file had records: ', hold_orig_address_count, 'Result of merge had records: ',
          after_address_count)

    orig_number_created = after_address_count - hold_orig_address_count
    if orig_number_created > 0:  # the merge added records
        exit_yes_no(str(orig_number_created) + " addresses were duplicated by merge.  Continue?",
                    "EXTRA RECORDS WERE CREATED BY MERGE",
                    display_exiting=False)


def zip_file_to_county_dict(zip_csv_path: Union[str, os.PathLike], xlsx_path: Union[str, os.PathLike]) -> dict:
    """
    Reads zip data from csv
    1.  writes an xlsx of unique state/county from purchased county data
    https://www.zip-codes.com/.
    2. returns a dictionary keying state/county to county_filename, countyToPrint, stateMixedCounty

    Parameters
    ----------
    zip_csv_path : csv of zip data: 80K rows with 93 fields, 'zip-codes-database-DELUXE-BUSINESS.csv'
    xlsx_path : path to xlsx 

    Returns
    -------
    dict : dictionary keying state/county to county_filename, countyToPrint, stateMixedCounty writes xlsx of unique
    states-counties
    """
    logger.debug("creating dictionary of zip to county")
    zip_rows_to_read = 999_999  # for testing
    # zip_rows_to_read = 999  # for testing

    # original county is uppercase
    main_zip_file = pd.read_csv(zip_csv_path, nrows=zip_rows_to_read, keep_default_na=False,
                                usecols=['State', 'County', 'CountyMixedCase'])

    main_zip_file.rename(
        columns={'ZipCode': 'zip', 'State': 'state', 'County': 'county', 'CountyMixedCase': 'county_mixedcase'},
        inplace=True)

    # military states like AA and AE have no county so remove
    main_zip_file = main_zip_file.loc[main_zip_file['county'].str.strip() != ""]

    # remove characters like ,- space
    main_zip_file['countyclean'] = main_zip_file['county'].apply(clean_field, case_convert='keep')
    main_zip_file['statecounty'] = main_zip_file['state'] + "-" + main_zip_file['countyclean']
    main_zip_file['county_filename'] = main_zip_file['county_mixedcase'].apply(clean_field, case_convert='keep')
    # add 'County' if 'City' is not at end of name
    main_zip_file['countyToPrint'] = np.where(main_zip_file['county_mixedcase'].str[-4:] != "City",
                                              main_zip_file['county_mixedcase'] + " County",
                                              main_zip_file['county_mixedcase'])
    main_zip_file['statecounty_mixed'] = main_zip_file['state'] + "-" + main_zip_file['county_filename']

    unique_county = main_zip_file.drop_duplicates(subset=['statecounty'], keep='last')
    sorted_unique_county = unique_county.sort_values(['statecounty'], ascending=[True])

    sorted_unique_county.to_excel(xlsx_path, index=False,
                                  columns=['statecounty', 'county_filename', 'countyToPrint', 'statecounty_mixed'])

    # create dict using zip function
    county_dict = dict([(k, [a, b, c]) for k, a, b, c in zip(sorted_unique_county['statecounty'],
                                                   sorted_unique_county['county_filename'],
                                                   sorted_unique_county['countyToPrint'],
                                                   sorted_unique_county['statecounty_mixed'])])

    return county_dict


def create_zip_to_county_list_dict(unique_zips, split_zips, text_file_for_created_dict):
    """
    Creates a dictionary with zip as key, list of counties as values using purchased zip data.
    Also writes the dictionary to a text file.
    Two input files: one-to-one zip/county and split zips.

    Data purchased from https://www.zip-codes.com/.
    Multi county and unique files are merged.
    Multi county does not contain unique zips.
    Unique contains multiple records for the same county for multiple cities in zip.
    """
    logger.info("creating zip to county lists")
    zip_rows_to_read = 999_999
    # zip_rows_to_read = 9_999  # for testing

    main_zip_temp = pd.read_csv(unique_zips, nrows=zip_rows_to_read, keep_default_na=False,
                                usecols=['State', 'County', 'ZipCode'])
    multi_county_temp = pd.read_csv(split_zips, nrows=zip_rows_to_read, keep_default_na=False,
                                    usecols=['State', 'County', 'ZipCode'])

    # combined_temp = main_zip_temp.append(multi_county_temp, ignore_index=True)
    combined_temp = pd.concat([main_zip_temp, multi_county_temp], ignore_index=True)

    combined_temp['countyclean'] = combined_temp['County'].apply(clean_field, case_convert='upper')

    combined_temp['statecounty'] = combined_temp['State'] + "-" + combined_temp['countyclean']

    # combined_temp2 = combined_temp[['ZipCode', 'countyclean']]  # keep only two cols
    combined_temp2 = combined_temp[['ZipCode', 'statecounty']]  # keep only two cols
    unique_zip_county = combined_temp2.drop_duplicates(subset=['ZipCode', 'statecounty'], keep='last')

    df_for_dict = unique_zip_county.groupby(["ZipCode"], as_index=False).agg({'statecounty': list})
    zip_to_county_list_dict = dict(df_for_dict.values.tolist())

    with open(ROV_SETUP['exe_path'] / text_file_for_created_dict, 'w') as f:
        print(zip_to_county_list_dict, file=f)

    return zip_to_county_list_dict


def find_header_row_in_file(file_with_path, header_string, header_col, sheet_name=None):
    """ identifies row with header by searching for header_string in header_col.  Used to skip blank and rows with titles.

    Parameters
    ----------
    file_with_path : input file being read in, csv or xlsx?
    header_string : string identifying header row, like 'pdiid'
    header_col : alpha col to search for string, 'B' 'AA'
    sheet_name : sheet name in input file in case multiple
    """
    if sheet_name is None:
        sheet_name = 0
    header_row = None

    # If the header identifying field is not in the first 30 rows assume something is wrong in the file
    df_temp = read_file_to_df(file_with_path, **{'header': None, 'sheet_name': sheet_name, 'nrows': 30,
                                                 'keep_default_na': True, 'dtype': str})

    excel_col_num = column_index_from_string(coordinate_from_string(header_col + '1')[0])  # -1 to 0 index

    for row in df_temp.itertuples():
        if type(row[excel_col_num]) == str:  # cell in input being checked is ok, otherwise blank cells/None cause
            # problems in compare
            if row[excel_col_num].strip().lower() == header_string.strip().lower():
                header_row = row.Index
                break
    if header_row is None:
        exit_yes((f"File may be bad.\n\nThe header check string '{header_string}' "
                  f"was not found in column '{header_col}' "
                  "in the first 30 lines of input file:"
                  f"\n\n'{file_with_path}'"
                  ))
    return header_row


def single_pivot_report(df, index_fields, value_fields, sheet_name, single_piv_writer, second_pivot_by_count=False):
    """ run one pivot report and write out to worksheet leaving room for titles to be added later """
    logger.debug(f"here '{sheet_name=}'")

    # ExcelWorkbook = py.load_workbook(FilePath)
    # writer = pd.ExcelWriter(FilePath, engine='openpyxl')
    # writer.book = ExcelWorkbook

    sheet_name2 = 'Cnt,' + sheet_name[:27]

    if sheet_name in single_piv_writer.book.sheetnames or sheet_name2 in single_piv_writer.book.sheetnames:
        file_counter_max = 0
        for sheet_ in single_piv_writer.book.sheetnames:
            if re.compile(r'-(\d+)$').search(sheet_):  # to avoid None if not found
                file_counter = re.compile(r'-(\d+)$').search(sheet_).group(1)
                if int(file_counter) > file_counter_max:
                    file_counter_max = file_counter
        file_counter_next = str(int(file_counter_max) + 1)

        sheet_name = sheet_name[:-(len(file_counter_next)+1)] + '-' + file_counter_next
        sheet_name2 = sheet_name2[:-(len(file_counter_next)+1)] + '-' + file_counter_next

    if isinstance(index_fields, str):
        index_fields = [index_fields]  # to generalize to list and not force user to enter []s
    fields_not_in_file = set(index_fields) - set(df.columns)  # some fields not in df
    if fields_not_in_file:
        pymsgbox.alert("Specified index field on pivot not in dataframe:\n" + ','.join(fields_not_in_file) +
                       "\n\nAvailable fields are:\n" + ', '.join(set(df.columns)),
                       'Warning:Tabulate field not on file, continuing')
    else:
        df_pt = pd.pivot_table(df,
                               index=index_fields,
                               values=value_fields,
                               aggfunc='count',
                               margins=True)
        # take the df created by pivot and add percent field by dividing count by total number of addresses
        df_pt['Pct_of_Total'] = round(df_pt[value_fields] / df.shape[0] * 100, 1)
        df_pt.to_excel(single_piv_writer, sheet_name=sheet_name, startrow=5)

        if second_pivot_by_count:
            # same as above but sorted by count
            df_pt = df_pt.sort_values(value_fields, ascending=False)
            df_pt.to_excel(single_piv_writer, sheet_name=sheet_name2, startrow=5)


def pivot_reports(df, output_wks, input_fn, dict_address_concentration):
    """
    :param df: input dataframe
    :param output_wks: output spreadsheet
    :param input_fn: name of file report is being run on, goes in title3 only
    :param dict_address_concentration report concentrated prop description and removal reason
    """
    logger.info("creating reports")

    # ExcelWorkbook = py.load_workbook(FilePath)
    # writer = pd.ExcelWriter(FilePath, engine='openpyxl')
    # writer.book = ExcelWorkbook
    # if not 'testSheet' in book.sheetnames:
    #     book.create_sheet('testSheet')

    writer = pd.ExcelWriter(output_wks, engine='openpyxl')
    df_clean = df[df['remove'] == '']

    # Create summary sheet of Rawdata, Formatted and Removed
    # for other states, roll all counties in to one called "All Counties'
    df['countysummed'] = np.where(df['state'] == ROV_SETUP['expectedstate'], df['statecounty'], "All Counties")

    # State by county for all including removed
    single_pivot_report(df, index_fields=['state', 'countysummed'], value_fields=['address'],
                        sheet_name='RawData by State-County', single_piv_writer=writer, second_pivot_by_count=False)

    if ROV_SETUP['add_filename_column_flag']:
        single_pivot_report(df, index_fields=['filename'], value_fields=['address'], sheet_name='RawData by Filename',
                            single_piv_writer=writer, second_pivot_by_count=False)

    single_pivot_report(df, index_fields=['remove'], value_fields=['address'], sheet_name='Removed Reasons',
                        single_piv_writer=writer, second_pivot_by_count=False)

    if ROV_SETUP['splitfield'] != "":
        single_pivot_report(df_clean, index_fields=ROV_SETUP['splitfield'], value_fields=['address'],
                            sheet_name='Clean by ' + ROV_SETUP['splitfield'][:22],
                            single_piv_writer=writer, second_pivot_by_count=True)

    # address concentration
    df_pt = pd.pivot_table(df, index=['state', 'county', 'city', 'address', 'remove'],
                           values=['lastname'],
                           aggfunc='count')
    df_from_query = df_pt.query("lastname >= " + str(PROP_CONCENTRATION))
    if len(df_from_query) > 0:
        df_from_query.rename(columns={'lastname': 'address_count'}, inplace=True)
        df_from_query.reset_index(inplace=True)

        # open browser windows for concentrated properties
        address_concentration_open_browser(df_from_query)

        # tried to eliminate error, "A value is trying to be set on a copy of a slice from a DataFrame", but couldn't so
        # isolated to with .copy() to show all is well
        dfq = df_from_query.copy()
        dfq['addrdesc'] = dfq.apply(lambda lam_row: get_addr_concentration(dict_address_concentration,
                                                                           lam_row.state, lam_row.county, lam_row.city,
                                                                           lam_row.address)[0], axis=1)

        dfq.to_excel(writer, sheet_name='Address GT ' + str(PROP_CONCENTRATION), startrow=5, index=False)

    # fields specific to county check - zip, county, zip/county match
    if ROV_SETUP['run_county_check_code_flag']:
        logger.debug('running countyCheck pivots')
        # recode opposite of mismatch field value so we can sum
        df['match_county'] = np.where(df["mismatch_county"] == 1, 0, 1)

        # % county mismatched
        df_pt = pd.pivot_table(df, index=['state', 'countysummed'],
                               values=['address', 'mismatch_county', 'match_county'],
                               aggfunc={'address': 'count', 'mismatch_county': 'sum', 'match_county': 'sum'},
                               margins=True)
        df_pt['Pct_Mis_Matched'] = round(df_pt['mismatch_county'] / df_pt['address'] * 100, 1)
        df_pt['Pct_Matched'] = round(df_pt['match_county'] / df_pt['address'] * 100, 1)
        df_pt.to_excel(writer, sheet_name='County Match Summary', startrow=5)

        # zips w/ # occurrences in ZIP_CONCENTRATION, mismatch first
        df_pt = pd.pivot_table(df, index=['state', 'statecounty', 'zip_county_list', 'zip', 'mismatch_county'],
                               values=['address'],
                               aggfunc={'address': 'count'})
        df_pt['Pct_of_Total'] = round(df_pt['address'] / df.shape[0] * 100, 1)
        df_from_query = df_pt.query("address >= " + str(ZIP_CONCENTRATION))
        df_from_query = df_from_query.sort_values(["mismatch_county", "address"], ascending=(False, False))
        df_from_query.rename(columns={'address': 'address_count'}, inplace=True)
        df_from_query.reset_index().to_excel(writer, sheet_name='Zips Over ' + str(ZIP_CONCENTRATION), startrow=5,
                                             index=False)

    # pivots for ad hoc fields
    for specs in ROV_SETUP['pivot_specs']:
        if specs['pivot_fields']:
            dfx = (df if specs['pivot_for_all'] else df_clean)
            universe = ('All-' if specs['pivot_for_all'] else 'Cln-')
            single_pivot_report(dfx, index_fields=specs['pivot_fields'], value_fields=['address'],
                                sheet_name=universe + ','.join(specs['pivot_fields'])[:27],
                                single_piv_writer=writer,
                                second_pivot_by_count=(True if specs['pivot_by_cnt'] else False))

    piv_wb = writer.book
    for ws in piv_wb.worksheets:
        autosize_xls_cols(ws)  # widen columns using BEK routine before wide titles

        ws["A1"] = "Address Summary Report"
        ws['A1'].font = Font(b=True, size=16)
        ws["A2"] = ws.title
        ws['A2'].font = Font(b=True, size=12)
        ws["A3"] = "Input File: " + input_fn
        ws['A3'].font = Font(b=True, size=12)
        ws["A4"] = datetime.now().strftime('%m/%d/%Y')

    logger.debug('out of countyCheck pivots')

    writer.close()


def create_import_code_from_sheet(sheet_with_python_code, output_file):
    """ Imports a spreadsheet sheet and creates properly indented python code from it based on sheet columns

    Parameters
    ----------
    sheet_with_python_code : the sheet containing the python code
    output_file : the .py that will receive the text python code
    """
    logger.debug("here")

    # sheet_with_python_code = setup_wb[sheet_with_python_code]
    with open(ROV_SETUP['exe_path'] / output_file, "w") as new_code_text:

        # Loop through sheet positions in excel file and indent as needed in python code (4 characters per col hardcoded)
        for rowidx, row_cells in enumerate(sheet_with_python_code.iter_rows()):
            # col and rows indexes needs +1 cause python 0 indexed, worksheet cells start at 1
            for colidx, cell in enumerate(row_cells):
                if sheet_with_python_code.cell(rowidx + 1, colidx + 1).value is not None and \
                        not sheet_with_python_code.cell(rowidx + 1, colidx + 1).value.strip().startswith('#'):
                    new_code_text.write(" " * colidx * 4 + sheet_with_python_code.cell(rowidx + 1, colidx + 1).value + '\n')
                    break  # dont go to subsequent columns once data is found to avoid extraneous info
    new_code_text.close()
    return new_code_text


def address_concentration_open_browser(df):
    """ uses address fields from query on address concentration pivot and google search with params to open
    browser windows for each address to decide if addreess should be excluded from carding.
    """
    logger.debug("here")

    openbrowser = True
    if ROV_SETUP['concentrated_address_browser_prompt_freq'] in [1, 2]:
        if ROV_SETUP['concentrated_address_browser_prompt_freq'] == 2:
            choice = pymsgbox.confirm("Open " + str(len(df)) + " tabs in browser?.  OK?\n\n",
                                      'Open browser windows?', ['Yes', 'No'])
            if choice == "No":
                openbrowser = False
        if openbrowser:
            df = df.reset_index()  # converts multi-index to columns
            for index, row in df.iterrows():
                st = row['state']
                cnt = row['county']
                cit = row['city']
                ad = row['address']
                webbrowser.open('https://www.google.com/search?q=' + st + "+" + cnt + "+" + "+" + cit + "+" + ad,
                                new=2)
        else:
            pass
    else:
        pass


def fields_to_list(base_list, new_fields):
    """ adds new fields to base lists.  exits if already in list."""
    new_fields_lst = [field.strip().lower() for field in new_fields.split(",")]
    same_fields = [field.strip().lower() for field in new_fields_lst if field in base_list]

    if not same_fields:
        base_list.extend(new_fields_lst)
    else:
        exit_yes(f"Field(s) '{str(same_fields)}' already exists on list.  Can not add it.")


def split_files_for_sincere(lim):
    """ splits main df into files by splitfield for loading into VoterLetters/Sincere.  splits large files into subs
    with counter if larger than limit.
    """
    logger.info("split files for Sincere")

    def chunk_split_file(df, limit, split_path_hold, split_filename):
        """ pass a split file and needed parts and it will chunk it into sizes specified in setup as
        ROV_SETUP['sub_split_limit'] and postfix name with file-counter.
        split_filename is the root which csv and 'file x' is postfixed to
        """
        logger.debug("here")

        addresses_to_write = len(df)  # don't use df reference to save time

        if addresses_to_write <= limit:  # write one file
            split_file = split_path_hold / (split_filename + ".csv")
            df.to_csv(split_file, index=False, columns=ROV_SETUP['splitfile_field_list'])
            logger.info(f"'{split_filename}' written, {len(df)} addresses.")
        else:  # need to create split files by looping
            for file_counter in range(1, math.ceil(addresses_to_write / limit) + 1):
                low_record = ((file_counter - 1) * limit)
                hi_record = (file_counter * limit) - 1

                split_file = split_path_hold / (split_filename + " file-" + str(file_counter) + '.csv')
                df_chunk = df[low_record: hi_record + 1]
                df_chunk.to_csv(split_file, index=False, columns=ROV_SETUP['splitfile_field_list'])
                # print("   split sub file ", file_counter)
                logger.debug(f" -'{split_filename + ' file-' + str(file_counter)}' written, {len(df_chunk)} "
                      f"addresses.")

    if lim == 0:
        lim = 99999999
    ip_stem = ROV_SETUP['OPFile'].stem
    op_stem = ROV_SETUP['splitfnbase'].stem

    # Ask if sorting by zip ok.  Assumes current sort order is how data was sorted under Combine.
    if ROV_SETUP['sortchoice'] in [1, 2]:
        exit_yes_no("Output is being sorted by zip. OK?",
                    'SORT BY ZIP?',
                    display_exiting=False)

    combinedfile_w_path = ROV_SETUP['combined_path'] / (ip_stem + '.csv')

    try:
        df_combo_w_no_remove = pd.read_csv(combinedfile_w_path, header=0, keep_default_na=False)
    except:
        exit_yes("Unable to read  the Combinedfile in the split step.  Was 'Combine' run?"
                 f"\n\nMissing file:\n\n{combinedfile_w_path}"
                 )

    df_combo_w_no_remove = df_combo_w_no_remove[df_combo_w_no_remove["remove"] == ""]
    exit_yes_no(f"Split will process {len(df_combo_w_no_remove)} clean addresses.",
                'SPLIT RECORDS',
                display_exiting=False)

    if ROV_SETUP['sort_list']:
        df_combo_w_no_remove.sort_values(by=ROV_SETUP['sort_list'], inplace=True)

    fields_missing_from_combinefile = set(ROV_SETUP['splitfile_field_list']) - set(df_combo_w_no_remove.columns)
    if fields_missing_from_combinefile:
        exit_yes("Combinefile is missing the following fields to write to Splitfiles:"
                 f"\n\n{', '.join(fields_missing_from_combinefile)}"
                 )

    if ROV_SETUP['splitfield'] == '':  # no split field specified so write out one file with name "Combined"
        split_filename = ROV_SETUP['expectedstate'] + '-' + "Combined " + op_stem
        chunk_split_file(df_combo_w_no_remove, lim, ROV_SETUP['split_path_hold'], split_filename)

    else:
        if ROV_SETUP['splitfield'].lower() == 'county':
            splitfield = 'statecounty'
        else:
            splitfield = ROV_SETUP['splitfield']
        unique_split_values = df_combo_w_no_remove[splitfield].unique()
        unique_split_values.sort()

        # for each - write out a csv file.
        for splitfield_value in unique_split_values:
            # print("split " + splitfield_value)
            df_one_splifield = df_combo_w_no_remove[df_combo_w_no_remove[splitfield] == splitfield_value]

            if ROV_SETUP['splitfield'].lower() == 'county':
                split_filename = ROV_SETUP['dict_statecounty_to_alt_formats'][splitfield_value][2]
                # get the format of county we want to use for filename using county lookup
            else:
                split_filename = ROV_SETUP['expectedstate'] + '-' + splitfield_value
            split_filename = split_filename + "- " + op_stem

            # write out file, broken into chinks if needed
            chunk_split_file(df_one_splifield, lim, ROV_SETUP['split_path_hold'], split_filename)


def get_addr_concentration(dict_addr_rem, state, county, city, address):
    # get_addr_remove_OLD has passed dictionary which is causing problems
    """
    Given location for a concentrated address, return the remove reason (blank if ok) or description based on dictionary
    Returns a tuple [0] is description, [1] is reason for removal
    :param dict_addr_rem: given a tuple of state, county, city, address, returns a two element list of remove code or description
    :param state, county, city, address
    """
    logger.info("get address concentration")

    state = state.lower().strip()
    county = county.lower().strip()
    city = city.lower().strip()
    address = address.lower().strip()
    addr_desc_rem_tuple = dict_addr_rem.get(tuple([state, county, city, address]), ['missing', 'missing'])

    return addr_desc_rem_tuple


def check_county_to_zips(df, zipskip_list, dict_statecounty):
    """ check county by comparing it to zip lookup.  sets flag for mismatches and saves vars along the way. """
    logger.info("check county to zip file")

    df['orig_county'] = df['county']
    df['clean_county'] = df['county'].apply(clean_field)

    df['statecounty'] = df[['state', 'clean_county']] \
        .apply(lambda row: (row['state'] + "-" + row['clean_county']).upper(),
               axis="columns")

    # collapse all statecounty into "all_counties" for bad states
    df.loc[(df['state'] != ROV_SETUP['expectedstate']), 'statecounty'] = "All_Counties"
    # df['clean_county'] = df['county'].apply(clean_field)  #  this is a repeat of line above

    df['numzip'] = df['zip'].map(lambda x: (int(x) if is_number(x) else 0))

    # TODO: below mixed state counties with counties and makes it impossible to produce simple summaries of data
    # TODO: add recode of all non-expected state to 'all counties in non-expected state'
    # replace county with standard version so it matches filename, etc (value [0] is clean, mixed-case)
    df['county'] = df['statecounty'].map(lambda x: dict_statecounty.get(x, [x+'-statecounty not found'])[0])

    # set 'zip_county_list' to string of statecounties based on zip
    df['zip_county_list'] = df['numzip'].map(lambda x: ","
                                             .join(ROV_SETUP['dict_zip_to_countylist'].get(x, ['zip not found'])))

    # check for county in list to account for split zips
    # df['mismatch_county'] = df.apply(lambda dfx: (1 if dfx['clean_county'] not in dfx['zip_county_list']
    #                                               else 0), axis=1)
    df['mismatch_county'] = df.apply(lambda row: (1 if row['statecounty'] not in row['zip_county_list']
                                                  else 0), axis=1)

    # if flag is set, use zip skip list to reset county/zip mismatches to 0
    if ROV_SETUP['skip_selected_zip_match_flag']:
        df['mismatch_county'] = df.apply(lambda row:
                                         (0 if (row['county'].lower(), row['zip'])
                                               in zipskip_list else row['mismatch_county']), axis='columns')

    logger.debug("Done filling zip and county info")


def read_file_to_df(file_with_path, **param_dict):
    """reads either xlsx or csv into a dataframe using parms passed in dictionary. Non-applicable parms
    are skipped."""
    logger.info(f"reading file to dataframe '{file_with_path.stem}'")

    if PurePath(file_with_path).suffix.lower() == '.xlsx':
        filtered_dict = {k: v for k, v in param_dict.items()
                         if k in [p.name for p in inspect.signature(pd.read_excel).parameters.values()]}
        df_temp = pd.read_excel(file_with_path, **filtered_dict)
    elif file_with_path.suffix.lower() == '.csv':
        filtered_dict = {k: v for k, v in param_dict.items()
                         if k in [p.name for p in inspect.signature(pd.read_csv).parameters.values()]}
        df_temp = pd.read_csv(file_with_path, **filtered_dict)
    else:
        df_temp = None
        logger.debug('here')
        exit_yes((f"Bad file type on input - not xlsx or csv."
                  f"\n\nFile: '{file_with_path}'"
                  ))
    return df_temp


def get_setup_file_name(initial_campaign_dir):
    """ use Tkinter to get name of setup workbook for desired campaign.  checks version of ROVCleaver program to
    version of setup xlsx using file name
    """
    logger.debug('picking setup file')
    Tk().withdraw()  # we don't want a full GUI, so keep the root window from appearing

    # Use this flag when testing - False allows hardcoding input from alternate starting directory
    # noinspection PyUnreachableCode
    if True:
        # show an "Open" dialog box and return the path to the selected file
        # V13.1 parameterize start directory and remove '/Users/Denise' reference
        # ROV.setup_file_name = askopenfilename(
        #     initialdir=INITIAL_CAMPAIGN_DIR,
        #     title="Select ROVCleaver setup file SetupFormat" + filename_ver, filetypes=(
        #         ("Excel files", "*.xlsx *.xls"),))

        xl_setup_file_name = get_file_name("Select Setup File",
                                           f"Select ROVCleaver setup file xlsx. Must have 'UniversalSetup' in "
                                           f"name",
                                            initial_campaign_dir)
        setup_file_name = pathlib.Path(xl_setup_file_name).expanduser()
    else:
        # Hardcode in TEST INPUT FILE directory for repetitive testing
        setup_file_name = pathlib.Path(
            "~/Dropbox/Postcard Files/TestInputFiles/TestCampaigns/BEK Test UniversalSetup/ROVCleaver "
            "UniversalSetup.xlsx").expanduser()

        exit_yes_no("Running hardcoded Setup file.  OK?\n\n" + str(setup_file_name),
                    'RUN IN TEST?',
                    display_exiting=False)
        # choice = bek_text_box(f"Running hardcoded Setup file.  OK?\n\n{str(setup_file_name)}",
        #                       "Choose an Action",
        #                       '',
        #                       )

    if not 'universalsetup' in str(setup_file_name).lower():
        exit_yes(("The chosen setup file does not contain 'UniversalSetup' in it's name.\n"
                  "\n\nYou need to pick a different setup file or "
                  ))

    return setup_file_name


# def assign_rov_variables():
#     """ assigns variables from cells in setup sheet and places them in global object.  set some global variables"""
#
#     ### Done inputing fields - make sure specified input paths exist or EXIT
#     bad_path_exit(ROV_SETUP['rawdata_path'])
#     if ROV_SETUP['run_merge_data_flag']:
#         bad_path_exit(ROV_SETUP['rawdata_path'])


def exit_for_unwanted_setup_options():
    """ verifies setup options when program is run allowing to exit and edit setup
    """
    logger.info('Checking setup options')

    if ROV_SETUP['run_county_check_code_flag'] and ROV_SETUP['splitfield'].lower() != 'county':
        exit_yes_no(("You are checking for zip/county mismatches"
                     "\nbut you are not splitting by county."
                     "\n\n\nIs this what you meant?"))

    if ROV_SETUP['skip_selected_zip_match_flag']:
        str_zipskip = str()
        for county, zip in ROV_SETUP['zipskip_list']:
            str_zipskip += f"{county:<20} {zip}\n"

        exit_yes_no("County and zip combinations to ignore when flagging mismatches "
                    f"(listed in 'ZipSkip' setup sheet): \n{str_zipskip}",
                    "Zip Codes to Ignore in County Matching")

    if ROV_SETUP['sortchoice'] in [1, 2]:
        exit_yes_no("Sorting by other than county / random number can cause ugly clumps of addresses "
                    "(eg all PO boxes).  OK?",
                    "Sort Order")

    if ROV_SETUP['copy_other_format_files_flag']:
        exit_yes_no("Copy FORMAT files listed in setup sheet 'FormatCopies' during combine?\n\n",
                    'COPY FORMAT FILES?')

        formatfile_copy(ROV_SETUP['copy_formatfile_filelist_sheet'], perform_copies=False)

    if ROV_SETUP['run_first_code_flag']:
        # display code and prompt if it should be run
        display_imported_code(ROV_SETUP['first_code_sheet'], ROV_SETUP['first_code'])

    if ROV_SETUP['run_middle_code_flag']:
        display_imported_code(ROV_SETUP['middle_code_sheet'], ROV_SETUP['middle_code'])

    if ROV_SETUP['run_last_code_flag']:
        display_imported_code(ROV_SETUP['last_code_sheet'], ROV_SETUP['last_code'])


def create_field_lists():
    """ fill array with default field names of ' ' and add fields required by options selected"""
    logger.info('Creating field lists')

    # create 'formatfile' field list by replacing blanks with inputfile names and replacing those to be renamed,
    # keep 'x' so we can zip with input fields
    ROV_SETUP['formatfile_field_list_to_zip'] = [input_field if rename_field == '' else rename_field
                                        for (input_field, rename_field)
                                        in itertools.zip_longest(ROV_SETUP['inputfile_orig_list'],
                                                                 ROV_SETUP['inputfile_renamed_list'])]

    ROV_SETUP['inputfile_delete_field_list'] = ['_' + ofield
                                       for ofield, new_field
                                       in zip(ROV_SETUP['inputfile_orig_list'], ROV_SETUP['formatfile_field_list_to_zip'])
                                       if new_field.strip().lower() == 'x']

    # create dict to rename fields that need it, otherwise leave with prefix
    ROV_SETUP['inputfile_rename_fields_dict'] = {('_' + orig_field): new_field
                                        for orig_field, new_field
                                        in zip(ROV_SETUP['inputfile_orig_list'], ROV_SETUP['formatfile_field_list_to_zip'])
                                        if new_field.strip().lower() != 'x'}

    # list of actual fields to be included on file ('x's removed)
    ROV_SETUP['formatfile_field_list'] = [field for field in ROV_SETUP['formatfile_field_list_to_zip'] if field != 'x']

    # check for duplicate field values
    duplicate_fields = [field for field, count in collections.Counter(ROV_SETUP['formatfile_field_list']).items() if count > 1]
    # duplicate_fields = [field for field, count in collections.Counter(ROV_SETUP['formatfile_field_list']).items() if count > 1]
    if duplicate_fields:
        exit_yes((f"The following fields are duplicated on the Format file.  Remove one occurance."
                  f"\n\n{', '.join(duplicate_fields)}"
                  ))

    ROV_SETUP['combinefile_field_list'] = [field
                                  for field in ROV_SETUP['formatfile_field_list']
                                  if field not in ROV_SETUP['combinefile_fields_to_delete_list']]  # dont use set to keep order

    # ROV_SETUP['splitfile_field_list'] = [field.strip().lower() for field in ROV_SETUP['splitfile_field_list']]

    if ROV_SETUP['run_county_check_code_flag']:  # V15.0
        fields_to_list(ROV_SETUP['formatfile_field_list'], "orig_county,clean_county,zip_county_list,statecounty,"
                                                  "mismatch_county,"
                                                  "numzip")

    # add 'filename' field if requested
    if ROV_SETUP['add_filename_column_flag']:  # V15.0
        fields_to_list(ROV_SETUP['formatfile_field_list'], "filename")

    # Always add 'remove' field - makes reporting easier and will almost always be used
    fields_to_list(ROV_SETUP['formatfile_field_list'], 'remove')
    fields_to_list(ROV_SETUP['formatfile_field_list'], 'rownum')  # V15.0
    fields_to_list(ROV_SETUP['formatfile_field_list'], 'dupe_key')  # V16.0
    fields_to_list(ROV_SETUP['formatfile_field_list'], 'dupe_id_field')  # V16.0
    fields_to_list(ROV_SETUP['formatfile_field_list'], 'pull_group')
    fields_to_list(ROV_SETUP['formatfile_field_list'], 'custom_field')

    if ROV_SETUP['add_random_number_column_flag']:
        # add field randnum so we can sort FORMAT files by county and randnum
        fields_to_list(ROV_SETUP['formatfile_field_list'], "randnum")  # V15.0

    # Check that no fields are specified on output that are not on input  # V15.0 commented out -
    fields_missing_from_input = set(ROV_SETUP['combinefile_field_list']) - \
                                set(ROV_SETUP['formatfile_field_list'])
    # this doesn't need to check input to format which is what it was
    if fields_missing_from_input:
        exit_yes((f"Field(s) are specified to output on Combine file but are not present on Format.  "
                  f"\n\nMissing field(s) specified are:\n\n"
                  f"'{', '.join(fields_missing_from_input)}'"
                  f"\n\nFields on input are:\n\n"
                  f"{', '.join(ROV_SETUP['formatfile_field_list'])}"
                  ))

    # Check if splitfield is on field list, error if not
    if ROV_SETUP['splitfield'] != '' and ROV_SETUP['splitfield'].lower() not in ROV_SETUP['formatfile_field_list']:
        exit_yes((f"Splitfield field '{ROV_SETUP['splitfield']}' is missing from Format file field list.\n\n"
                  f"Available fields are:\n{', '.join(ROV_SETUP['formatfile_field_list'])}"
                  ))


def create_dicts():
    """ create dicts, lists, need to run """
    logger.info('Creating dictionaries')

    # create county dict returning various formats with
    # GA-CHATHAM as key: [0] is filename format; [1] print; [2] state-mixed
    ROV_SETUP['dict_statecounty_to_alt_formats'] = zip_file_to_county_dict(
        ROV_SETUP['exe_path'] / 'zip-codes-database-DELUXE-BUSINESS.csv',
        ROV_SETUP['exe_path'] / 'Unique_County_List.xlsx')

    logger.debug('Ran Counties_to_xls')

    if ROV_SETUP['run_county_check_code_flag']:
        with open(ROV_SETUP['exe_path'] / ZIP_TO_COUNTY_LIST_FILE, "r") as dict_file:
            ROV_SETUP['dict_zip_to_countylist'] = ast.literal_eval(dict_file.read())
            logger.debug("Imported " + ZIP_TO_COUNTY_LIST_FILE)

    concentrated_addresses_data = range_to_list(ROV_SETUP['concentrated_addresses_sheet'], 2,
                                                len(ROV_SETUP['concentrated_addresses_sheet']['A']), 1, 7)

    address_desc_list = [
        [tuple([state.strip().lower(), county.strip().lower(), city.strip().lower(), address.strip().lower()]),
         [addressdesc.strip(), removeReason.strip()]]
        for state, county, city, address, removeReason, freq, addressdesc
        in concentrated_addresses_data]  # NOTE: [0] desc, [1] remove

    ROV_SETUP['dict_concentrated_addresses'] = dict(address_desc_list)
    # this dict can be used in remove code sheet; returns a two-tuple [0] is address description, [1] remove reason
    # code like this: if ROV_SETUP['dict_concentrated_addresses'].get(('NC','new hanover','wilmington', '811 martin st'),
    # 'Other') != 'Other' => set code
    # or df['remove'] = ROV_SETUP['dict_concentrated_addresses'].get(('NC','new hanover','wilmington', '811 martin st'), 'Other')[1]

    # Create list of county/zips to not flag as mismatched for a particular county.  Key is tuple of county and zip.
    if ROV_SETUP['skip_selected_zip_match_flag']:
        zip_skip_range = range_to_list(ROV_SETUP['skip_selected_zip_sheet'], 2, len(ROV_SETUP['skip_selected_zip_sheet']['A']), 1, 2)
        ROV_SETUP['zipskip_list'] = [(cnty.lower(), int(zipcode)) for cnty, zipcode in zip_skip_range]


def display_imported_code(sheet_name, py_file_name):
    """ reads python code contained in a workbook sheet, writes it to a py file, displays the contents on screen,
    and writes it to console for fixing program mistakes

    Parameters
    ----------
    sheet_name : the sheet in the setup.xlxs where code is found
    py_file_name : the .py text python file where the code is saved
    """
    logger.debug('here')

    # put the text code in py_file_name
    # create_import_code_from_sheet(sheet_name, os.path.abspath(py_file_name))
    create_import_code_from_sheet(sheet_name, py_file_name)

    with open(ROV_SETUP['exe_path'] / py_file_name, "r") as myfile:  # this copies in the code but does not execute it
        logger.debug(f"compiling {py_file_name} - taken from sheet: {sheet_name}")
        try:
            codeobj = compile(myfile.read(), py_file_name, 'exec')
        except Exception as e:
            msg = (f"There was an error compiling '{py_file_name}', line number {e.lineno}, from sheet: "
                        f"'{sheet_name.title}'.  "
                        f"The text of the line is: '{e.text.strip()}'.  "
                        f"Fix and rerun.")
            logger.info(msg)
            bek_text_box(msg,'Import Code Error','')
            raise Exception

        # create a compiled object to list the lines to be executed for debugging with ine nums and comments/blanks
        # removed

        # code_lines = [(str(index + 1) + ' ' + line)
        #               for index, line in enumerate(inspect.getsourcelines(codeobj)[0])]
        # f"{index + 1:<4   this left justifies index+1
        code_lines = [f"{index + 1:<4}  {line}"
                      for index, line in enumerate(inspect.getsourcelines(codeobj)[0])]
        if len(code_lines) > 40:  # too many to display on screen, disables the 'ok' box
            skipped = len(code_lines) - 40
            code_lines = code_lines[:40]
            code_lines.append(f"<{skipped} lines not shown>")
        code_lines = ' '.join(code_lines)
        logger.info(f"\n{py_file_name} to be run - taken from sheet: {sheet_name}")
        [print(f"{index + 1:<4}  {line}", end=' ') for index, line in enumerate(inspect.getsourcelines(codeobj)[0])]

        # Below names module based on py code name, eg 'first_code_to_be_run_module'; puts pointer to module in
        # ROV_SETUP
        # A function from this module (can even have passed parameters) will be called later in the program.
        ROV_SETUP[pathlib.Path(py_file_name).stem + '_module'] = \
            importlib.import_module(pathlib.Path(py_file_name).stem)

    exit_yes_no(f"'{py_file_name}' to be run - taken from setup sheet: '{sheet_name}'"
                f"\n\nCode and line numbers printed in python console log, too.\n\n"
                f"{code_lines}",
                f"Check '{py_file_name}' Code",
                display_exiting=False)



def check_file_headers(ws, vals):
    """
    Check list of (cell, val) tuples representing header labels in ws_to_chk and error if val not found in cell.
    eg vals = [('A1', 'use'), ('B1', 'fromFilePath'), ('C1', 'fromfilename'), ....]
    """

    def chk_header_vals(ws_to_chk, cell, val):
        """ error if val not found in wks cell. """
        if str(ws_to_chk[cell].value).strip().lower() != str(val).lower():
            exit_yes((f"Column heading '{cell}' on Setup sheet '{ws_to_chk.title}' not equal to literal '{val}'."
                      f"\n\nIt is '{str(ws_to_chk[cell].value)}'."),
                     )

    for pairs in vals:
        chk_header_vals(ws, pairs[0], pairs[1])


def process_format_files(filelist_wks):
    """ loops through filelist sheet and creates format for specified x """
    logger.info('process format files')

    cumulative_missing_counties_list = []

    for fn, format_flag, combine_flag, update_fn, update_fields, pull_group, custom_field, notes, *_ \
            in filelist_wks.iter_rows(min_row=2, values_only=True):

        if str(format_flag).strip().lower() == "x":
            ip = process_format_file(fn, pull_group, custom_field, ROV_SETUP['rawdata_path'],
                                     ROV_SETUP['format_path'],
                                     cumulative_missing_counties_list)  # file appended to in function

            if str(update_fn).strip().lower() != 'none':
                merge_into_format_file(ip, str(update_fn), ROV_SETUP['format_path'])

    missing_counties_file = ROV_SETUP['root_path'] / 'missing_counties.csv'
    if os.path.isfile(missing_counties_file):
        os.remove(missing_counties_file)

    if cumulative_missing_counties_list:
        pymsgbox.alert("The following counties are not in the lookup file, written to 'missing_counties.csv':\n\n" +
                       ",".join(cumulative_missing_counties_list),
                       "Alert")
        with open(ROV_SETUP['exe_path'] / missing_counties_file, mode='wt', encoding='utf-8') as myfile:
            myfile.write('\n'.join(cumulative_missing_counties_list))


def process_format_file(fn, pull_group, custom_field, input_path, op_path,
                        missing_counties_list_all_formatfiles):
    """
    Given a rawdata csv or xlsx with filename fn and path input_path, places a transformed csv file with the same
    name prefixed with 'FORMATTED' to the directory output_path, and returns a dataframe of the info.  The list
    missing_counties_list_all_formatfiles accumulates all counties not found in
    dictionary ROV_SETUP['dict_statecounty_to_alt_formats'] so they can be recoded in first_code.

    Parameters
    ----------
    pull_group :
    custom_field :
    """
    logger.info(f"Creating formatted file for: '{fn}'" )

    start_time = datetime.now()

    ip_file_w_path = input_path / fn

    # Make sure data file exists
    # bad_file_exit(ip_file_w_path, "Data file does not exist. Change filelist or "
    #                               "copy into Rawdata directory.\n\n" + str(ip_file_w_path))
    bad_file_exit(input_path / fn, "Data file does not exist. Change filelist or "
                                  "copy into Rawdata directory.\n\n" + str(input_path / fn))
    # find the header row in the input file using an expected string and column.  error if > 30.  Defaults to search sheet0.
    header_row = find_header_row_in_file(ip_file_w_path, ROV_SETUP['check_header_string'], ROV_SETUP['strcheck_header_col'])

    ip = read_file_to_df(ip_file_w_path, **{'header': header_row, 'sheet_name': 0, 'nrows': ROV_SETUP['rows_to_read_limit'],
                                            'keep_default_na': False, 'dtype': str})

    # convert column names to lower case
    ip.columns = [field.strip().lower() for field in ip.columns.values]

    # do all fields on file being read match expected?
    bad_header_fields = set(ip.columns) ^ set([field for field in ROV_SETUP['inputfile_orig_list'] if field != ''])
    if bad_header_fields:
        exit_yes((f"The fields in header for '{fn}' do not match the ones in orig format files."
                  f"\nDifferences:\n\n{os.linesep.join(bad_header_fields)}"
                  # had to use os.linesep cause \n not allowed in fstrings
                  ))

    # convert columns identified as numeric in setup field list
    numeric_cols = [column_index for column_index, col_type in enumerate(ROV_SETUP['inputfile_type_list']) if
                    col_type == 'int']
    # non-blank are set to integer; code options here if move to other types
    for col in numeric_cols:
        ip[ip.columns[col]] = pd.to_numeric(ip[ip.columns[col]], errors='coerce', downcast='integer')
        # try to force specified numeric columns to integer
    logger.debug("Done reading in one file in process_format_file, csv or xlsx")

    ## rename fields with prefix of '_' so not conflicted with renamed field of same name
    logger.debug("Ready to prefix fields to with '_'")
    ip = ip.add_prefix('_')

    ip.rename(ROV_SETUP['inputfile_rename_fields_dict'], axis='columns', inplace=True)

    add_fields_list = [field for field in ROV_SETUP['formatfile_field_list'] if field not in ip.columns]
    # add new fields. check for duplicate field names on output is already done above
    ip = ip.reindex(columns=ip.columns.tolist() + add_fields_list, fill_value='')

    # fill rownum field
    ip['rownum'] = ip.index + 2  # +1 for 0index, +1 for header
    logger.debug("Done fill rownum")

    # Check for literal of TRUE or FALSE in fields because it gets converted to boolean and kills the program
    logger.debug("Ready to fill firstname true/false")
    replace_boolean_column_vals(ip, ['firstname', 'lastname', 'lastnametemp', 'address',
                                     'addresstemp', 'address2', 'city'])
    logger.debug("Done filling specified fields true/false with blank")

    # create filename filed here so we can use in recode below
    if ROV_SETUP['add_filename_column_flag']:
        ip['filename'] = fn

    # V16.1 add 'custom_field' to df to use in recoding sections if desired
    ip['pull_group'] = pull_group
    ip['custom_field'] = custom_field

    # create random number so we can sort
    if ROV_SETUP['add_random_number_column_flag']:
        np.random.seed(0)
        ip['randnum'] = np.random.random(ip.shape[0])

    if ROV_SETUP['run_first_code_flag']:
        # *** Only run imported first_code and middle_code in format, not combine to keep things like remove
        # assignment, random number from being overwritten.
        # last_code can be run in combine since it's only setting remove code.
        logger.info('Ready to run first_code')

        ROV_SETUP['first_code_to_import_module'].first_code_func(ip)

        # This is the function from the sheet with any parameters it needs
        logger.debug('Ran first_code')  # these prompts help if error in imported code

    if ROV_SETUP['run_county_check_code_flag']:
        check_county_to_zips(ip, ROV_SETUP['zipskip_list'], ROV_SETUP['dict_statecounty_to_alt_formats'])
        # FIXME pymsgbox list of mismatched counties not showing

    if ROV_SETUP['run_middle_code_flag']:
        # *** Only run imported first_code and middle_code in format, not combine to keep things like remove
        # assignment, random number from being overwritten.
        # last_code can be run in combine since it's only setting remove code.
        logger.debug('Ready to run middle_code')  # these prompts help if error in imported code

        ROV_SETUP['middle_code_to_import_module'].middle_code_func(ip)

        # This is the function from the sheet with any parameters it needs
        logger.debug('Ran middle_code')  # these prompts help if error in imported code

    # if requested, RUN code to remove unwanted records
    if ROV_SETUP['run_last_code_flag']:
        # *** Only run imported first_code and middle_code in format, not combine to keep things like remove
        # assignment, random number from being overwritten.
        # last_code can be run in combine since it's only setting remove code.
        logger.debug('Ready to run last_code (remove code)')  # these prompts help if error in imported code

        ROV_SETUP['last_code_to_import_module'].last_code_func(ip, ROV_SETUP['dict_concentrated_addresses'],
                                                            ROV_SETUP['expectedstate'])
        # This is the function from the sheet with any parameters it needs

        logger.debug('Ran last_code')  # these prompts help if error in imported code

    ip.drop(ROV_SETUP['inputfile_delete_field_list'], axis=1, inplace=True)

    # Sort file by randnum if flag is set
    if ROV_SETUP['sort_list']:  # true if not empty
        ip.sort_values(by=['remove', 'zip', 'address', 'randnum'], inplace=True)

    # Make sure county is on file, find mismatched counties and return for accumulating
    if ROV_SETUP['run_county_check_code_flag']:
        if 'county' not in ip.columns:
            exit_yes("'CHECK COUNTY INFO FLAG' option is specified as 'TRUE' but 'county' field is not present.")

        unique_statecounties = ip.loc[ip['remove'] == '', 'statecounty'].unique()
        # unique_statecounties = ip['statecounty'].unique()

        unique_statecounties.sort()
        missing_counties_this_formatfile = [chkfield
                                            for chkfield in unique_statecounties
                                            if chkfield.upper() not in ROV_SETUP['dict_statecounty_to_alt_formats']]

        # [print(chkfield) for chkfield in unique_statecounties]

        missing_counties_list_all_formatfiles.extend(missing_counties_this_formatfile)

    op_file = op_path / ("FORMATTED " + str(PurePath(fn).stem) + ".csv")
    # must check if remove flag is set otherwise field remove field is not on df

    ip.to_csv(op_file, index=False, columns=ROV_SETUP['formatfile_field_list'])

    if ROV_SETUP['run_last_code_flag']:
        # write out the 'removed' file
        op_file = ROV_SETUP['format_path'] / "Removed" / ("REMOVED " + str(PurePath(fn).stem) + ".csv")
        ip[ip["remove"] != ""].to_csv(op_file, index=False, columns=ROV_SETUP['formatfile_field_list'])

    logger.info("\nFormatted file: " + fn)
    logger.info("   Input records: " + str(len(ip)))
    if ROV_SETUP['run_last_code_flag']:
        logger.info("   Kept: " + str(len(ip[ip["remove"] == ""])))
        logger.info("   Removed: " + str(len(ip[ip["remove"] != ""])))

    logger.debug("\n", ip.head(5), "\n\n")

    logger.debug("End time is ", datetime.now().strftime("%H:%M:%S"), "  Elapsed time is", str(datetime.now() - start_time),
          " (H:M:S.s)")

    # Create summary sheet of Rawdata, Formatted and Removed
    pivot_file = ROV_SETUP['format_path'] / "Summary" / ("SUMMARY " + str(PurePath(fn).stem) + ".xlsx")
    pivot_reports(ip, pivot_file, fn, ROV_SETUP['dict_concentrated_addresses'])

    logger.debug("Leaving process_format_file")
    return ip


def formatfile_copy(ws_copy_formatfile_filelist, perform_copies=True):
    """
    Copies FORMATed files from a separate directory (likely came in a different format so needed their own read) into the current format directory
    :param ws_copy_formatfile_filelist: sheet containing rows of fromFile, toFile, processFlag
    If perform_copies = False, paths and file existence is checked but copies do not take place.
    :return: files copied to directory
    """
    logger.info('In formatfile_copy')

    def copy_file(source, destination):
        """ func to copy  with error handling pymsgbox"""
        try:
            shutil.copy(os.path.expanduser(source), destination)
            logger.info(f"File '{source}' copied successfully to\n'{destination}'.")
            # print("File " + source + " copied successfully to\n" + destination + ".")
        except Exception:
            exit_yes((f"File not copied:\n\n {source} \n\nto\n\n {destination}"
                      ))

    # Make sure column heading/locations are as expected and nothing was moved
    check_file_headers(ws_copy_formatfile_filelist,
                       [('A1', 'use'),
                        ('B1', 'fromFilePath'),
                        ('C1', 'fromfilename'),
                        ('D1', 'tofilepath'),
                        ('E1', 'tofilename'),
                        ])

    # make sure all files and directories exist for row in ws_filelist:
    for copy_formatfile_flag, from_path, from_fn, to_path, to_fn, *_ in islice(ws_copy_formatfile_filelist, 1, None):
        # islice starts at row index 1 not 0 to skip header; *_ discards unused cols (notes, etc)
        if str(copy_formatfile_flag.value).strip().lower() == "x":
            src_path, src_file_w_path, dest_path, dest_file_w_path = \
                assign_formatcopy_vars(from_path, from_fn, to_path, to_fn)

            bad_path_exit(src_path)
            bad_file_exit(file=src_file_w_path)
            bad_path_exit(dest_path)
            # bad_file_exit(file=dest_file_w_path)  # cant check cause doesn't exist

            # if we got to here, paths and files are ok so copy
            if perform_copies:
                copy_file(src_file_w_path, dest_path)


def assign_formatcopy_vars(from_path, from_fn, to_path, to_fn):
    """
    Takes to and from paths and filenames and converts to simple files w path. Paths can be code to be evaluated,
    eg referencing variable rootPathOneUp, where rooPath is the current Format dir.
    :return: list of two variable, source file and destination file or path
    """
    logger.debug('here')
    # from_path, from_fn, to_path, to_fn
    src_path_expr = from_path.value
    src_file = str(from_fn.value).strip()
    xl_dest_path = str(to_path.value).strip()
    xl_dest_file = str(to_fn.value).strip()

    if "," in src_path_expr or "(" in src_path_expr or " / " in src_path_expr:
        src_path = eval(src_path_expr)
    elif "/" in src_path_expr or "\\" in src_path_expr:
        src_path = src_path_expr
    else:
        src_path = eval(os.path.expanduser(src_path_expr))

    src_file_w_path = os.path.join(src_path, src_file)  # FIXME test formatcopy, remove os.path.join

    if xl_dest_path == "" or xl_dest_path.lower() == 'none':  # no "to path" specified so use default, format dir
        dest_path = ROV_SETUP['format_path'] / ""  # FIXME NEEDED?
    else:
        dest_path = os.path.join(xl_dest_path, "")

    if xl_dest_file == "" or xl_dest_file.lower() == 'none':  # no "to path" specified so use default, format dir
        dest_file_w_path = os.path.join(dest_path, src_file)
    else:
        dest_file_w_path = os.path.join(dest_path, xl_dest_file)

    return [src_path, src_file_w_path, dest_path, dest_file_w_path]


def combine_formatfiles(orig_df, copy_formatfile_list_sheet):
    """
    gets files from FormatCopy sheet and puts contents into combined dataframe
    :param copy_formatfile_list_sheet: sheet containing rows of fromFile, toFile, processFlag
    :return: dataframe with data loaded
    """
    logger.info('combining format files')

    # Make sure column heading/locations are as expected and nothing was moved in Formatfile copy sheet
    check_file_headers(copy_formatfile_list_sheet,
                       [('A1', 'use'),
                        ('B1', 'fromFilePath'),
                        ('C1', 'fromfilename'),
                        ('D1', 'tofilepath')])

    # append records from FORMAT file into dataframe
    for copy_formatfile_flag, from_path, from_fn, to_path, to_fn, *_ in islice(copy_formatfile_list_sheet, 1, None):
        # Filelist is a sheet in setup opened using openpyxl
        if str(copy_formatfile_flag.value).strip().lower() == "x":
            src_path, src_file_w_path, dest_path, \
                dest_file_w_path = assign_formatcopy_vars(from_path, from_fn, to_path, to_fn)

            # make sure file and directories exist
            bad_path_exit(src_path)
            bad_file_exit(file=src_file_w_path)
            bad_path_exit(dest_path)
            bad_file_exit(file=dest_file_w_path)

            logger.debug("Ready to combine copied format file in combine_formatfiles: ", src_file_w_path)

            df_one_file = pd.read_csv(src_file_w_path, nrows=ROV_SETUP['rows_to_read_limit'],
                                      keep_default_na=False)

            missing_split_fields = set(ROV_SETUP['splitfile_field_list']) - set(df_one_file.columns.tolist())
            # missing_split_fields = set(df_one_file.columns.tolist()) ^ set(orig_df.columns.tolist())
            if missing_split_fields:
                # should we exit if field lists dont match exactly?  just show mismatches? what if one has a ton?
                exit_yes((f"The input fields of the copied Formatfile: \n\n'{src_file_w_path}' "
                          f"\n\nare missing the following fields to Split:\n\n "
                          f"{', '.join(missing_split_fields)}"
                          ))
            # orig_df = orig_df.append(df_one_file, ignore_index=True)
            orig_df = pd.concat([orig_df, df_one_file], ignore_index=True)

    return orig_df


def combine_files():
    """ combines all files specified in Setup > filelist into one and runs pivots """
    logger.info('starting combine_files')

    df_combined = pd.DataFrame()  # empty dataframe

    for fn, format_flag, combine_flag, update_fn, update_fields_cell, *_ in islice(ROV_SETUP['filelist_sheet'], 1, None):
        # islice starts in row 2 (index 1)

        if str(combine_flag.value).strip().lower() == "x":
            logger.debug(f"Ready to combine file '{fn.value}'")
            fnstem = pathlib.Path(fn.value).stem
            formatfile_w_path = ROV_SETUP['format_path'] / ('FORMATTED ' + fnstem + '.csv')

            bad_file_exit(formatfile_w_path)

            df_one_file = pd.read_csv(formatfile_w_path, nrows=ROV_SETUP['rows_to_read_limit'], keep_default_na=False)

            if df_one_file.columns.tolist() != ROV_SETUP['formatfile_field_list']:
                exit_yes((f"The input fields of '{fn.value}' do not match specified OP fields:\n\n"
                          ",".join(df_one_file.columns.tolist())
                          ))
            # df_combined = df_combined.append(df_one_file, ignore_index=True)
            df_combined = pd.concat([df_combined, df_one_file], ignore_index=True)

    # if flag set to copy external Format files, copy them after to match fields
    if ROV_SETUP['copy_other_format_files_flag']:
        df_combined = combine_formatfiles(df_combined, ROV_SETUP['copy_formatfile_filelist_sheet'])

    # delete fields specified in setup
    try:
        df_combined.drop(ROV_SETUP['combinefile_fields_to_delete_list'], axis=1, inplace=True)
    except KeyError:
        fields_missing_from_formatfile = set(ROV_SETUP['combinefile_fields_to_delete_list']) - set(df_combined.columns)
        exit_yes("Formatfile is missing the following fields to delete when creating Combinefile:"
                 f"\n\n{', '.join(fields_missing_from_formatfile)}"
                 )

    return df_combined


def bek_text_box(txt, title='', box_title="", buttons=None):
    """ Display text block with lines separated by \n and choice of buttons at bottom.

    Parameters
    ----------
    box_title :
    title :
    txt :
    buttons :

    """

    if buttons is None:
        buttons = ["OK", "Exit"]

    col_factor = 3  # to scale window equally
    row_factor = 30  # to scale window equally
    max_cols = len(max(txt.split("\n"), key=len)) * col_factor
    cols = max_cols
    # v_scroll = False
    col_limit = 80 * col_factor
    col_min = 50 * col_factor
    if cols > col_limit:
        # v_scroll = True
        cols = col_limit
    elif cols < col_min:
        cols = col_min

    noscroll = True
    row_limit = 80
    row_min = 6
    # max_rows = len(txt.split("\n"))
    # rows = max_rows
    rows = len(txt.split("\n"))
    if rows > row_limit:
        noscroll = False
        rows = row_limit
    elif rows < row_min:
        rows = row_min
#horizontal_scroll=h_scroll,
    layout = [
        [sg.Text(title, font=("Arial", 18))],
        [sg.Multiline(txt, autoscroll=False,  expand_x=True,no_scrollbar=noscroll,
                      expand_y=True, enable_events=True)],
        [sg.Button(text) for text in buttons],
    ]

    event, values = sg.Window(box_title, layout, titlebar_font=("Arial", 20), font=("Arial", 14),
                              use_custom_titlebar=True, size=(600, rows*row_factor), disable_close=True,
                              resizable=True, grab_anywhere=True).read(close=True)
    if event is not None:
        event = event.lower()
    return event


def get_dir_name(box_title, title2, initial_dir):
    """ show an "Open" dialog box and return the selected directory. Replaced askdirectory with PySimpleGUI
    :param title2:
    :type title2:
    """
    logger.debug('here')

    layout = [
        [sg.Text(title2, font=("Arial", 18))],
        [
         sg.Input(key="-IN-", expand_x=True),
         sg.FolderBrowse(initial_folder=os.path.expanduser(initial_dir))
         ],
        [sg.Button("Choose")],
    ]

    # event, values = sg.Window(heading_in_box, layout, size=(600, 100)).read(close=True)
    event, values = sg.Window(box_title, layout, titlebar_font=("Arial", 20), font=("Arial", 14),
                              size=(1000, 150), use_custom_titlebar=True).read(close=True)

    dir_name = values['-IN-']
    if dir_name == "":
        exit_yes("No directory name chosen")

    return dir_name


def get_file_name(box_title, title2, initial_dir):
    """ show an "Open" dialog box and return the selected file name. Replaced askopenfilename with pyeasygui
    :param title2: heading of the box
    :type title2: text next to input field
    """
    logger.debug('here')
    # "Select Sincere address export file 'all-parent-campaign-requests-yyyy-mm-dd.csv'"
    layout = [
        [sg.Text(title2, font=("Arial", 18))],
        [
         sg.Input(key="-IN-", expand_x=True),
         sg.FileBrowse(initial_folder=os.path.expanduser(initial_dir))
         ],
        [sg.Button("Choose")],
    ]

    # event, values = sg.Window(heading_in_box, layout, size=(600, 100)).read(close=True)
    event, values = sg.Window(box_title, layout, titlebar_font=("Arial", 20), font=("Arial", 14),
                              size=(1000, 150), use_custom_titlebar=True).read(close=True)
    # sg.Window.close()

    file_name = values['-IN-']
    if file_name == "":
        exit_yes("No file name chosen")

    return file_name


def convert_xlsx_to_csvs():
    logger.info('starting convert_xlsx_to_csv')

    str_xls_dir = get_dir_name("Select a DIRECTORY containing XLSX files to convert to CSVs",
                                            "XLSX Directory",
                                            INITIAL_CAMPAIGN_DIR)
    # str_xls_dir = "/Users/Denise/Library/CloudStorage/Dropbox/Postcard Files/TestInputFiles/TestCampaigns/TestXlsToCsv/Rawdata"

    str_csv_dir = get_dir_name("Select a DIRECTORY where converted CSVs will be placed",
                                            "CSV Directory",
                                            INITIAL_CAMPAIGN_DIR)
    # str_csv_dir = "/Users/Denise/Library/CloudStorage/Dropbox/Postcard Files/TestInputFiles/TestCampaigns/TestXlsToCsv/csv"

    xls_dir = pathlib.Path(str_xls_dir)
    csv_dir = pathlib.Path(str_csv_dir)

    # get list of files with xls or xlsx
    xls_files_w_path = list(xls_dir.glob("*.xls?"))

    # create list of dicts, each key a variable name corresponding to a file value,
    # eg xls_files_dict[0]['xls_name'] = 'abc.xlsx'.
    xls_files_dict = [{'xls_w_path': xls_w_path, 'xls_stem': xls_w_path.stem, 'xls_name': xls_w_path.name}
                      for xls_w_path
                      in xls_files_w_path]

    csv_files_w_path = list(csv_dir.glob("*.csv"))

    # same as done with xls above
    csv_files_dict = [{'csv_w_path': csv_w_path, 'csv_stem': csv_w_path.stem, 'csv_name': csv_w_path.name}
                      for csv_w_path
                      in csv_files_w_path]

    # set of file names (without .xls), csv below
    # xls_stem = {file_info['xls_stem'] for file_info in xls_files_dict}
    csv_stem = {file_info['csv_stem'] for file_info in csv_files_dict}

    # compare list of xls files without csv
    # xls_stem_wo_csv = xls_stem - csv_stem
    # compare list of xls files with matching csv
    # xls_stem_w_csv = xls_stem & csv_stem

    # get lists of file_info for xls with and without csvs
    xls_wo_csv = [xls_info for xls_info in xls_files_dict if xls_info['xls_stem'] not in csv_stem]
    xls_w_csv = [xls_info for xls_info in xls_files_dict if xls_info['xls_stem'] in csv_stem]

    newline = '\n'  # cause can't use \n in f-string
    # prompt stating two lists - skip with_csv, process no_csv - ok to continue?
    exit_yes_no(f"The following files WILL BE converted to csvs: {newline}" +
                newline.join([file_info['xls_name'] for file_info in xls_wo_csv]) +
                f"{newline}{newline}The following files already HAVE CSVs and will be SKIPPED: {newline}" +
                newline.join([file_info['xls_name'] for file_info in xls_w_csv]),
                "Continue with convert?"
                )

    # read each xlsx into dataframe with options and write out with same name
    for file_info in xls_wo_csv:
        logger.debug(f"Reading {file_info['xls_name']}")
        df = pd.read_excel(file_info['xls_w_path'])
        logger.debug(f"Writing {(file_info['xls_stem'] + '.csv')}")
        df.to_csv(csv_dir / (file_info['xls_stem'] + '.csv'), index=False)
        logger.debug('')


def main():
    """ processes multiple rawdata files using parameters set in a setup spreadsheet
    """
    def set_logfile_path(log_path):
        """ set log file path to location based on whether setup file is used or not.  If not, use EXE.
        using loguru could put in downloads (next line)
        but I'm locating it once we know what section
        is running, either with the setup file or the EXE if no setup, eg updating zip file.
        logfile = pathlib.Path.home() / "Downloads" / (Path(__file__).name + ".log")
    """

        logger.remove(0)
        logfile = log_path / (Path(__file__).name + ".log")
        try:
            os.remove(logfile)
        except Exception:
            pass

        logger.add(open(logfile, 'w'), level=log_level)
        logger.add(sys.stdout, level='INFO')


    logger.info("starting main")

    # shows all cols for dataframe head instead of truncating to first and last few
    pd.set_option('display.max_columns', None)

    # need path of exe/script for location of zip csv files
    if getattr(sys, 'frozen', False):
        EXE_PATH = pathlib.Path(sys.executable).parent
    elif __file__:
        EXE_PATH = pathlib.Path(__file__).parent
    else:
        EXE_PATH = None
    logger.debug(f"({EXE_PATH=}")

    ROV_SETUP['exe_path'] = EXE_PATH

    # choice = pymsgbox.confirm("What do you want to do?",
    #                           'Choose Action',
    #                           ['Format', 'Combine', 'Split', 'XLSXs to CSVs', 'Update Zip File', 'Exit'])

    choice = bek_text_box("What do you want to do?","Choose an Action",
                              '',
                              ['Format', 'Combine', 'Split', 'XLSXs to CSVs', 'Update Zip File', 'Exit'])

    if choice == 'xlsxs to csvs':
        set_logfile_path(EXE_PATH)
        logger.debug("chose 'XLSXs to CSVs'")

        convert_xlsx_to_csvs()
        pymsgbox.alert("Ran convert_xlsx to csv", "Convert Xlsxs to CSVs")

    elif choice == 'update zip file':
        set_logfile_path(EXE_PATH)
        logger.debug("chose 'Update Zip File'")

        import textwrap
        msg = f'''\
          'Updating Zip File' will read data from zip files purchased from zip-codes.com.

          Two files are read:
               {MAIN_ZIP_FILE}
               {MULTI_COUNTY_ZIP_FILE}

          Download the csv files.  These files are updated periodically 
          and the newest versions should be downloaded into the same
          directory as ROVCleaver.py before this update is run.
          A dictionary will be created and saved to 'ZIP_TO_COUNTY_LIST_FILE'
          and used in future runs of ROVCleaver.

          Aside: the file 'Unique_County_List.xlsx' in the ROVCleaver
          directory is updated every time ROVCleaver is run
          and can be used to get the required county names
          for those which need remapping.'''

        msg = textwrap.dedent(msg)
        exit_yes_no(msg,
                    'Update Zip Files',
                    display_exiting=False)

        # create py dict file; file defined at top of program
        create_zip_to_county_list_dict(ROV_SETUP['exe_path'] / MAIN_ZIP_FILE,
                                       ROV_SETUP['exe_path'] / MULTI_COUNTY_ZIP_FILE,
                                       ROV_SETUP['exe_path'] / ZIP_TO_COUNTY_LIST_FILE)
        logger.debug("done 'Update Zip File'")
        pymsgbox.alert("Ran Zip Dict file update", "Update zip files")

    elif choice == 'exit':
        logger.debug("chose 'exit'")
        exit()

    else:

        ROV_SETUP['setup_file_name'] = get_setup_file_name(INITIAL_CAMPAIGN_DIR)  # use TKInter to get the file/path of setup in campaign
        set_logfile_path(ROV_SETUP['setup_file_name'].parent)
        logger.debug("in 'else' to pick format, combine, split")

        # new for uniformat
        init_setup_dict()
        read_setup_vars(FIELD_DEF_COL_NUMERIC)
        format_setup_vars()

        logger.debug(f"{ROV_SETUP=}")

        # assign_rov_variables()  # read all the variables from the setup file and put the into ROV object

        create_field_lists()  # fill array with default field names of ' ' and add fields required by options selected

        create_dicts()  # create dicts and other file set up needed to run

        if choice == 'format':
            logger.debug("picked 'Format'")

            # check dir structure for formatted file creation. Dont create combine or split to indicate if dir is for
            # Format files copied in.
            bad_path_create(ROV_SETUP['format_path'])
            bad_path_create(ROV_SETUP['format_path'] / "Summary")
            bad_path_create(ROV_SETUP['format_path'] / "Removed")
            bad_path_create(ROV_SETUP['format_path'] / "Duplicates")

            # allow to exit if desired, eg flag not correct, imported code not right
            exit_for_unwanted_setup_options()

            # Loop through all files in filelist to be formatted
            process_format_files(ROV_SETUP['filelist_sheet'])

            pymsgbox.alert("Ran format section of main", "Alert")

        elif choice == 'combine':
            bad_path_create(os.path.expanduser(ROV_SETUP['combined_path']))
            logger.debug("picked 'Combine'")

            if ROV_SETUP['copy_other_format_files_flag']:
                exit_yes_no("COPY FORMAT files from other directory in combine?  OK?\n\n",
                            'COPY FORMAT FILES?')

                # COPY FILES via parm 'perform_copies=True' in addition to checking files/path existence
                formatfile_copy(ROV_SETUP['copy_formatfile_filelist_sheet'], perform_copies=True)

            if ROV_SETUP['run_last_code_flag']:
                display_imported_code(ROV_SETUP['last_code_sheet'], ROV_SETUP['last_code'])

            df = combine_files()  # V16.1 no longer writes out file

            if ROV_SETUP['id_dupes']:
                logger.debug(f"{ROV_SETUP['dupe_key_formula']=}")
                df['dupe_key'] = eval(ROV_SETUP['dupe_key_formula'])

                # sort for dupe check
                # sort_fields, ascending_vals = split_tuples(ROV.dupe_key_sort_tuples)
                # df.sort_values(by=['carol', 'dupe_key'], ascending=[False,True], inplace=True)
                df.sort_values(by=[k for k, v in ROV_SETUP['dupe_key_sort_tuples']],
                               ascending=[v for k, v in ROV_SETUP['dupe_key_sort_tuples']], inplace=True)

                identify_duplicates(df, 'dupe_key', 'dupe_id_field')

                if ROV_SETUP['run_last_code_flag']:
                    # *** Only run imported first_code and middle_code in format, not combine to keep things like remove
                    # assignment, random number from being overwritten.
                    # last_code can be run in combine since it's only setting remove code.
                    logger.debug('Ready to run last_code (remove code)')  # these prompts help if error in imported code

                    ROV_SETUP['last_code_to_import_module'].last_code_func(df, ROV_SETUP['dict_concentrated_addresses'],
                                                                  ROV_SETUP['expectedstate'])
                    # This is the function from the sheet with any parameters it needs

                    logger.debug('Ran last_code')  # these prompts help if error in imported code

                dupfile = ROV_SETUP['combined_path'] / ('DUPLICATES in ' + ROV_SETUP['OPFile'].stem + ".csv")

                logger.debug("Ready to sort by ['dupe_id_field','dupe_key']")  # to speed up copy to excel
                df.sort_values(by=['dupe_id_field', 'dupe_key'], inplace=True)

                logger.debug('Ready to copy dupes to CSV')  # these prompts help if error in imported code
                # df[df['dupe_id_field'] != 'X'].to_excel(dupfile, index=False)
                df[df['dupe_id_field'] != 'X'].to_csv(dupfile, index=False)

            test_df_clean = df[df['remove'] == '']

            if ROV_SETUP['sort_list']:  # true if not empty
                df.sort_values(by=ROV_SETUP['sort_list'], inplace=True)

            # run pivot reports on combined
            logger.debug('Ready to run pivots')  # these prompts help if error in imported code
            file = ROV_SETUP['combined_path'] / ('Summary of ' + ROV_SETUP['OPFile'].stem + ".xlsx")
            pivot_reports(df,
                          file,
                          ROV_SETUP['OPFile'].stem + '.csv',
                          ROV_SETUP['dict_concentrated_addresses'])

            # V16.1 moved writeing of combined file to after dedupe
            # Write out combined file
            combine_file = ROV_SETUP['combined_path'] / (ROV_SETUP['OPFile'].stem + '.csv')

            df.to_csv(combine_file, index=False)

            pd.set_option('display.max_columns', None)
            logger.info("\n", df.head(5), "\n\n")

            pymsgbox.alert("Ran combine section of main", "Alert")

        elif choice == 'split':
            logger.debug("picked 'Split'")

            # V11.1 moved to check only where needed
            bad_path_create(ROV_SETUP['split_path'])
            bad_path_create(ROV_SETUP['split_path_hold'])
            bad_path_create(ROV_SETUP['split_path_done'])

            split_files_for_sincere(ROV_SETUP['sub_split_limit'])
            logger.info('here')
            pymsgbox.alert("Ran split section of main", "Alert")


def init_setup_dict():
    """ assigns variables from cells in setup sheet and places them in global dictionary.  set some global variables"""
    logger.info('starting init_setup_dict')

    ROV_SETUP['setup_wb'] = load_workbook(filename=ROV_SETUP['setup_file_name'])
    ROV_SETUP['setup_sheet'] = ROV_SETUP['setup_wb']["Setup"]

    ROV_SETUP['filelist_sheet'] = ROV_SETUP['setup_wb']["FileList"]
    ROV_SETUP['copy_formatfile_filelist_sheet'] = ROV_SETUP['setup_wb']["FormatCopies"]

    # root_path is
    ROV_SETUP['root_path'] = ROV_SETUP['setup_file_name'].parent  # dir containing the setup file
    ROV_SETUP['root_path_one_level_up'] = ROV_SETUP['root_path'].parent
    ROV_SETUP['rawdata_path'] = ROV_SETUP['root_path'] / 'Rawdata'
    ROV_SETUP['format_path'] = ROV_SETUP['root_path'] / 'Formatted'
    ROV_SETUP['split_path'] = ROV_SETUP['root_path'] / 'Split'
    ROV_SETUP['split_path_hold'] = ROV_SETUP['root_path'] / 'Split' / 'Hold'
    ROV_SETUP['split_path_done'] = ROV_SETUP['root_path'] / 'Split' / 'Done'
    ROV_SETUP['combined_path'] = ROV_SETUP['root_path'] / 'Combined'
    ROV_SETUP['op_path'] = ROV_SETUP['root_path'] / 'UpdateFiles'

    # exit if no path exists for rawdata
    bad_path_exit(ROV_SETUP['rawdata_path'])

    # hardcoded file names created from first_code, middle_code, and last_code sheets
    ROV_SETUP['first_code'] = "first_code_to_import.py"
    ROV_SETUP['middle_code'] = "middle_code_to_import.py"
    ROV_SETUP['last_code'] = "last_code_to_import.py"

    # Check heading fields in filelist sheet of setup file to make sure it didn't move/change
    check_file_headers(ROV_SETUP['filelist_sheet'],
                       [('A1', 'file'),
                        ('B1', 'formatfile'),
                        ('C1', 'concatfile'),
                        ('D1', 'updatefile'),
                        ('E1', 'updatefilenames'),
                        ('F1', 'pull_group'),
                        ('G1', 'custom_field'),
                        ])
    a = 1


def format_setup_vars():
    """ reformat vars in setup_dict and erase temporary (starting with xl_)"""
    logger.debug(f"in format_setup_vars top {id(ROV_SETUP)=}")
    logger.info(f"starting format_setup_vars")

    # put worksheet object of string in ROV_DICT
    if ROV_SETUP['skip_selected_zip_match_flag']:
        ROV_SETUP['skip_selected_zip_sheet'] = ROV_SETUP['setup_wb'][ROV_SETUP['xl_skip_selected_zip_sheet']]
    ROV_SETUP['first_code_sheet'] = ROV_SETUP['setup_wb'][ROV_SETUP['xl_first_code_sheet']]
    ROV_SETUP['middle_code_sheet'] = ROV_SETUP['setup_wb'][ROV_SETUP['xl_middle_code_sheet']]
    ROV_SETUP['last_code_sheet'] = ROV_SETUP['setup_wb'][ROV_SETUP['xl_last_code_sheet']]

    max_len = max(len(ROV_SETUP['xl_inputfile_orig_list']),
                  len(ROV_SETUP['xl_inputfile_renamed_list']),
                  len(ROV_SETUP['xl_inputfile_type_list']),
                  )

    ROV_SETUP['inputfile_orig_list'] = pad_list(ROV_SETUP['xl_inputfile_orig_list'], max_len, pad_val="")
    ROV_SETUP['inputfile_renamed_list'] = pad_list(ROV_SETUP['xl_inputfile_renamed_list'], max_len, pad_val="")
    ROV_SETUP['inputfile_type_list'] = pad_list(ROV_SETUP['xl_inputfile_type_list'], max_len, pad_val="")

    ROV_SETUP['concentrated_addresses_wb'] = load_workbook(filename=ROV_SETUP['concentrated_addresses_file'])
    ROV_SETUP['concentrated_addresses_sheet'] = ROV_SETUP['concentrated_addresses_wb']["Addresses"]  # sheet "Addresses" hardcoded

    if ROV_SETUP['run_merge_data_flag']:
        bad_path_exit(ROV_SETUP['rawdata_path'])

    if ROV_SETUP['splitfield'] == '' and ROV_SETUP['sortchoice'] in [1, 2, 3]:
        exit_yes(f"Sort choice must be '4' if no no split field is specified"
                 f"\n\nSort choice is: {ROV_SETUP['sortchoice']}")
    else:
        if ROV_SETUP['sortchoice'] == 1:
            ROV_SETUP['sort_list'] = [ROV_SETUP['splitfield'], 'remove', 'zip', 'address', 'randnum']
        elif ROV_SETUP['sortchoice'] == 2:
            ROV_SETUP['sort_list'] = [ROV_SETUP['splitfield'], 'remove', 'zip', 'randnum']
        elif ROV_SETUP['sortchoice'] == 3:
            ROV_SETUP['sort_list'] = [ROV_SETUP['splitfield'], 'remove', 'randnum']
        else:
            ROV_SETUP['sort_list'] = []

    # combine pivot info into an object we can loop through
    ROV_SETUP['pivot_specs'] = []
    for piv_num in islice(range(6), 1, None):
        d = dict()
        d['pivot_fields'] = [fld.strip() for fld
                             in ROV_SETUP['xl_str_pivot_field' + str(piv_num)].split(',') if fld != '']
        d['pivot_by_cnt'] = ROV_SETUP['xl_str_pivot_field' + str(piv_num) + '_by_cnt']
        d['pivot_for_all'] = ROV_SETUP['xl_str_pivot_field' + str(piv_num) + '_all']
        ROV_SETUP['pivot_specs'].append(d)

    a = 1

    # sort dictionary DO NOT DO THIS - creates a new object even though reference is the same! https://stackoverflow.com/questions/61645769/sort-dict-in-place
    # local_dict = dict(sorted(local_dict.items()))

    # delete entries for master_dict that are temporary (begin with xl_)
    logger.info(f"{id(ROV_SETUP)=}")
    if REMOVE_XL_FROM_SETUP:
        # noinspection PyUnreachableCode
        for k, v in list(ROV_SETUP.items()):
            if k.startswith('xl_'):
                del ROV_SETUP[k]

    logger.debug(f"in format_setup_vars - finished")

    a = 1


def read_setup_vars(field_col):
    """ assigns variables from cells in setup sheet and places them in global dictionary.  set some global variables"""
    logger.info('starting read_setup_vars')

    # for testing read one row
    # row = list(ROV_SETUP['setup_sheet'].rows)[37]  # 4, 7, 12, 18, 37, 97, 140, 41 pythobj
    # logger.debug(f"{row[5].value=}")
    # row_list = row_to_list(row)
    # read_setup_var(ROV_SETUP, row_list)
    # logger.info(f"{ROV_SETUP=}")
    # exit()

    # returns a tuple of 0-indexed cell references.  row[0] = 1, first_row[0] is 'A'. use .value to get value in cell.
    for row in islice(list(ROV_SETUP['setup_sheet'].rows), 1, None):
        # print(f"{row[1].value=}")
        row_list = row_to_list(row)
        read_setup_var(row_list)


def read_setup_var(row_data):
    """ assigns variables from cells in setup sheet and places them in global dictionary.  set some global variables"""
    logger.debug(f"starting read_setup_var {row_data[ min(len(row_data)-1,FIELD_DEF_COL_NUMERIC) ]}")
    def len_tuple(tuple):
        """ check len of tuple where single value might not have a len and throw error (like bool)"""
        try:
            len(tuple)
        except:
            return -99
        return len(tuple)

    def return_func(var_type, str_case='l', str_strip='b', **kwargs):
        """ returns a function to convert a string to the passed type """

        def convert_bool(bool_val):
            """ bool('FALSE') return True so need better """
            if isinstance(bool_val, bool):
                return_val = bool_val
            else:
                if bool_val is None or bool_val.lower() not in ['true', 'false']:
                    raise ValueError('only allowable booleans are any case of true and false.  0/1 could be added to '
                                     'convert_bool code')
                elif bool_val.lower() == 'true':
                    return_val = True
                else:
                    return_val = False
            return return_val

        def my_expanduser(file_str):
            """ apply path and expanduser when expanduser does not take argument"""
            p = pathlib.Path(file_str)
            return p.expanduser()

        def my_python_obj(code_str):
            """ takes string of code and returns python object.
            Note: double quotes, tuples not supported - must be lists, multi items must be in list/[], bool must be
            true not True """
            python_object = json.loads(code_str)
            return python_object

        def my_lower_strip(valx, *extra_args, str_case='l', str_strip='b', **extra_kwargs):
            """ chain them together """
            val = valx
            if val is None:
                val = ''  # TODO is this better None?
            elif str_strip == 'b':
                val = val.strip()
            elif str_strip == 'l':
                val = val.lstrip()
            elif str_strip == 'r':
                val = val.rstrip()
            else:
                raise Exception  # 'Invalid option for str_strip parameter'

            if val is None:
                val = ''  # TODO is this better None?
            elif str_case == 'l':
                val = val.lower()
            elif str_case == 'u':
                val = val.upper()
            elif str_case == 'k':  # (k)eep the same
                pass
            else:
                raise Exception  # 'Invalid option for str_lower parameter'

            return val

        if var_type == 'str':
            return_fnc = my_lower_strip
        elif var_type == 'int':
            return_fnc = int
        elif var_type == 'bool':
            return_fnc = convert_bool
        elif var_type == 'float':
            return_fnc = float
        elif var_type == 'file':
            return_fnc = my_expanduser
        elif var_type == 'pythobj':
            return_fnc = my_python_obj
        else:
            raise ValueError
        return return_fnc

    def check_vars_string_for_errors(vars_string):
        """ check if var_string in proper format; raise exception if not."""
        try:
            vars_temp = eval(vars_string)
        except:
            logger.info(f"Field_vals '{vars_string}' is not a valid format.  check missing quotes.")
            raise Exception  # FIXME print this f"Field_vals '{vars_string}' is not a valid format.  check missing
            # quotes."

        # TODO optional/missing dicts caused problems in list comp so used loop.  way to use list comp?
        vars_list = [vars_temp[0]]  # put boolean at front
        for var, var_type, *other in islice(vars_temp, 1, None):
            if not other:
                vars_list.append([var.strip(), var_type.strip()])
            else:
                vars_list.append([var.strip(), var_type.strip()] + other)

        # must be tuple or list
        if not isinstance(vars_list, list):
            logger.info(
                f"Field_vals '{vars_string}' is not an expected tuple 'True,('xxx','yyy'),('aaa','bbb')...' format")
            raise Exception  # catch one explicit tuple like (a,int) as opposed to

        # first element must be bool (whether var is list of not)
        if not isinstance(vars_list[0], bool):
            logger.info(
                f"First part of Field_vals '{vars_string}' must be True=field with list of values or False=one or more "
                f"single variables "
                f"'('var', 'type')'")
            raise Exception  # catch one explicit tuple like (a,int) as opposed to

        # must have at least two element, bool for list and one variable specification
        if len(vars_list) < 2:
            logger.info(
                f"Field_vals '{vars_string}' must have at least two elements, the first True/False 'True,('xxx','yyy')...)")
            raise Exception  # catch one explicit tuple like (a,int) as opposed to

        # list (bool = True) but more than one var and type specified
        if vars_list[0] and len(vars_list) != 2:  # can only have one variable definition if type is a list (True)
            logger.info(
                f"If field_type is a list (first element is True) '{vars_string}' can only have one other element "
                f"('var_name','var_type') and all values must be of that type")
            raise Exception

        # check that var file_types are in accepted list
        bad_tuples = [(var, my_type) for var, my_type, *other
                      in islice(vars_list, 1, None)
                      if not isinstance(var, str) or my_type not in ['str', 'int', 'float', 'bool', 'file', 'pythobj']]
        # started adding 'pyth' for python code or objects but code needed to execute (eg have df defined)
        if bad_tuples:
            logger.info(
                f"Field_vals var/type in '{vars_string}' must have two elements, the first (the variable name) a string, "
                f"the second, the type, either str, int, float, file, or bool")
            raise Exception  # catch one explicit tuple like (a,int) as opposed to

        return vars_list

    # ROV_SETUP['rows_to_read_limit'] = ROV_SETUP['setup_sheet']['A19']
    if len(row_data) <= FIELD_DEF_COL_NUMERIC:  # no field data in field_vals column
        pass
    else:
        vars_string = row_data[FIELD_DEF_COL_NUMERIC]
        if vars_string is None:  # skip row - no variable info specified - blank or info row
            pass
        # only keep if field_keep is true
        elif str(row_data[FIELD_DEF_COL_NUMERIC - 1]).lower().strip() != 'true':
            pass
        else:

            # check structure and var tuples for input errors
            vars_list = check_vars_string_for_errors(vars_string)  # raises exceptions it error found

            if vars_list[0]:  # if True process list (first item is whether vars a list of values)
                my_dict = ({} if len(vars_list[1]) < 3 else vars_list[1][2])  # len()<3 means no dict if dict not
                # supplied
                func = return_func(vars_list[1][1], **my_dict)  # elements 1 (type) and 2(dict) for first and only var
                row_list = []
                for cell_val in islice(row_data, DATA_STARTS_COL_NUMERIC, None):  # TODO replace loop with list comprehension
                    if cell_val is not None:
                        row_list.append(func(cell_val, **my_dict))
                    else:
                        row_list.append("")
                ROV_SETUP[vars_list[1][0]] = row_list
            else:
                for index, (field_name, field_type, *my_dict) in enumerate(list(islice(vars_list, 1, None))):
                    my_dict = ({} if my_dict == [] else my_dict[0])
                    func = return_func(field_type, **my_dict)
                    # print(f"{index=}, {field_name=}, {field_type=}")
                    ROV_SETUP[field_name] = func(row_data[DATA_STARTS_COL_NUMERIC + index], **my_dict)
    a = 1


if __name__ == '__main__':

    main()
    a = 1
