""" ROVCleaver - V14.0 is a redo under the hood!  Taken from working V13.1
  Setup file stays basically the same BUT
  Code is cleaned up splitting out functions, PEP guides followed for var and function names
  Set up variables moved out of main scope and placed in a class to make them self contained
  Documentation in V14 setup file in sheet SetupDocumentation
"""

# # # ### deleted version comments prior to V14.0
# # # # V14.0 see top line - code rewritten: setup vars in common class, functions split, PEP followed, doc in V14 setup file
# # # #   Renamed imported code sher_concentration to return a tuple with (desc,remove reason) and removed flag
# V14.1 Clean up pivots reports
#  create exit_yes and raise exceptions instead of exit_yes_no for fatal flaws
# V15.0 2/5/23 Redo all field lists and process using zip
#   replaced xl_inputfields[0-2] with separate more descriptive lists and use zip to iterate over multi at a time
# V15.1 redo split to limit into routine chunk.  Filenames had to be kept in parallel and didn't chunk with no
# splitfield
#   6/29/23 replaced df.append with concat for zip update
# V16.0 code to identify and remove duplicates based on key field in setup
#   Included running 'last_code' in combine to flag duplicates
# V16.1 Move writing of combined csv to after dupes are removed
# V16.2 Add utility to convert xlsx to csv
# V16.3 Move csv and update zip utility outside of requiring setup file


# TODO: put filetype into select file and specify xlsx for Setup.
# TODO: change all tkinter & pymsgbox to simplegui

# TODO: replace prints with logger info/debug
# FIXME: list of imported code (first_code, etc) doesn't print on console if error found (raises first)
# FIXME make sure Format merge works - which fields missing are ok?

import ast
import collections
import copy
import datetime
import importlib
import inspect
import itertools
import math  # for ceil function
import os
import pathlib
import re
import shutil  # to copy file from other dirs
import webbrowser
from dataclasses import dataclass
from datetime import datetime
from itertools import islice  # to skip 1st row of iterated spreadsheet
from pathlib import *
from tkinter import Tk  # from tkinter import Tk for Python 3.x
from tkinter.filedialog import askopenfilename

import numpy as np
import pandas as pd
import pymsgbox
from openpyxl import load_workbook
from openpyxl.styles import Font
import PySimpleGUI as sg
import sys

# INITIAL_CAMPAIGN_DIR = os.path.expanduser(r"/Users/Denise/Dropbox/Postcard Files/InputFiles/Campaigns")
INITIAL_CAMPAIGN_DIR = os.path.expanduser(r"~/Dropbox/Postcard Files/InputFiles/Campaigns")
MAIN_ZIP_FILE = 'zip-codes-database-DELUXE-BUSINESS.csv'
MULTI_COUNTY_ZIP_FILE = 'zip-codes-database-MULTI-COUNTY.csv'
ZIP_TO_COUNTY_LIST_FILE = 'dictZipToCountyList.py'  # file where the numeric zip to county list is stored (ie  1011: ['hampden', 'hampshire'])
PROP_CONCENTRATION = 50
ZIP_CONCENTRATION = 10


def exit_yes_no(msg, title=None, display_exiting=False):
    """ makes this choice to continue one line"""
    if not title:
        title = "Exit?"
    choice = pymsgbox.confirm(msg, title, ['Yes', 'No'])
    if choice == "No":
        if display_exiting:
            pymsgbox.alert("Exiting", "Alert")
        exit()


def exit_yes(msg: str, title: str = None, *, errmsg: str = None) -> None:
    """ exits program after giving user a popup window and raising an error. """
    msg = (msg + "\n\n\nExiting." +
           f"\n\nCalled from {calling_func(level=3)}"
           f"\nCalled from {calling_func(level=2)}"
           f"\nCalled from {calling_func(level=1)}"
           )
    if not errmsg:
        errmsg = msg.replace("\n", " ")  # dont fill the console with linefeeds
    if not title:
        title = "** Exiting Program **"
    pymsgbox.alert(msg, title)
    raise Exception(errmsg)


def is_number(s: str) -> bool:
    """  Used as check, particularly before trying to set zip to numeric for lookup.
    expects param to be a string to trap all types of data.   """
    if s == np.NAN:  # this is needed- np.nan are int which are numbers
        return False
    try:
        float(s)
        return True
    except ValueError:
        return False


def clean_path(original_path):
    """ Return a path cleaned: adds ~ to start, returns path object using os.path """
    new_path = original_path
    if new_path[1] != '~':
        new_path = "~" + new_path
    if original_path[-1] != '/':
        new_path = new_path + '/'
    new_path = os.path.expanduser(new_path)
    return new_path


def max_used_col(ws, rw):
    """ Returns the column number (1 indexed) maximum non-none column in the input row of a sheet. """
    mxcol = 0
    for cell in reversed(ws[rw]):
        if cell.value is not None:
            mxcol = cell.col_idx
            break
    return mxcol


def range_to_list(ws, start_row, end_row, start_col, end_col):
    """ converts range of cells to a list of values
    V2.0 Worksheet to list. if one col, multi lists in list rather than one list of elements.  Best?
    """
    # tried using openpyl's cells = setup['A18': 'E19'] but parsing letter cols would be ugly so iterate rows.
    # openpyxl does not offer cells((r,c),(r,c)) using numerics
    final_list = []
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col,
                            max_col=end_col):
        row_list = []
        for cell in row:
            if cell.value is not None:
                row_list.append(str(cell.value).lower().strip())
            else:
                row_list.append("")
        final_list.append(row_list)
    if start_row == end_row:  # produce a one dimensional list instead of a one element list in a list
        final_list = final_list[0]
    return final_list


def clean_field(fld, case_convert='lower'):
    """
    returns a string in lower, strip, no space, no -, no ., no '
    can be used with dataframe like IP['clean2'] = IP['B'].apply(clean_field, convert_case='keep')
    1/28/23 added optional paramter convert_case deafualting to lower, as was done before, but allowing 'upper' or 'keep'.
    """
    return_fld = str(fld).strip().replace(" ", "").replace("'", "").replace(".", "").replace("-", "")
    if case_convert == 'lower':
        return_fld = return_fld.lower()
    elif case_convert == 'upper':
        return_fld = return_fld.upper()
    elif case_convert == 'keep':
        pass
    else:
        exit_yes(f"wrong value fed to clean_field parameter case_convert, '{case_convert}' - exiting")
    return return_fld


def autosize_xls_cols(ws):
    """ BEKs routine that works on the wks rather than df.  Datetime format set to width of 10. """
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                if cell.data_type == 'd':
                    date_width = 10
                else:
                    date_width = len(str(cell.value))
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), date_width))

    for col, value in dims.items():
        ws.column_dimensions[col].width = value + 1


def replace_boolean_column_vals(df, field_list):
    """ Check field in list for literal TRUE or FALSE because it gets converted to boolean and kills the program"""
    for field in field_list:
        if field in df.columns:
            df.loc[(df[field].str.upper() == "TRUE") | (df[field].str.upper() == "FALSE"), field] = ''


def bad_file_exit(file, msg=None):
    """ checks for file existence and exits if not found"""
    if msg is None:
        msg = f"File:\n\n'{file}'\n\ndoes not exist."
    if not Path(os.path.expanduser(file)).is_file():
        exit_yes(msg)


def bad_path_exit(path, msg=None):
    """ checks for directory existence and exits if not found"""
    if msg is None:
        msg = f"Directory:\n\n'{path}'\n\ndoes not exist."
    if not Path(os.path.expanduser(path)).exists():  # need expanduser for ~; only os works (not pathlib)
        # pymsgbox.alert(msg, "** Exiting via bad_path_exit **")
        # # FIXME: close TKINTER window here.  https://stackoverflow.com/questions/8009176/function-to-close-the-window-in-tkinter
        # exit()
        exit_yes(msg)


def bad_path_create(path, msg=None):
    """ checks for directory existence and creates if not found"""
    if msg is None:
        msg = ("Directory:\n\n" + path + "\n\ndoes not exist.  Creating." +
               "\n\nCalled from " + calling_func(level=2))
    if not os.path.isdir(path):
        pymsgbox.alert(msg, "Adding Directory via bad_path_create")
        os.makedirs(path)


def calling_func(level=0):
    """ returns the various levels of calling function.  0 is current, 1 is caller of current, etc """
    try:
        func = f"'{inspect.stack()[level][3]}', line #: {inspect.stack()[level][2]}"
    except Exception:
        func = f"** error ** inspect level too deep: {str(level)} called from {inspect.stack()[level][3]}"
    return func


def identify_duplicates(df, key, dupe_id_field):
    """
    creates a field in input df identifying duplicates as: X:not dupe;F:first,L:Last;D:other dupe;O:Other
    Parameters
    ----------
    df : input dataframe
    key : field checked as duplicate
    dupe_id_field : field in df filled with identifiers above
    """
    any_dupe_bool = df.duplicated([key], keep=False)  # all dupes as true/false
    # 'First' returns true for non-dupes as well as 'real' firsts, so must 'and' with all dupes
    first_bool = ~df.duplicated([key], keep='first') & any_dupe_bool  # first AND dupe as true
    last_bool = ~df.duplicated([key], keep='last') & any_dupe_bool  # last AND dupe as true

    # set different values to first and last so we can identify them as well as other dupes later
    first_numeric = np.where(first_bool, 2, 0)
    last_numeric = np.where(last_bool, 1, 0)
    any_dupe_numeric = np.where(any_dupe_bool, 1, 0)
    dupe_numeric = any_dupe_numeric + first_numeric + last_numeric  # sets each to different value
    dupe_alpha = [{0: 'X', 1: 'D', 2: 'L', 3: 'F', 4: 'O'}[element] for element in dupe_numeric]  #
    # note that setting non-dupe, 0, to blank above removes them from list so can not be merged into df
    if len(dupe_alpha) != len(df):
        print(f"*** assigning duplicate identifier error:  "
              f"Orig dataframe has {len(df)} rows but merged list has {len(dupe_alpha)}")
        raise Exception
    df[dupe_id_field] = dupe_alpha

    return


def split_tuples(tuple_list):
    """
    Accepts a list of tuples, ex for sort fields and ascending/descending, checks to make sure all list elements are
    tuples, then splits them into list of tuples keys and list of tuple values
    Parameters
    ----------
    tuple_list : list of tuples containing pairs of key/values

    Returns
    -------
    key_list: list of tuples keys
    val_list: list of tuple values
    """
    if tuple_list is None or tuple_list.strip() == '':
        return None, None
    else:
        tuple_list = eval(tuple_list)
        tuple_list = list(tuple_list)
        # check for list items that are not tuples - user error
        tuple_check_list = [True if isinstance(val, tuple) else False
                                 for val in tuple_list]
        if False in tuple_check_list:
            exit_yes('List of tuples contained a non-tuple', 'Error in Tuple List')
            raise ValueError
        key_list = [tuple[0] for tuple in tuple_list]
        val_list = [tuple[1] for tuple in tuple_list]

        return key_list, val_list


def merge_into_format_file(orig_df, update_file, cur_path):
    # TODO:  THIS NEEDS TO BE RE-WRITTEN AND CHECKED.  Copied and some refactoring but no logi or working checked.
    """ updates fields in format file from update file"""

    update_file_w_path = os.path.join(cur_path, update_file)
    # NOTE: update file is expected to have headers row 1 unlike raw data which may have title lines at top.

    update_df = read_file_to_df(update_file_w_path, **{'sheet_name': 0, 'nrows': ROV.xl_rows_to_read_limit,
                                                       'keep_default_na': False})

    update_df.columns = [x.strip().lower() for x in list(update_df.columns)]
    # rename df fields as lower() because col matching is case sensitive

    if set(ROV.xl_update_field_list).issubset(set(update_df.columns)):  # update field list all contained in df fields
        update_df = update_df[update_df.columns[update_df.columns.isin(ROV.xl_update_field_list)]]
    else:  # some specified update fields not on df
        extraelems = set(ROV.xl_update_field_list) - set(update_df.columns)
        exit_yes((f"Following field(s) in update list not in input.\n\n"
                  f"{os.linesep.join(extraelems)}"
                  ))

    # add the key fields specified in setup to both files
    # update_df['update_key'] = address + city
    update_df['update_key'] = eval(
        ROV.xl_update_file_key_formula)  # TODO: try sorting by keys to speed up dropping duplicates
    orig_df['master_key'] = eval(ROV.xl_orig_file_key_formula)

    # BEK2/22 hardcode field input as variable updateField to 'updateField'
    # update_df.rename(columns={updateField : 'update_field'}, inplace=True)  # why specify field name to create in setup then rename to hardcoded 'update_field'?

    # update_df.columns += '_updt' # add suffix to all columns names to keep separate from master. Prefix is df.columns = 'prefix_' + df.columns
    update_df.columns = [str(col) + '_updt' if col not in ['update_key'] else 'update_key' for col in
                         update_df.columns]  # add suffix to all columns names except 'update_key' to keep separate from master.

    # create df after dropping  duplicates based on key
    updt_w_duplicates_removed = update_df.drop_duplicates(subset=['update_key'])
    update_count_before_dups_removed = len(update_df)
    update_count_after_dups_removed = len(updt_w_duplicates_removed)
    num_dups_in_updt = update_count_before_dups_removed - update_count_after_dups_removed

    # Output a df of duplicate key values
    print("list of counts of updt duplicate key values in ", update_file)
    # dups = update_df.duplicated(subset=['update_key'], keep='first')  # not sure what this shows.  needed?
    dups_in_update = update_df[update_df.duplicated(subset=['update_key'],
                                                    keep=False)]  # Keep = false will show all dups; only first is kept above
    dups_in_update = dups_in_update.sort_values("update_key")
    print("list of updt duplicate key values (showing the first) in ", update_file)
    print(dups_in_update)  # does this work for df or do we need str() or something to avoid object?
    # write all duplicates to a file so we can take a look if desired
    file_of_dups = os.path.join(ROV.format_path, "Duplicates",
                                "UPDATE DUPLICATES " + str(PurePath(update_file).stem) + ".xlsx")
    dups_in_update.to_excel(file_of_dups, index=False)

    # must check for dups because key might not be unique (eg using truncated name)
    if num_dups_in_updt > 0:  # calculate above by comparing before and after de-dup
        exit_yes_no("Continue?  Dups in update will be removed.\n\n\nUpdate file\n" +
                    orig_df + "\n contains " + str(num_dups_in_updt) + " duplicates.",
                    "CHECK FOR DUPLICATE KEYS IN UPDATE",
                    display_exiting=False)

        print('Count before de-dup in update file', len(update_df))
        print('Count after de-dup in update file', len(updt_w_duplicates_removed))
        update_df = updt_w_duplicates_removed
        # print('New update_df without dupes - should be old updt_w_duplicates_removed ',update_df.shape[0])
        print('New update_df without dupes - should be old updt_w_duplicates_removed ', len(update_df))

    # update_df = update_df[['update_key', 'update_field']]  # keep only the key to prevent dup fields from being renamed in df with suffix (_x)

    master_w_dupkey_removed = orig_df.drop_duplicates(subset=['master_key'])  # new df with all duplicate master_key

    num_of_master_dups = len(orig_df) - len(master_w_dupkey_removed)

    # prompt if wanting to update if dupe on master present
    if num_of_master_dups > 0:
        exit_yes_no("Continue with dups?\n\n\nMaster file\n" + orig_df + "\ncontains " +
                    str(num_of_master_dups) + " duplicates.",
                    "CHECK FOR DUPLICATE KEYS IN MASTER",
                    display_exiting=False)

        # write dupes in master to file for reviewing
        file_of_dups = os.path.join(ROV.format_path, "Duplicates",
                                    "MASTER DUPLICATES " + str(PurePath(orig_df).stem) + ".xlsx")
        # writer = pd.ExcelWriter(file_of_dups)
        orig_dups = orig_df[orig_df.duplicated(subset=['master_key'], keep=False)]
        orig_dups = orig_dups.sort_values("master_key")
        orig_dups.to_excel(file_of_dups, index=False)

    #### Do the actual update!
    # In perfect world could use ", validate='1:1'" in merge to ensure integrity, but input data has dups,
    # so code was added to point out how many and allow continuance
    # Get counts before and after merge
    hold_orig_address_count = len(orig_df)
    orig_df = pd.merge(orig_df, update_df, how="left", left_on=['master_key'], right_on=['update_key'])
    after_address_count = len(orig_df)
    print('Orig address file had records: ', hold_orig_address_count, 'Result of merge had records: ',
          after_address_count)

    orig_number_created = after_address_count - hold_orig_address_count
    if orig_number_created > 0:  # the merge added records
        exit_yes_no(str(orig_number_created) + " addresses were duplicated by merge.  Continue?",
                    "EXTRA RECORDS WERE CREATED BY MERGE",
                    display_exiting=False)


def counties_to_xlsx(zipcsv, fn):
    """
    Returns an xlsx, fn, of unique state/county form purchased county data  https://www.zip-codes.com/.
    And a dictionary keying state/county to county_filename, countyToPrint, stateMixedCounty
    """
    # zipcsv = 'zip-codes-database-DELUXE-BUSINESS.csv'
    # multi_county_zip_file = 'zip-codes-database-MULTI-COUNTY.csv'
    zip_rows_to_read = 999999  # for testing
    # zip_rows_to_read = 999  # for testing

    main_zip_file = pd.read_csv(zipcsv, nrows=zip_rows_to_read, keep_default_na=False,
                                usecols=['State', 'County', 'CountyMixedCase'])

    main_zip_file.rename(
        columns={'ZipCode': 'zip', 'State': 'state', 'County': 'county', 'CountyMixedCase': 'county_mixedcase'},
        inplace=True)

    main_zip_file = main_zip_file.loc[main_zip_file['county'].str.strip() != ""]
    # military states like AA and AE have no county so remove

    main_zip_file['countyclean'] = main_zip_file['county'].apply(clean_field, case_convert='keep')
    main_zip_file['statecounty'] = main_zip_file['state'] + "-" + main_zip_file['countyclean']
    main_zip_file['county_filename'] = main_zip_file['county_mixedcase'].apply(clean_field, case_convert='keep')
    main_zip_file['countyToPrint'] = np.where(main_zip_file['county_mixedcase'].str[-4:] != "City",
                                              main_zip_file['county_mixedcase'] + " County",
                                              main_zip_file['county_mixedcase'])
    main_zip_file['statecounty_mixed'] = main_zip_file['state'] + "-" + main_zip_file['county_filename']

    unique_county = main_zip_file.drop_duplicates(subset=['statecounty'], keep='last')
    sorted_unique_county = unique_county.sort_values(['statecounty'], ascending=[True])

    sorted_unique_county.to_excel(fn, index=False,
                                  columns=['statecounty', 'county_filename', 'countyToPrint',
                                           'statecounty_mixed'])  # directory defaults to current python work

    # convert df to list so we can set up dictionary tuples easily
    df_list = unique_county[['statecounty', 'county_filename', 'countyToPrint', 'statecounty_mixed']].values.tolist()

    # set up list of tuples
    dict_list = [(statecounty, [county_filename, countyToPrint, statecounty_mixed])
                 for statecounty, county_filename, countyToPrint, statecounty_mixed
                 in df_list]

    return dict(dict_list)


def create_zip_to_county_list_dict(unique_zips, split_zips, text_file_for_created_dict):
    """
    Creates a dictionary with zip as key, list of counties as values using purchased zip data.
    Also writes the dictionary to a text file.
    Two input files: one-to-one zip/county and split zips.

    Data purchased from https://www.zip-codes.com/.
    Multi county and unique files are merged.
    Multi county does not contain unique zips.
    Unique contains multiple records for the same county for multiple cities in zip.
    """
    zip_rows_to_read = 999999  # for testing

    #os.path.join(ROV.root_path,
    main_zip_temp = pd.read_csv(unique_zips, nrows=zip_rows_to_read, keep_default_na=False,
                                usecols=['County', 'ZipCode'])
    multi_county_temp = pd.read_csv(split_zips, nrows=zip_rows_to_read, keep_default_na=False,
                                    usecols=['County', 'ZipCode'])

    # combined_temp = main_zip_temp.append(multi_county_temp, ignore_index=True)
    combined_temp = pd.concat([main_zip_temp, multi_county_temp], ignore_index=True)

    combined_temp['countyclean'] = combined_temp['County'].apply(clean_field, case_convert='lower')

    combined_temp2 = combined_temp[['ZipCode', 'countyclean']]  # keep only two cols
    unique_zip_county = combined_temp2.drop_duplicates(subset=['ZipCode', 'countyclean'], keep='last')

    df_for_dict = unique_zip_county.groupby(["ZipCode"], as_index=False).agg({'countyclean': list})
    zip_to_county_list_dict = dict(df_for_dict.values.tolist())

    with open(os.path.join(ROV.exe_path,text_file_for_created_dict), 'w') as f:
        print(zip_to_county_list_dict, file=f)

    return zip_to_county_list_dict


def find_header_row_in_file(file_with_path, header_string, header_col, sheet_name=None):
    """ identifies row with header by searching for header_string in header_col.  Used to skip blank and rows with titles."""
    if sheet_name is None:
        sheet_name = 0
    header_row = None

    # If the header identifying field is not in the first 30 rows assume something is wrong in the file
    df_temp = read_file_to_df(file_with_path, **{'header': None, 'sheet_name': sheet_name, 'nrows': 30,
                                                 'keep_default_na': True, 'dtype': str})

    excel_col_num = 1 + ord(header_col) - ord('A')  # converts alpha column to number. df index is 1st element.

    for row in df_temp.itertuples():
        if type(row[excel_col_num]) == str:  # otherwise blank cells/None cause problems in compare
            if row[excel_col_num].strip().lower() == header_string.strip().lower():
                header_row = row.Index
                break
    if header_row is None:
        exit_yes((f"File may be bad.\n\nThe header check string '{header_string}' "
                  f"was not found in column '{header_col}' "
                  "in the first 30 lines of input file:"
                  f"\n\n'{file_with_path}'"
                  ))
    return header_row


def single_pivot_report(df, index_fields, value_fields, sheet_name, single_piv_writer, second_pivot_by_count=False):
    """ run one pivot report and write out to worksheet leaving room for titles to be added later """

    # ExcelWorkbook = py.load_workbook(FilePath)
    # writer = pd.ExcelWriter(FilePath, engine='openpyxl')
    # writer.book = ExcelWorkbook

    sheet_name2 = sheet_name[:27] + '-Cnt'

    if sheet_name in single_piv_writer.book.sheetnames or sheet_name2 in single_piv_writer.book.sheetnames:
        file_counter_max = 0
        for sheet_ in single_piv_writer.book.sheetnames:
            if re.compile(r'-(\d+)$').search(sheet_):  # to avoid None if not found
                file_counter = re.compile(r'-(\d+)$').search(sheet_).group(1)
                if int(file_counter) > file_counter_max:
                    file_counter_max = file_counter
        file_counter_next = str(int(file_counter_max) + 1)

        sheet_name = sheet_name[:-(len(file_counter_next)+1)] + '-' + file_counter_next
        sheet_name2 = sheet_name2[:-(len(file_counter_next)+1)] + '-' + file_counter_next
    print(sheet_name2)

    if isinstance(index_fields, str):
        index_fields = [index_fields]  # to generalize to list and not force user to enter []s
    fields_not_in_file = set(index_fields) - set(df.columns)  # some fields not in df
    if fields_not_in_file:
        pymsgbox.alert("Specified index field on pivot not in dataframe:\n" + ','.join(fields_not_in_file) +
                       "\n\nAvailable fields are:\n" + ', '.join(set(df.columns)),
                       'Warning:Tabulate field not on file, continuing')
    else:
        df_pt = pd.pivot_table(df,
                               index=index_fields,
                               values=value_fields,
                               aggfunc='count',
                               margins=True)
        # take the df created by pivot and add percent field by dividing count by total number of addresses
        df_pt['Pct_of_Total'] = round(df_pt[value_fields] / df.shape[0] * 100, 1)
        df_pt.to_excel(single_piv_writer, sheet_name=sheet_name, startrow=5)

        if second_pivot_by_count:
            # same as above but sorted by count
            df_pt = df_pt.sort_values(value_fields, ascending=False)
            df_pt.to_excel(single_piv_writer, sheet_name=sheet_name2, startrow=5)


def pivot_reports(df, output_wks, input_fn, dict_address_concentration):
    """
    :param df: input dataframe
    :param output_wks: output spreadsheet
    :param input_fn: name of file report is being run on, goes in title3 only
    :param dict_address_concentration report concentrated prop description and removal reason
    """

    # ExcelWorkbook = py.load_workbook(FilePath)
    # writer = pd.ExcelWriter(FilePath, engine='openpyxl')
    # writer.book = ExcelWorkbook
    # if not 'testSheet' in book.sheetnames:
    #     book.create_sheet('testSheet')

    writer = pd.ExcelWriter(output_wks, engine='openpyxl')
    df_clean = df[df['remove'] == '']

    # Create summary sheet of Rawdata, Formatted and Removed
    # for other states, roll all counties in to one called "All Counties'
    df['countysummed'] = np.where(df['state'] == ROV.expectedstate, df['county'], "All Counties")

    # State by county for all including removed
    single_pivot_report(df, index_fields=['state', 'countysummed'], value_fields=['address'],
                        sheet_name='RawData by State-County', single_piv_writer=writer, second_pivot_by_count=False)

    if ROV.xl_add_filename_column_flag:
        single_pivot_report(df, index_fields=['filename'], value_fields=['address'], sheet_name='RawData by Filename',
                            single_piv_writer=writer, second_pivot_by_count=False)

    single_pivot_report(df, index_fields=['remove'], value_fields=['address'], sheet_name='Removed Reasons',
                        single_piv_writer=writer, second_pivot_by_count=False)

    if ROV.xl_splitfield != "none":
        single_pivot_report(df_clean, index_fields=ROV.xl_splitfield, value_fields=['address'],
                            sheet_name='Clean by ' + ROV.xl_splitfield[:22],
                            single_piv_writer=writer, second_pivot_by_count=True)

    # address concentration
    df_pt = pd.pivot_table(df, index=['state', 'county', 'city', 'address', 'remove'],
                           values=['lastname'],
                           aggfunc='count')
    df_from_query = df_pt.query("lastname >= " + str(PROP_CONCENTRATION))
    if len(df_from_query) > 0:
        df_from_query.rename(columns={'lastname': 'address_count'}, inplace=True)
        df_from_query.reset_index(inplace=True)

        # open browser windows for concentrated properties
        address_concentration_open_browser(df_from_query)

        # tried to eliminate error, "A value is trying to be set on a copy of a slice from a DataFrame", but couldn't so
        # isolated to with .copy() to show all is well
        dfq = df_from_query.copy()
        dfq['addrdesc'] = dfq.apply(lambda lam_row: get_addr_concentration(dict_address_concentration,
                                                                           lam_row.state, lam_row.county, lam_row.city,
                                                                           lam_row.address)[0], axis=1)

        dfq.to_excel(writer, sheet_name='Address GT ' + str(PROP_CONCENTRATION), startrow=5, index=False)

    # fields specific to county check - zip, county, zip/county match
    if ROV.xl_run_county_check_code_flag:
        print('running countyCheck pivots')
        # recode opposite of mismatch field value so we can sum
        df['match_county'] = np.where(df["mismatch_county"] == 1, 0, 1)

        # % county mismatched
        df_pt = pd.pivot_table(df, index=['state', 'countysummed'],
                               values=['address', 'mismatch_county', 'match_county'],
                               aggfunc={'address': 'count', 'mismatch_county': 'sum', 'match_county': 'sum'},
                               margins=True)
        df_pt['Pct_Mis_Matched'] = round(df_pt['mismatch_county'] / df_pt['address'] * 100, 1)
        df_pt['Pct_Matched'] = round(df_pt['match_county'] / df_pt['address'] * 100, 1)
        df_pt.to_excel(writer, sheet_name='County Match Summary', startrow=5)

        # zips w/ # occurrences in ZIP_CONCENTRATION, mismatch first
        df_pt = pd.pivot_table(df, index=['state', 'county', 'zip_county_list', 'zip', 'mismatch_county'],
                               values=['address'],
                               aggfunc={'address': 'count'})
        df_pt['Pct_of_Total'] = round(df_pt['address'] / df.shape[0] * 100, 1)
        df_from_query = df_pt.query("address >= " + str(ZIP_CONCENTRATION))
        df_from_query = df_from_query.sort_values(["mismatch_county", "address"], ascending=(False, False))
        df_from_query.rename(columns={'address': 'address_count'}, inplace=True)
        df_from_query.reset_index().to_excel(writer, sheet_name='Zips Over ' + str(ZIP_CONCENTRATION), startrow=5,
                                             index=False)

    # by report fields
    if ROV.xl_str_pivot_field1 != "none":
        dfx = (df if ROV.xl_str_pivot_field1_all.value else df_clean)
        universe = ('All-' if ROV.xl_str_pivot_field1_all.value else 'Cln-')
        single_pivot_report(dfx, index_fields=ROV.lst_pivot_field1, value_fields=['address'],
                            sheet_name=universe + ROV.xl_str_pivot_field1[:27],
                            single_piv_writer=writer,
                            second_pivot_by_count=(True if ROV.xl_str_pivot_field1_by_cnt.value else False))

    if ROV.xl_str_pivot_field2 != "none":
        dfx = (df if ROV.xl_str_pivot_field2_all.value else df_clean)
        universe = ('All-' if ROV.xl_str_pivot_field2_all.value else 'Cln-')
        single_pivot_report(dfx, index_fields=ROV.lst_pivot_field2, value_fields=['address'],
                            sheet_name=universe + ROV.xl_str_pivot_field2[:22],
                            single_piv_writer=writer,
                            second_pivot_by_count=(True if ROV.xl_str_pivot_field2_by_cnt.value else False))

    if ROV.xl_str_pivot_field3 != "none":
        dfx = (df if ROV.xl_str_pivot_field3_all.value else df_clean)
        universe = ('All-' if ROV.xl_str_pivot_field3_all.value else 'Cln-')
        single_pivot_report(dfx, index_fields=ROV.lst_pivot_field3, value_fields=['address'],
                            sheet_name=universe + ROV.xl_str_pivot_field3[:27],
                            single_piv_writer=writer,
                            second_pivot_by_count=(True if ROV.xl_str_pivot_field3_by_cnt.value else False))

    if ROV.xl_str_pivot_field4 != "none":
        dfx = (df if ROV.xl_str_pivot_field4_all.value else df_clean)
        universe = ('All-' if ROV.xl_str_pivot_field4_all.value else 'Cln-')
        single_pivot_report(dfx, index_fields=ROV.lst_pivot_field4, value_fields=['address'],
                            sheet_name=universe + ROV.xl_str_pivot_field4[:27],
                            single_piv_writer=writer,
                            second_pivot_by_count=(True if ROV.xl_str_pivot_field4_by_cnt.value else False))

    if ROV.xl_str_pivot_field5 != "none":
        dfx = (df if ROV.xl_str_pivot_field5_all.value else df_clean)
        universe = ('All-' if ROV.xl_str_pivot_field5_all.value else 'Cln-')
        single_pivot_report(dfx, index_fields=ROV.lst_pivot_field5, value_fields=['address'],
                            sheet_name=universe + ROV.xl_str_pivot_field5[:27],
                            single_piv_writer=writer,
                            second_pivot_by_count=(True if ROV.xl_str_pivot_field5_by_cnt.value else False))

    piv_wb = writer.book
    for ws in piv_wb.worksheets:
        autosize_xls_cols(ws)  # widen columns using BEK routine before wide titles

        ws["A1"] = "Address Summary Report"
        ws['A1'].font = Font(b=True, size=16)
        ws["A2"] = ws.title
        ws['A2'].font = Font(b=True, size=12)
        ws["A3"] = "Input File: " + input_fn
        ws['A3'].font = Font(b=True, size=12)
        ws["A4"] = datetime.now().strftime('%m/%d/%Y')

    print('out of countyCheck pivots')

    writer.close()


def create_import_code_from_sheet(setup_wb, sheet_with_python_code, output_file):
    """ Imports a spreadsheet sheet and creates properly indented python code from it based on sheet columns

    Parameters
    ----------
    setup_wb : the workbook object (setup.xlsx) that contains the code sheet
    sheet_with_python_code : the sheet containing the python code
    output_file : the .py that will receive the text python code
    """

    sheet = setup_wb[sheet_with_python_code]
    with open(os.path.join(ROV.exe_path,output_file), "w") as new_code_text:

        # Loop through sheet positions in excel file and indent as needed in python code (4 characters per col hardcoded)
        for rowidx, row_cells in enumerate(sheet.iter_rows()):
            # col and rows indexes needs +1 cause python 0 indexed, worksheet cells start at 1
            for colidx, cell in enumerate(row_cells):
                if sheet.cell(rowidx + 1, colidx + 1).value is not None and \
                        not sheet.cell(rowidx + 1, colidx + 1).value.strip().startswith('#'):
                    new_code_text.write(" " * colidx * 4 + sheet.cell(rowidx + 1, colidx + 1).value + '\n')
                    break  # dont go to subsequent columns once data is found to avoid extraneous info
    new_code_text.close()
    return new_code_text


def address_concentration_open_browser(df):
    """ uses address fields from query on address concentration pivot and google search with params to open
    browser windows for each address to decide if addreess should be excluded from carding."""
    openbrowser = True
    if ROV.xl_concentrated_address_browser_prompt_freq in [1, 2]:
        if ROV.xl_concentrated_address_browser_prompt_freq == 2:
            choice = pymsgbox.confirm("Open " + str(len(df)) + " tabs in browser?.  OK?\n\n",
                                      'Open browser windows?', ['Yes', 'No'])
            if choice == "No":
                openbrowser = False
        if openbrowser:
            df = df.reset_index()  # converts multi-index to columns
            for index, row in df.iterrows():
                st = row['state']
                cnt = row['county']
                cit = row['city']
                ad = row['address']
                webbrowser.open('https://www.google.com/search?q=' + st + "+" + cnt + "+" + "+" + cit + "+" + ad,
                                new=2)
        else:
            pass
    else:
        pass


def fields_to_list(base_list, new_fields):
    """ adds new fields to base lists.  exits if already in list."""
    new_fields_lst = [field.strip().lower() for field in new_fields.split(",")]
    same_fields = [field.strip().lower() for field in new_fields_lst if field in base_list]

    if not same_fields:
        base_list.extend(new_fields_lst)
    else:
        exit_yes(f"Field(s) '{str(same_fields)}' already exists on list.  Can not add it.")


def split_files_for_sincere(lim):
    """ splits main df into files by splitfield for loading into VoterLetters/Sincere.  splits large files into subs
    with counter if larger than limit.
    """
    def chunk_split_file(df, limit, split_path_hold, split_filename):
        """ pass a split file and needed parts and it will chunk it into sizes specified in setup as
        ROV.xl_sub_split_limit and postfix name with file-counter.
        split_filename is the root which csv and 'file x' is postfixed to """
        addresses_to_write = len(df)  # don't use df reference to save time

        if addresses_to_write <= limit:  # write one file
            split_file = os.path.join(split_path_hold, split_filename + ".csv")
            df.to_csv(split_file, index=False, columns=ROV.xl_splitfile_field_list)
            print(f"'{split_filename}' written, {len(df)} addresses.")
        else:  # need to create split files by looping
            for file_counter in range(1, math.ceil(addresses_to_write / limit) + 1):
                low_record = ((file_counter - 1) * limit)
                hi_record = (file_counter * limit) - 1

                split_file = os.path.join(split_path_hold, split_filename + " file-" + str(file_counter) + '.csv')
                df_chunk = df[low_record: hi_record + 1]
                df_chunk.to_csv(split_file, index=False, columns=ROV.xl_splitfile_field_list)
                # print("   split sub file ", file_counter)
                print(f" -'{split_filename + ' file-' + str(file_counter)}' written, {len(df_chunk)} "
                      f"addresses.")

    if lim == 0:
        lim = 99999999
    ip_stem = pathlib.Path(ROV.xl_OPFile).stem
    op_stem = pathlib.Path(ROV.xl_splitfnbase).stem

    # Ask if sorting by zip ok.  Assumes current sort order is how data was sorted under Combine.
    if ROV.xl_sortchoice in [1, 2]:
        exit_yes_no("Output is being sorted by zip. OK?",
                    'SORT BY ZIP?',
                    display_exiting=False)

    combinedfile_w_path = os.path.join(ROV.combined_path, ip_stem + '.csv')

    try:
        df_combo_w_no_remove = pd.read_csv(combinedfile_w_path, header=0, keep_default_na=False)
    except:
        exit_yes("Unable to read  the Combinedfile in the split step.  Was 'Combine' run?"
                 f"\n\nMissing file:\n\n{combinedfile_w_path}"
                 )

    df_combo_w_no_remove = df_combo_w_no_remove[df_combo_w_no_remove["remove"] == ""]
    exit_yes_no(f"Split will process {len(df_combo_w_no_remove)} clean addresses.",
                'SPLIT RECORDS',
                display_exiting=False)

    if ROV.sort_list:
        df_combo_w_no_remove.sort_values(by=ROV.sort_list, inplace=True)

    fields_missing_from_combinefile = set(ROV.xl_splitfile_field_list) - set(df_combo_w_no_remove.columns)
    if fields_missing_from_combinefile:
        exit_yes("Combinefile is missing the following fields to write to Splitfiles:"
                 f"\n\n{', '.join(fields_missing_from_combinefile)}"
                 )

    if ROV.xl_splitfield == 'none':  # no split field specified so write out one file with name "Combined"
        split_filename = ROV.expectedstate + '-' + "Combined " + op_stem
        chunk_split_file(df_combo_w_no_remove, lim, ROV.split_path_hold, split_filename)

    else:
        if ROV.xl_splitfield.lower() == 'county':
            splitfield = 'statecounty'
        else:
            splitfield = ROV.xl_splitfield
        unique_split_values = df_combo_w_no_remove[splitfield].unique()
        unique_split_values.sort()

        # for each - write out a csv file.
        for splitfield_value in unique_split_values:
            # print("split " + splitfield_value)
            df_one_splifield = df_combo_w_no_remove[df_combo_w_no_remove[splitfield] == splitfield_value]

            if ROV.xl_splitfield.lower() == 'county':
                split_filename = ROV.dict_statecounty_to_alt_formats[splitfield_value][2]
                # get the format of county we want to use for filename using county lookup
            else:
                split_filename = ROV.expectedstate + '-' + splitfield_value
            split_filename = split_filename + "- " + op_stem

            # write out file, broken into chinks if needed
            chunk_split_file(df_one_splifield, lim, ROV.split_path_hold, split_filename)


def get_addr_concentration(dict_addr_rem, state, county, city, address):
    # get_addr_remove_OLD has passed dictionary which is causing problems
    """
    Given location for a concentrated address, return the remove reason (blank if ok) or description based on dictionary
    Returns a tuple [0] is description, [1] is reason for removal
    :param dict_addr_rem: given a tuple of state, county, city, address, returns a two element list of remove code or description
    :param state, county, city, address
    """
    state = state.lower().strip()
    county = county.lower().strip()
    city = city.lower().strip()
    address = address.lower().strip()
    addr_desc_rem_tuple = dict_addr_rem.get(tuple([state, county, city, address]), ['missing', 'missing'])

    return addr_desc_rem_tuple


def check_county_to_zips(df, zipskip_list, dict_statecounty):
    """ check county by comparing it to zip lookup.  sets flag for mismatches and saves vars along the way. """
    print("\nFilling zip and county info")

    df['orig_county'] = df['county']
    df['clean_county'] = df['county'].apply(clean_field)
    df['numzip'] = df['zip'].map(lambda x: (int(x) if is_number(x) else 0))

    df['statecounty'] = df[['state', 'clean_county']] \
        .apply(lambda row: (row['state'] + "-" + row['clean_county']).upper(),
               axis="columns")

    df['clean_county'] = df['county'].apply(clean_field)  # todo this is a repeat of line above

    # TODO: below mixed state counties with counties and makes it impossible to produce simple summaries of data
    # TODO: add recode of all non-expected state to 'all counties in non-expected state'
    df['county'] = df['statecounty'].map(lambda x: dict_statecounty.get(x, [x+'-statecounty not found'])[0])
    # df['county'] = np.where(df['state'] == ROV.expectedstate, df['county'], "All counties in non-expected-state")

    df['zip_county_list'] = df['numzip'].map(lambda x: ",".join(ROV.dict_zip_to_countylist.get(x, ['mismatch'])))
    # set res_county to string of zip lookup list

    df['mismatch_county'] = df.apply(lambda dfx: (1 if dfx['clean_county'] not in dfx['zip_county_list']
                                                  else 0), axis=1)
    # check for county in list to account for split zips

    if ROV.xl_skip_selected_zip_match_flag:  # use zip skip list to reset county/zip mismatches to 0
        df['mismatch_county'] = df.apply(lambda row:
                                         (0 if (row['county'].lower(), row['zip']) in zipskip_list else row[
                                             'mismatch_county']), axis='columns')

    print("Done filling zip and county info")


def read_file_to_df(file_with_path, **param_dict):
    """reads either xlsx or csv into a dataframe using parms passed in dictionary. Non-applicable parms
    are skipped."""
    if PurePath(file_with_path).suffix.lower() == '.xlsx':
        filtered_dict = {k: v for k, v in param_dict.items()
                         if k in [p.name for p in inspect.signature(pd.read_excel).parameters.values()]}
        df_temp = pd.read_excel(file_with_path, **filtered_dict)
    elif PurePath(file_with_path).suffix.lower() == '.csv':
        filtered_dict = {k: v for k, v in param_dict.items()
                         if k in [p.name for p in inspect.signature(pd.read_csv).parameters.values()]}
        df_temp = pd.read_csv(file_with_path, **filtered_dict)
    else:
        df_temp = None
        exit_yes((f"Bad file type on input - not xlsx or csv."
                  f"\n\nFile: '{file_with_path}'"
                  ))
    return df_temp


def get_setup_file_name():
    """ use Tkinter to get name of setup workbook for desired campaign.  checks version of ROVCleaver program to
    version of setup xlsx using file name
    """
    Tk().withdraw()  # we don't want a full GUI, so keep the root window from appearing

    # Get version of ROVCleaver being run to use in setup prompt and later as check
    filename_ver_string = (re.search(r' v[1-9]+[ .-]', os.path.basename(__file__), flags=re.IGNORECASE))
    if filename_ver_string:  # returns Non if no value returned from RegEx
        filename_ver = filename_ver_string.group()[2:-1]
    else:
        filename_ver = None
    if not filename_ver:
        exit_yes((f"An improper ROVCleaver version numer was specified.\n"
                  f"The ROVCleaver name must contain Vxx (with space before, space or . after)\n"
                  f"where xx is the version number."
                  ))

    # Use this flag when testing - False allows hardcoding input from alternate starting directory
    if True:
        # show an "Open" dialog box and return the path to the selected file
        # V13.1 parameterize start directory and remove '/Users/Denise' reference
        # ROV.setup_file_name = askopenfilename(
        #     initialdir=INITIAL_CAMPAIGN_DIR,
        #     title="Select ROVCleaver setup file SetupFormat" + filename_ver, filetypes=(
        #         ("Excel files", "*.xlsx *.xls"),))

        # V16.2 use PySimpleGUI instead of tkinter
        ROV.setup_file_name = get_file_name("Select Setup File",
                                           f"Select ROVCleaver setup file SetupFormat{filename_ver}.x",
                                            INITIAL_CAMPAIGN_DIR)
    else:
        # Hardcode in TEST INPUT FILE directory for repetitive testing
        ROV.setup_file_name = os.path.expanduser(
            "/Users/Denise/Dropbox/Postcard Files/TestInputFiles/Campaigns/TEST GA Runoff 11-2022/ROVCleaver PDI GA Runoff GOTV 11-2022 SetupFormat12.0.xlsx")

        exit_yes_no("Running hardcoded Setup file.  OK?\n\n" + ROV.setup_file_name,
                    'RUN IN TEST?',
                    display_exiting=False)

    # V13.1 Ensure match between extracted version number from ROVCleaver being run and setup file
    setup_ver_string = (re.search(r'setupformat[1-9]+[ .]', ROV.setup_file_name, flags=re.IGNORECASE))
    if setup_ver_string:  # returns Non if no value returned from RegEx
        setup_ver = setup_ver_string.group()[11:-1]
    else:
        setup_ver = None

    if not setup_ver_string or not filename_ver_string or setup_ver != filename_ver:
        exit_yes(("The format of the ROVCleaver being run (xx in Vxx in name)\n"
                  "does not match the setup file version (ROV.setupFormatxx)"
                  "\n\n   or"
                  "\n\nThe version format specification(s) are not correct in one or both files."
                  "\n\nEither pick a different setup file or "
                  "run a compatible version of ROVCleaver."
                  ))


def assign_rov_variables():
    """ assigns variables from cells in setup sheet and places them in global object.  set some global variables"""

    # assign references to setup sheet in object ROV
    ROV.setup_wb = load_workbook(filename=ROV.setup_file_name)  # use openpyxl
    ROV.setup_sheet = ROV.setup_wb["Setup"]
    ROV.filelist_sheet = ROV.setup_wb["FileList"]
    ROV.copy_formatfile_filelist_sheet = ROV.setup_wb["FormatCopies"]

    # Check heading fields in filelist sheet of setup file to make sure it didn't move/change
    check_file_headers(ROV.filelist_sheet,
                       [('A1', 'file'),
                        ('B1', 'formatfile'),
                        ('C1', 'concatfile'),
                        ('D1', 'updatefile'),
                        ('E1', 'updatefilenames')])

    ### assign paths
    # need path of exe/script for location of zip csv files
    if getattr(sys, 'frozen', False):
        EXE_PATH = pathlib.Path(sys.executable).parents[0]
    elif __file__:
        EXE_PATH = pathlib.Path(__file__).parents[0]
    else:
        EXE_PATH = None
    print(f"({EXE_PATH=}")


    ROV.exe_path = str(EXE_PATH)
    ROV.root_path = os.path.split(os.path.abspath(ROV.setup_file_name))[0]
    ROV.root_path_one_level_up = str(pathlib.PurePath(ROV.root_path).parent)
    ROV.rawdata_path = os.path.join(ROV.root_path, 'Rawdata')
    ROV.format_path = os.path.join(ROV.root_path, 'Formatted')
    ROV.split_path = os.path.join(ROV.root_path, 'Split')
    ROV.split_path_hold = os.path.join(ROV.root_path, 'Split', 'Hold')
    ROV.split_path_done = os.path.join(ROV.root_path, 'Split', 'Done')
    ROV.combined_path = os.path.join(ROV.root_path, 'Combined')
    ROV.op_path = os.path.join(ROV.root_path, 'UpdateFiles')

    # read setup fields
    ROV.expectedstate = str(ROV.setup_sheet['A8'].value.strip())
    ROV.xl_concentrated_addresses_file = os.path.expanduser(ROV.setup_sheet['A13'].value)
    ROV.concentrated_addresses_wb = load_workbook(filename=ROV.xl_concentrated_addresses_file)
    ROV.concentrated_addresses_sheet = ROV.concentrated_addresses_wb["Addresses"]  # sheet "Addresses" hardcoded
    ROV.xl_rows_to_read_limit = ROV.setup_sheet['A19'].value
    ROV.xl_concentrated_address_browser_prompt_freq = ROV.setup_sheet['A24'].value
    ROV.xl_sortchoice = ROV.setup_sheet['A29'].value

    ROV.xl_id_dupes = ROV.setup_sheet['A35'].value
    ROV.xl_dupe_key_formula = ROV.setup_sheet['A38'].value
    ROV.xl_dupe_key_sort_tuples = ROV.setup_sheet['A42'].value
    if ROV.xl_id_dupes != 'none':
        # ROV.dupe_key_sort_fields / order use function that splits tuples = [field.strip().upper() for field in
        # ROV.xl_dupe_key_sort_tuples.split(",")]
        ROV.dupe_key_sort_fields, ROV.dupe_key_sort_orders = split_tuples(ROV.xl_dupe_key_sort_tuples)
        # ROV.dupe_keys_to_keep.append('X')

    ROV.xl_OPFile = str(ROV.setup_sheet['A62'].value).strip()
    ROV.xl_splitfield = str(ROV.setup_sheet['A94'].value).strip().lower()

    ROV.xl_str_pivot_field1 = str(ROV.setup_sheet['A98'].value).strip().lower()
    ROV.xl_str_pivot_field1_by_cnt = ROV.setup_sheet['B98']  # set to True to report by count
    ROV.xl_str_pivot_field1_all = ROV.setup_sheet['C98']  # True reports all; False clean only
    if ROV.xl_str_pivot_field1 != 'none':
        ROV.lst_pivot_field1 = [field.strip().lower() for field in ROV.xl_str_pivot_field1.split(",")]

    ROV.xl_str_pivot_field2 = str(ROV.setup_sheet['A99'].value).strip().lower()
    ROV.xl_str_pivot_field2_by_cnt = ROV.setup_sheet['B99']  # set to True to report by count
    ROV.xl_str_pivot_field2_all = ROV.setup_sheet['C99']  # True reports all; False clean only
    if ROV.xl_str_pivot_field2 != 'none':
        ROV.lst_pivot_field2 = [field.strip().lower() for field in ROV.xl_str_pivot_field2.split(",")]

    ROV.xl_str_pivot_field3 = str(ROV.setup_sheet['A100'].value).strip().lower()
    ROV.xl_str_pivot_field3_by_cnt = ROV.setup_sheet['B100']  # set to True to report by count
    ROV.xl_str_pivot_field3_all = ROV.setup_sheet['C100']  # True reports all; False clean only
    if ROV.xl_str_pivot_field3 != 'none':
        ROV.lst_pivot_field3 = [field.strip().lower() for field in ROV.xl_str_pivot_field3.split(",")]

    ROV.xl_str_pivot_field4 = str(ROV.setup_sheet['A101'].value).strip().lower()
    ROV.xl_str_pivot_field4_by_cnt = ROV.setup_sheet['B101']  # set to True to report by count
    ROV.xl_str_pivot_field4_all = ROV.setup_sheet['C101']  # True reports all; False clean only
    if ROV.xl_str_pivot_field4 != 'none':
        ROV.lst_pivot_field4 = [field.strip().lower() for field in ROV.xl_str_pivot_field4.split(",")]

    ROV.xl_str_pivot_field5 = str(ROV.setup_sheet['A102'].value).strip().lower()
    ROV.xl_str_pivot_field5_by_cnt = ROV.setup_sheet['B102']  # set to True to report by count
    ROV.xl_str_pivot_field5_all = ROV.setup_sheet['C102']  # True reports all; False clean only
    if ROV.xl_str_pivot_field5 != 'none':
        ROV.lst_pivot_field5 = [field.strip().lower() for field in ROV.xl_str_pivot_field5.split(",")]

    ROV.xl_splitfnbase = str(ROV.setup_sheet['A105'].value.strip())
    ROV.xl_sub_split_limit = ROV.setup_sheet['A110'].value

    # Read setup file variables
    ROV.xl_inputfield_list = range_to_list(ROV.setup_sheet, 141, 143, 1,
                                           max(max_used_col(ROV.setup_sheet, 141),
                                               max_used_col(ROV.setup_sheet, 142),
                                               max_used_col(ROV.setup_sheet, 143)
                                               )
                                           )

    ROV.xl_inputfile_orig_list = \
        range_to_list(ROV.setup_sheet, 141, 141, 1, max_used_col(ROV.setup_sheet, 141))
    ROV.xl_inputfile_renamed_list = \
        range_to_list(ROV.setup_sheet, 142, 142, 1, max_used_col(ROV.setup_sheet, 142))
    ROV.xl_inputfile_type_list = \
        range_to_list(ROV.setup_sheet, 143, 143, 1, max_used_col(ROV.setup_sheet, 143))

    ROV.expected_field_list = copy.deepcopy(ROV.xl_inputfield_list)

    ROV.xl_combinefile_fields_to_delete_list = \
        range_to_list(ROV.setup_sheet, 151, 151, 1, max_used_col(ROV.setup_sheet, 151))
    # USAGE CHANGED: Now optional list of fields to delete from Formatfile when creating Combine.
    # Default is to copy all fields into Combine, but this can be used to slim file.

    ROV.xl_splitfile_field_list = range_to_list(ROV.setup_sheet, 156, 156, 1, max_used_col(ROV.setup_sheet, 156))
    ROV.xl_copy_other_format_files_flag = ROV.setup_sheet['A188'].value
    ROV.xl_check_header_string = str(ROV.setup_sheet['A192'].value.strip()).lower()
    ROV.xl_strcheck_header_col = str(ROV.setup_sheet['A195'].value.strip())
    ROV.xl_run_county_check_code_flag = ROV.setup_sheet['A200'].value
    # ROV.xl_check_statecounty = ROV.setup_sheet['A205'].value
    ROV.xl_skip_selected_zip_match_flag = ROV.setup_sheet['A211'].value
    ROV.xl_skip_selected_zip_sheet = ROV.setup_sheet['A215'].value
    ROV.xl_run_first_code_flag = ROV.setup_sheet['A220'].value
    ROV.xl_first_code_sheet = ROV.setup_sheet['A224'].value
    ROV.xl_run_middle_code_flag = ROV.setup_sheet['A229'].value
    ROV.xl_middle_code_sheet = ROV.setup_sheet['A233'].value
    ROV.xl_run_last_code_flag = ROV.setup_sheet['A238'].value
    ROV.xl_last_code_sheet = ROV.setup_sheet['A242'].value
    ROV.xl_add_filename_column_flag = ROV.setup_sheet['A248'].value  # BOOLEAN Add input filename as column
    ROV.xl_add_random_number_column_flag = ROV.setup_sheet['A253'].value
    ROV.xl_run_merge_data_flag = ROV.setup_sheet['A289'].value
    ROV.xl_orig_file_key_formula = ROV.setup_sheet['A294'].value
    ROV.xl_update_file_key_formula = ROV.setup_sheet['A299'].value
    ROV.xl_update_field_list = range_to_list(ROV.setup_sheet, 304, 304, 1, max_used_col(ROV.setup_sheet, 304))
    # list of fields kept on the update files immediately after they are read.

    ROV.xl_update_field_list = [x.strip().lower() for x in ROV.xl_update_field_list]

    # hardcoded file names created from first_code, middle_code, and last_code sheets
    ROV.first_code = "first_code_to_import.py"
    ROV.middle_code = "middle_code_to_import.py"
    ROV.last_code = "last_code_to_import.py"

    # BEK3/5/22 set sortvalues then sort in format, combine and split
    if ROV.xl_splitfield == 'none' and ROV.xl_sortchoice in [1, 2, 3]:
        exit_yes(f"Sort choice must be '4' if no no split field is specified"
                 f"\n\nSort choice is: {ROV.xl_sortchoice}")
    else:
        if ROV.xl_sortchoice == 1:
            ROV.sort_list = [ROV.xl_splitfield, 'remove', 'zip', 'address', 'randnum']
        elif ROV.xl_sortchoice == 2:
            ROV.sort_list = [ROV.xl_splitfield, 'remove', 'zip', 'randnum']
        elif ROV.xl_sortchoice == 3:
            ROV.sort_list = [ROV.xl_splitfield, 'remove', 'randnum']
        else:
            ROV.sort_list = []

    ### Done inputing fields - make sure specified input paths exist or EXIT
    bad_path_exit(ROV.rawdata_path)
    if ROV.xl_run_merge_data_flag:
        bad_path_exit(ROV.rawdata_path)


def exit_for_unwanted_setup_options():
    """ verifies setup options when program is run allowing to exit and edit setup
    """
    if ROV.xl_run_county_check_code_flag and ROV.xl_splitfield.lower() != 'county':
        exit_yes_no(("You are checking for zip/county mismatches"
                     "\nbut you are not splitting by county."
                     "\n\n\nIs this what you meant?"))

    if ROV.xl_skip_selected_zip_match_flag:
        str_zipskip = str()
        for county, zip in ROV.zipskip_list:
            str_zipskip += f"{county:<20} {zip}\n"

        exit_yes_no("Counties and zip to ignore in matching "
                    f"(listed in 'ZipSkip' setup sheet): \n{str_zipskip}",
                    "Zip Codes to Ignore in County Matching")

    if ROV.xl_sortchoice in [1, 2]:
        exit_yes_no("Sorting by other than county / random number can cause ugly clumps of addresses "
                    "(eg all PO boxes).  OK?",
                    "Sort Order")

    if ROV.xl_copy_other_format_files_flag:
        exit_yes_no("Copy FORMAT files listed in setup sheet 'FormatCopies' during combine?\n\n",
                    'COPY FORMAT FILES?')

        formatfile_copy(ROV.copy_formatfile_filelist_sheet, perform_copies=False)

    if ROV.xl_run_first_code_flag:
        # display code and prompt if it should be run
        display_imported_code(ROV.xl_first_code_sheet, ROV.first_code)

    if ROV.xl_run_middle_code_flag:
        display_imported_code(ROV.xl_middle_code_sheet, ROV.middle_code)

    if ROV.xl_run_last_code_flag:
        display_imported_code(ROV.xl_last_code_sheet, ROV.last_code)


def create_field_lists():
    """ fill array with default field names of ' ' and add fields required by options selected"""

    ROV.xl_inputfile_orig_list = [field.strip().lower() for field in ROV.xl_inputfile_orig_list]
    ROV.xl_inputfile_renamed_list = [field.strip().lower() for field in ROV.xl_inputfile_renamed_list]
    ROV.xl_inputfile_type_list = [field.strip().lower() for field in ROV.xl_inputfile_type_list]

    ROV.xl_splitfile_field_list = [field.strip().lower() for field in ROV.xl_splitfile_field_list]

    # create 'formatfile' field list by replacing blanks with inputfile names and replacing those to be renamed,
    # keep 'x' so we can zip with input fields
    ROV.formatfile_field_list_to_zip = [input_field if rename_field == '' else rename_field
                                        for (input_field, rename_field)
                                        in itertools.zip_longest(ROV.xl_inputfile_orig_list,
                                                                 ROV.xl_inputfile_renamed_list)]

    ROV.inputfile_delete_field_list = ['_' + ofield
                                       for ofield, new_field
                                       in zip(ROV.xl_inputfile_orig_list, ROV.formatfile_field_list_to_zip)
                                       if new_field.strip().lower() == 'x']

    # rename fields that need it, otherwise leave with prefix
    ROV.inputfile_rename_fields_dict = {('_' + ofield): new_field
                                        for ofield, new_field
                                        in zip(ROV.xl_inputfile_orig_list, ROV.formatfile_field_list_to_zip)
                                        if new_field.strip().lower() != 'x'}

    # list of actual fields to be included on file ('x's removed)
    ROV.formatfile_field_list = [field for field in ROV.formatfile_field_list_to_zip if field != 'x']

    # check for duplicate field values
    duplicate_fields = [field for field, count in collections.Counter(ROV.formatfile_field_list).items() if count > 1]
    # duplicate_fields = [field for field, count in collections.Counter(ROV.formatfile_field_list).items() if count > 1]
    if duplicate_fields:
        exit_yes((f"The following fields are duplicated on the Format file.  Remove one occurance."
                  f"\n\n{', '.join(duplicate_fields)}"
                  ))

    ROV.inputfile_orig_list = [field.strip().lower() for field in ROV.xl_inputfile_orig_list]
    ROV.inputfile_type_list = [field.strip().lower() for field in ROV.xl_inputfile_type_list]

    ROV.combinefile_fields_to_delete_list = [field.strip().lower() for field in
                                             ROV.xl_combinefile_fields_to_delete_list]

    ROV.combinefile_field_list = [field
                                  for field in ROV.formatfile_field_list
                                  if field not in ROV.combinefile_fields_to_delete_list]  # dont use set to keep order

    ROV.splitfile_field_list = [field.strip().lower() for field in ROV.xl_splitfile_field_list]

    if ROV.xl_run_county_check_code_flag:  # V15.0
        fields_to_list(ROV.formatfile_field_list, "orig_county,clean_county,zip_county_list,statecounty,"
                                                  "mismatch_county,"
                                                  "numzip")

    # add 'filename' field if requested
    if ROV.xl_add_filename_column_flag:  # V15.0
        fields_to_list(ROV.formatfile_field_list, "filename")

    # Always add 'remove' field - makes reporting easier and will almost always be used
    fields_to_list(ROV.formatfile_field_list, 'remove')
    fields_to_list(ROV.formatfile_field_list, 'rownum')  # V15.0
    fields_to_list(ROV.formatfile_field_list, 'dupe_key')  # V16.0
    fields_to_list(ROV.formatfile_field_list, 'dupe_id_field')  # V16.0
    fields_to_list(ROV.formatfile_field_list, 'custom_file_field')  # V16.1

    if ROV.xl_add_random_number_column_flag:
        # add field randnum so we can sort FORMAT files by county and randnum
        fields_to_list(ROV.formatfile_field_list, "randnum")  # V15.0

    # Check that no fields are specified on output that are not on input  # V15.0 commented out -
    fields_missing_from_input = set(ROV.combinefile_field_list) - \
                                set(ROV.formatfile_field_list)
    # this doesnt need to check input to format which is what it was
    if fields_missing_from_input:
        exit_yes((f"Field(s) are specified to output on Combine file but are not present on Format.  "
                  f"\n\nMissing field(s) specified are:\n\n"
                  f"'{', '.join(fields_missing_from_input)}'"
                  f"\n\nFields on input are:\n\n"
                  f"{', '.join(ROV.formatfile_field_list)}"
                  ))

    # Check if splitfield is on field list, error if not
    if ROV.xl_splitfield != 'none' and ROV.xl_splitfield.lower() not in ROV.formatfile_field_list:
        exit_yes((f"Splitfield field '{ROV.xl_splitfield}' is missing from Format file field list.\n\n"
                  f"Available fields are:\n{', '.join(ROV.formatfile_field_list)}"
                  ))


def create_dicts():
    """ create dicts, lists, need to run """

    # create county dict returning various formats with
    # GA-CHATHAM as key: [0] is filename format; [1] print; [2] state-mixed
    # ROV.dict_statecounty_to_alt_formats = zip_file_to_county_dict('zip-codes-database-DELUXE-BUSINESS.csv',
    #                                                        'Unique_County_List.xlsx')
    ROV.dict_statecounty_to_alt_formats = counties_to_xlsx(
        os.path.join(ROV.exe_path,'zip-codes-database-DELUXE-BUSINESS.csv'),
        os.path.join(ROV.exe_path,'Unique_County_List.xlsx'))

    print('Ran Counties_to_xls')

    if ROV.xl_run_county_check_code_flag:
        with open(os.path.join(ROV.exe_path,ZIP_TO_COUNTY_LIST_FILE), "r") as dict_file:
            ROV.dict_zip_to_countylist = ast.literal_eval(dict_file.read())
            print("Imported " + ZIP_TO_COUNTY_LIST_FILE)

    concentrated_addresses_data = range_to_list(ROV.concentrated_addresses_sheet, 2,
                                                len(ROV.concentrated_addresses_sheet['A']), 1, 7)

    address_desc_list = [
        [tuple([state.strip().lower(), county.strip().lower(), city.strip().lower(), address.strip().lower()]),
         [addressdesc.strip(), removeReason.strip()]]
        for state, county, city, address, removeReason, freq, addressdesc
        in concentrated_addresses_data]  # NOTE: [0] desc, [1] remove

    ROV.dict_concentrated_addresses = dict(address_desc_list)
    # this dict can be used in remove code sheet; returns a two-tuple [0] is address description, [1] remove reason
    # code like this: if ROV.dict_concentrated_addresses.get(('NC','new hanover','wilmington', '811 martin st'),
    # 'Other') != 'Other' => set code
    # or df['remove'] = ROV.dict_concentrated_addresses.get(('NC','new hanover','wilmington', '811 martin st'), 'Other')[1]

    # Create list of county/zips to not flag as mismatched for a particular county.  Key is tuple of county and zip.
    if ROV.xl_skip_selected_zip_match_flag:
        sh_zip_skip = ROV.setup_wb[ROV.xl_skip_selected_zip_sheet]
        zip_skip_range = range_to_list(sh_zip_skip, 2, len(sh_zip_skip['A']), 1, 2)
        ROV.zipskip_list = [(cnty.lower(), zipcode) for cnty, zipcode in zip_skip_range]


def display_imported_code(sheet_name, py_file_name):
    """ reads python code contained in a workbook sheet, writes it to a py file, displays the contents on screen,
    and writes it to console for fixing program mistakes

    Parameters
    ----------
    sheet_name : the sheet in the setup.xlxs where code is found
    py_file_name : the .py text python file where the code is saved
    """
    # put the text code in py_file_name
    create_import_code_from_sheet(ROV.setup_wb, sheet_name, os.path.abspath(py_file_name))

    with open(os.path.join(ROV.exe_path,py_file_name), "r") as myfile:  # this copies in the code but does not
        # execute it
        codeobj = compile(myfile.read(), py_file_name, 'exec')
        # create a compiled object to list the lines to be executed for debugging with ine nums and comments/blanks
        # removed

        # code_lines = [(str(index + 1) + ' ' + line)
        #               for index, line in enumerate(inspect.getsourcelines(codeobj)[0])]
        if False:
            code_lines = [f"{index + 1:<4}  {line}"
                          for index, line in enumerate(inspect.getsourcelines(codeobj)[0])]
                          # for index, line in enumerate(inspect.getsourcelines(os.path.join(ROV.exe_path,codeobj))[0])]
            if len(code_lines) > 40: #  too many to display on screen, disables the 'ok' box
                skipped = len(code_lines) - 40
                code_lines = code_lines[:40]
                code_lines.append(f"<{skipped} lines not shown>")
            code_lines = ' '.join(code_lines)
            print(f"\n{py_file_name} to be run - taken from sheet: {sheet_name}")
            [print(f"{index + 1:<4}  {line}", end=' ') for index, line in enumerate(inspect.getsourcelines(codeobj)[0])]

        # Below names module based on py code name, eg 'first_code_to_be_run_module"; puts pointer to module in ROV
        # setattr allows setting object attributes using a variable
        # (i.e. this would error: xxx = 'module' =>  ROV.xxx = "some value' because xxx is not an attribute of ROV).
        # This goes one further and imports a module (.py) using importlib, and places the object into an attribute!
        # A function from this module (can even have passed parameters) will be called later in the program.
        setattr(ROV, PurePath(py_file_name).stem + '_module', importlib.import_module(PurePath(py_file_name).stem))

    exit_yes_no(f"'{py_file_name}' to be run - taken from setup sheet: '{sheet_name}'"
                f"\n\nCode and line numbers printed in python console log, too.\n\n"
                f"{code_lines}",
                f"Check '{py_file_name}' Code",
                display_exiting=False)


def loop_through_format_files(filelist_wks):
    """ loops through filelist sheet and creates format for specified x """
    cumulative_missing_counties_list = []

    for fn, format_flag, combine_flag, update_fn, update_fields, notes, custom_file_field, *_ \
            in islice(filelist_wks, 1, None):  # V16.1 read in notes and new field 'custom_file_field'
        # islice starts at row index 1 not 0 to skip header; *_ discards unused cols after custom_file_field
        if str(format_flag.value).strip().lower() == "x":

            ip = create_one_format_file(fn.value, custom_file_field.value, ROV.rawdata_path, ROV.format_path,
                                        cumulative_missing_counties_list)  # file appended to in function

            if str(update_fn.value).strip().lower() != 'none':
                merge_into_format_file(ip, str(update_fn.value), ROV.format_path)

    missing_counties_file = os.path.join(ROV.root_path, 'missing_counties.csv')
    if os.path.isfile(missing_counties_file):
        os.remove(missing_counties_file)

    if cumulative_missing_counties_list:
        pymsgbox.alert("The following counties are not in the lookup file, written to 'missing_counties.csv':\n\n" +
                       ",".join(cumulative_missing_counties_list),
                       "Alert")
        with open(os.path.join(ROV.exe_path,missing_counties_file), mode='wt', encoding='utf-8') as myfile:
            myfile.write('\n'.join(cumulative_missing_counties_list))


def check_file_headers(ws, vals):
    """
    Check list of (cell, val) tuples representing header labels in ws_to_chk and error if val not found in cell.
    eg vals = [('A1', 'use'), ('B1', 'fromFilePath'), ('C1', 'fromfilename'), ....]
    """

    def chk_header_vals(ws_to_chk, cell, val):
        """ error if val not found in wks cell. """
        if str(ws_to_chk[cell].value).strip().lower() != str(val).lower():
            exit_yes((f"Column heading '{cell}' on Setup sheet '{ws_to_chk.title}' not equal to literal '{val}'."
                      f"\n\nIt is '{str(ws_to_chk[cell].value)}'."),
                     )

    for pairs in vals:
        chk_header_vals(ws, pairs[0], pairs[1])


def formatfile_copy(ws_copy_formatfile_filelist, perform_copies=True):
    """
    Copies FORMATed files from a separate directory (likely came in a different format so needed their own read) into the current format directory
    :param ws_copy_formatfile_filelist: sheet containing rows of fromFile, toFile, processFlag
    If perform_copies = False, paths and file existence is checked but copies do not take place.
    :return: files copied to directory
    """

    def copy_file(source, destination):
        """ func to copy  with error handling pymsgbox"""
        try:
            shutil.copy(os.path.expanduser(source), destination)
            print(f"File '{source}' copied successfully to\n'{destination}'.")
            # print("File " + source + " copied successfully to\n" + destination + ".")
        except Exception:
            exit_yes((f"File not copied:\n\n {source} \n\nto\n\n {destination}"
                      ))

    # Make sure column heading/locations are as expected and nothing was moved
    check_file_headers(ws_copy_formatfile_filelist,
                       [('A1', 'use'),
                        ('B1', 'fromFilePath'),
                        ('C1', 'fromfilename'),
                        ('D1', 'tofilepath'),
                        ('E1', 'tofilename'),
                        ])

    # make sure all files and directories exist for row in ws_filelist:
    for copy_formatfile_flag, from_path, from_fn, to_path, to_fn, *_ in islice(ws_copy_formatfile_filelist, 1, None):
        # islice starts at row index 1 not 0 to skip header; *_ discards unused cols (notes, etc)
        if str(copy_formatfile_flag.value).strip().lower() == "x":
            src_path, src_file_w_path, dest_path, dest_file_w_path = \
                assign_formatcopy_vars(from_path, from_fn, to_path, to_fn)

            bad_path_exit(src_path)
            bad_file_exit(file=src_file_w_path)
            bad_path_exit(dest_path)
            # bad_file_exit(file=dest_file_w_path)  # cant check cause doesn't exist

            # if we got to here, paths and files are ok so copy
            if perform_copies:
                copy_file(src_file_w_path, dest_path)


def create_one_format_file(fn, custom_file_field, input_path, op_path, missing_counties_list_all_formatfiles):
    """
    Given a rawdata csv or xlsx with filename fn and path input_path, places a transformed csv file with the same
    name prefixed with 'FORMATTED' to the directory output_path, and returns a dataframe of the info.  The list
    missing_counties_list_all_formatfiles accumulates all counties not found in
    dictionary ROV.dict_statecounty_to_alt_formats so they can be recoded in first_code.

    Parameters
    ----------
    custom_file_field :
    """
    print("Creating formatted file for: ", fn)

    start_time = datetime.now()
    print("Start time is ", start_time.strftime("%H:%M:%S"))

    ip_file_w_path = os.path.join(input_path, fn)

    # Make sure data file exists
    bad_file_exit(ip_file_w_path, "Data file does not exist. Change filelist or "
                                  "copy into Rawdata directory.\n\n" + ip_file_w_path)

    # find the header row in the input file using an expected string and column.  error if > 30.  Defaults to search sheet0.
    header_row = find_header_row_in_file(ip_file_w_path, ROV.xl_check_header_string, ROV.xl_strcheck_header_col)

    ip = read_file_to_df(ip_file_w_path, **{'header': header_row, 'sheet_name': 0, 'nrows': ROV.xl_rows_to_read_limit,
                                            'keep_default_na': False, 'dtype': str})

    # convert column names to lower case
    ip.columns = [field.strip().lower() for field in ip.columns.values]

    # do all fields on file being read match expected?
    bad_header_fields = set(ip.columns) ^ set([field for field in ROV.xl_inputfile_orig_list])
    if bad_header_fields:
        exit_yes((f"The fields in header for '{fn}' do not match the ones in orig format files."
                  f"\nDifferences:\n\n{os.linesep.join(bad_header_fields)}"
                  # had to use os.linesep cause \n not allowed in fstrings
                  ))

    # convert columns identified as numeric in setup field list
    numeric_cols = [column_index for column_index, col_type in enumerate(ROV.xl_inputfile_type_list) if col_type != '']
    # non-blank are set to integer; code options here if move to other types
    for col in numeric_cols:
        ip[ip.columns[col]] = pd.to_numeric(ip[ip.columns[col]], errors='coerce', downcast='integer')
        # try to force specified numeric columns to integer
    print("Done reading in one file in process_format_file, csv or xlsx")

    ## rename fields with prefix of '_' so not conflicted with renamed field of same name
    print("Ready to prefix fields to with '_'")
    ip = ip.add_prefix('_')

    ip.rename(ROV.inputfile_rename_fields_dict, axis='columns', inplace=True)

    add_fields_list = [field for field in ROV.formatfile_field_list if field not in ip.columns]

    ip = ip.reindex(columns=ip.columns.tolist() + add_fields_list, fill_value='')

    ip['rownum'] = ip.index + 1  # +1 to coincide with row including header
    print("Done fill rownum")

    # Check for literal of TRUE or FALSE in fields because it gets converted to boolean and kills the program
    print("Ready to fill firstname true/false")
    replace_boolean_column_vals(ip, ['firstname', 'lastname', 'lastnametemp', 'address',
                                     'addresstemp', 'address2', 'city'])
    print("Done filling specified fields true/false with blank")

    # capture filename here so we can use in recode below
    if ROV.xl_add_filename_column_flag:
        ip['filename'] = fn

    # V16.1 add 'custom_file_field' to df to use in recoding sections if desired
    ip['custom_file_field'] = custom_file_field

    # capture random number so we can sort
    if ROV.xl_add_random_number_column_flag:
        np.random.seed(0)
        ip['randnum'] = np.random.random(ip.shape[0])

    if ROV.xl_run_first_code_flag:
        # *** Only run imported first_code and middle_code in format, not combine to keep things like remove
        # assignment, random number from being overwritten.
        # last_code can be run in combine since it's only setting remove code.
        print('Ready to run first_code')

        ROV.first_code_to_import_module.first_code_func(ip)

        # This is the function from the sheet with any parameters it needs
        print('Ran first_code')  # these prompts help if error in imported code

    if ROV.xl_run_county_check_code_flag:
        check_county_to_zips(ip, ROV.zipskip_list, ROV.dict_statecounty_to_alt_formats)
        # FIXME pymsgbox list of mismatched counties not showing

    if ROV.xl_run_middle_code_flag:
        # *** Only run imported first_code and middle_code in format, not combine to keep things like remove
        # assignment, random number from being overwritten.
        # last_code can be run in combine since it's only setting remove code.
        print('Ready to run middle_code')  # these prompts help if error in imported code

        ROV.middle_code_to_import_module.middle_code_func(ip)

        # This is the function from the sheet with any parameters it needs
        print('Ran middle_code')  # these prompts help if error in imported code

    # if requested, RUN code to remove unwanted records
    if ROV.xl_run_last_code_flag:
        # *** Only run imported first_code and middle_code in format, not combine to keep things like remove
        # assignment, random number from being overwritten.
        # last_code can be run in combine since it's only setting remove code.
        print('Ready to run last_code (remove code)')  # these prompts help if error in imported code

        ROV.last_code_to_import_module.last_code_func(ip, ROV.dict_concentrated_addresses, ROV.expectedstate)
        # This is the function from the sheet with any parameters it needs

        print('Ran last_code')  # these prompts help if error in imported code

    ip.drop(ROV.inputfile_delete_field_list, axis=1, inplace=True)

    # Sort file by randnum if flag is set
    if ROV.sort_list:  # true if not empty
        ip.sort_values(by=['remove', 'zip', 'address', 'randnum'], inplace=True)

    # Make sure county is on file, find mismatched counties and return for accumulating
    if ROV.xl_run_county_check_code_flag:
        if 'county' not in ip.columns:
            exit_yes("'CHECK COUNTY INFO FLAG' option is specified as 'TRUE' but 'county' field is not present.")

        unique_statecounties = ip.loc[ip['remove'] == '', 'statecounty'].unique()
        # unique_statecounties = ip['statecounty'].unique()

        unique_statecounties.sort()
        missing_counties_this_formatfile = [chkfield
                                            for chkfield in unique_statecounties
                                            if chkfield.upper() not in ROV.dict_statecounty_to_alt_formats]

        # [print(chkfield) for chkfield in unique_statecounties]

        missing_counties_list_all_formatfiles.extend(missing_counties_this_formatfile)

    op_file = os.path.join(op_path, "FORMATTED " + str(PurePath(fn).stem) + ".csv")
    # must check if remove flag is set otherwise field remove field is not on df

    ip.to_csv(op_file, index=False, columns=ROV.formatfile_field_list)

    if ROV.xl_run_last_code_flag:
        # write out the 'removed' file
        op_file = os.path.join(ROV.format_path, "Removed", "REMOVED " + str(PurePath(fn).stem) + ".csv")
        ip[ip["remove"] != ""].to_csv(op_file, index=False, columns=ROV.formatfile_field_list)

    print("\nFormatted file: " + fn)
    print("   Input records: " + str(len(ip)))
    if ROV.xl_run_last_code_flag:
        print("   Kept: " + str(len(ip[ip["remove"] == ""])))
        print("   Removed: " + str(len(ip[ip["remove"] != ""])))

    print("\n", ip.head(5), "\n\n")

    print("End time is ", datetime.now().strftime("%H:%M:%S"), "  Elapsed time is", str(datetime.now() - start_time),
          " (H:M:S.s)")

    # Create summary sheet of Rawdata, Formatted and Removed
    pivot_file = os.path.join(ROV.format_path, "Summary", "SUMMARY " + str(PurePath(fn).stem) + ".xlsx")
    pivot_reports(ip, pivot_file, fn, ROV.dict_concentrated_addresses)

    print("Leaving process_format_file")
    return ip


def assign_formatcopy_vars(from_path, from_fn, to_path, to_fn):
    """
    Takes to and from paths and filenames and converts to simple files w path. Paths can be code to be evaluated,
    eg referencing variable rootPathOneUp, where rooPath is the current Format dir.
    :return: list of two variable, source file and destination file or path
    """
    # from_path, from_fn, to_path, to_fn
    src_path_expr = from_path.value
    src_file = str(from_fn.value).strip()
    xl_dest_path = str(to_path.value).strip()
    xl_dest_file = str(to_fn.value).strip()

    if "," in src_path_expr or "(" in src_path_expr or " / " in src_path_expr:
        src_path = eval(src_path_expr)
    elif "/" in src_path_expr or "\\" in src_path_expr:
        src_path = src_path_expr
    else:
        src_path = eval(os.path.expanduser(src_path_expr))

    src_file_w_path = os.path.join(src_path, src_file)

    if xl_dest_path == "" or xl_dest_path.lower() == 'none':  # no "to path" specified so use default, format dir
        dest_path = os.path.join(ROV.format_path, "")
    else:
        dest_path = os.path.join(xl_dest_path, "")

    if xl_dest_file == "" or xl_dest_file.lower() == 'none':  # no "to path" specified so use default, format dir
        dest_file_w_path = os.path.join(dest_path, src_file)
    else:
        dest_file_w_path = os.path.join(dest_path, xl_dest_file)

    return [src_path, src_file_w_path, dest_path, dest_file_w_path]


def combine_formatfiles(orig_df, copy_formatfile_list_sheet):
    """
    gets files from FormatCopy sheet and puts contents into combined dataframe
    :param copy_formatfile_list_sheet: sheet containing rows of fromFile, toFile, processFlag
    :return: dataframe with data loaded
    """

    # Make sure column heading/locations are as expected and nothing was moved in Formatfile copy sheet
    check_file_headers(copy_formatfile_list_sheet,
                       [('A1', 'use'),
                        ('B1', 'fromFilePath'),
                        ('C1', 'fromfilename'),
                        ('D1', 'tofilepath')])

    # append records from FORMAT file into dataframe
    for copy_formatfile_flag, from_path, from_fn, to_path, to_fn, *_ in islice(copy_formatfile_list_sheet, 1, None):
        # Filelist is a sheet in setup opened using openpyxl
        if str(copy_formatfile_flag.value).strip().lower() == "x":
            src_path, src_file_w_path, dest_path, \
                dest_file_w_path = assign_formatcopy_vars(from_path, from_fn, to_path, to_fn)

            # make sure file and directories exist
            bad_path_exit(src_path)
            bad_file_exit(file=src_file_w_path)
            bad_path_exit(dest_path)
            bad_file_exit(file=dest_file_w_path)

            print("Ready to combine copied format file in combine_formatfiles: ", src_file_w_path)

            df_one_file = pd.read_csv(src_file_w_path, nrows=ROV.xl_rows_to_read_limit,
                                      keep_default_na=False)

            missing_split_fields = set(ROV.splitfile_field_list) - set(df_one_file.columns.tolist())
            # missing_split_fields = set(df_one_file.columns.tolist()) ^ set(orig_df.columns.tolist())
            if missing_split_fields:
                # should we exit if field lists dont match exactly?  just show mismatches? what if one has a ton?
                exit_yes((f"The input fields of the copied Formatfile: \n\n'{src_file_w_path}' "
                          f"\n\nare missing the following fields to Split:\n\n "
                          f"{', '.join(missing_split_fields)}"
                          ))
            # orig_df = orig_df.append(df_one_file, ignore_index=True)
            orig_df = pd.concat([orig_df, df_one_file], ignore_index=True)

    return orig_df


def combine_files():
    """ combines all files specified in Setup > filelist into one and runs pivots """

    df_combined = pd.DataFrame()  # empty dataframe

    for fn, format_flag, combine_flag, update_fn, update_fields_cell, *_ in islice(ROV.filelist_sheet, 1, None):
        # islice starts in row 2 (index 1)

        if str(combine_flag.value).strip().lower() == "x":
            print(f"Ready to combine file '{fn.value}'")
            fnstem = pathlib.Path(fn.value).stem
            formatfile_w_path = os.path.join(ROV.format_path, 'FORMATTED ' + fnstem + '.csv')

            bad_file_exit(formatfile_w_path)

            df_one_file = pd.read_csv(formatfile_w_path, nrows=ROV.xl_rows_to_read_limit, keep_default_na=False)

            if df_one_file.columns.tolist() != ROV.formatfile_field_list:
                exit_yes((f"The input fields of '{fn.value}' do not match specified OP fields:\n\n"
                          ",".join(df_one_file.columns.tolist())
                          ))
            # df_combined = df_combined.append(df_one_file, ignore_index=True)
            df_combined = pd.concat([df_combined, df_one_file], ignore_index=True)

    # if flag set to copy external Format files, copy them after to match fields
    if ROV.xl_copy_other_format_files_flag:
        df_combined = combine_formatfiles(df_combined, ROV.copy_formatfile_filelist_sheet)

    # delete fields specified in setup
    try:
        df_combined.drop(ROV.xl_combinefile_fields_to_delete_list, axis=1, inplace=True)
    except KeyError:
        fields_missing_from_formatfile = set(ROV.xl_combinefile_fields_to_delete_list) - set(df_combined.columns)
        exit_yes("Formatfile is missing the following fields to delete when creating Combinefile:"
                 f"\n\n{', '.join(fields_missing_from_formatfile)}"
                 )

    # V16.1 moved to main so dedup can be performed first
    # # Write out combined file
    # combine_file = os.path.join(ROV.combined_path, PurePath(ROV.xl_OPFile).stem + '.csv')
    #
    # df_combined.to_csv(combine_file, index=False)
    #
    # pd.set_option('display.max_columns', None)
    # print("\n", df_combined.head(5), "\n\n")

    return df_combined


def bek_text_box(box_title, title2, txt, buttons=None):
    """" Display text block with lines separated by \n and choice of buttons at bottom.
    :param box_title: main heading on box
    :type box_title: str
    :param title2: 2nd title above text
    :type title2: str
    :param txt: text block with lines separated by \n
    :type txt: str
    :param buttons: list of button text, defaults to ['OK', 'Exit']
    :type buttons: list of str
    :return: lower case value of selected button
    :rtype: str
    """

    if buttons is None:
        buttons = ["OK", "Exit"]

    col_factor = 3  # to scale window equally
    row_factor = 25  # to scale window equally
    max_cols = len(max(txt.split("\n"), key=len)) * col_factor
    cols = max_cols
    # v_scroll = False
    col_limit = 80 * col_factor
    col_min = 50 * col_factor
    if cols > col_limit:
        # v_scroll = True
        cols = col_limit
    elif cols < col_min:
        cols = col_min

    h_scroll = False
    row_limit = 80
    row_min = 10
    max_rows = len(txt.split("\n"))
    rows = max_rows
    if rows > row_limit:
        h_scroll = True
        rows = row_limit
    elif rows < row_min:
        rows = row_min

    layout = [
        [sg.Text(title2, font=("Arial", 18))],
        [sg.Multiline(txt, autoscroll=False, horizontal_scroll=h_scroll, expand_x=True,
                      expand_y=True, enable_events=True)],
        [sg.Button(text) for text in buttons],
    ]

    event, values = sg.Window(box_title, layout, titlebar_font=("Arial", 20), font=("Arial", 14),
                              use_custom_titlebar=True, size=(600, rows*row_factor), disable_close=True,
                              resizable=True, grab_anywhere=True).read(close=True)
    if event is not None:
        event = event.lower()
    return event


def get_dir_name(box_title, title2, initial_dir):
    """ show an "Open" dialog box and return the selected directory. Replaced askdirectory with PySimpleGUI
    :param title2:
    :type title2:
    """

    layout = [
        [sg.Text(title2, font=("Arial", 18))],
        [
         sg.Input(key="-IN-", expand_x=True),
         sg.FolderBrowse(initial_folder=os.path.expanduser(initial_dir))
         ],
        [sg.Button("Choose")],
    ]

    # event, values = sg.Window(heading_in_box, layout, size=(600, 100)).read(close=True)
    event, values = sg.Window(box_title, layout, titlebar_font=("Arial", 20), font=("Arial", 14),
                              size=(1000, 150), use_custom_titlebar=True).read(close=True)

    dir_name = values['-IN-']
    if dir_name == "":
        exit_yes("No directory name chosen")

    return dir_name


def get_file_name(box_title, title2, initial_dir):
    """ show an "Open" dialog box and return the selected file name. Replaced askopenfilename with pyeasygui
    :param title2: heading of the box
    :type title2: text next to input field
    """
    # "Select Sincere address export file 'all-parent-campaign-requests-yyyy-mm-dd.csv'"
    layout = [
        [sg.Text(title2, font=("Arial", 18))],
        [
         sg.Input(key="-IN-", expand_x=True),
         sg.FileBrowse(initial_folder=os.path.expanduser(initial_dir))
         ],
        [sg.Button("Choose")],
    ]

    # event, values = sg.Window(heading_in_box, layout, size=(600, 100)).read(close=True)
    event, values = sg.Window(box_title, layout, titlebar_font=("Arial", 20), font=("Arial", 14),
                              size=(1000, 150), use_custom_titlebar=True).read(close=True)
    # sg.Window.close()

    file_name = values['-IN-']
    if file_name == "":
        exit_yes("No file name chosen")

    return file_name


def convert_xlsx_to_csvs():

    str_xls_dir = get_dir_name("Select a DIRECTORY containing XLSX files to convert to CSVs",
                                            "XLSX Directory",
                                            INITIAL_CAMPAIGN_DIR)
    # str_xls_dir = "/Users/Denise/Library/CloudStorage/Dropbox/Postcard Files/TestInputFiles/TestCampaigns/TestXlsToCsv/Rawdata"

    str_csv_dir = get_dir_name("Select a DIRECTORY where converted CSVs will be placed",
                                            "CSV Directory",
                                            INITIAL_CAMPAIGN_DIR)
    # str_csv_dir = "/Users/Denise/Library/CloudStorage/Dropbox/Postcard Files/TestInputFiles/TestCampaigns/TestXlsToCsv/csv"

    xls_dir = pathlib.Path(str_xls_dir)
    csv_dir = pathlib.Path(str_csv_dir)

    # get list of files with xls or xlsx
    xls_files_w_path = list(xls_dir.glob("*.xls?"))

    # create list of dicts, each key a variable name corresponding to a file value,
    # eg xls_files_dict[0]['xls_name'] = 'abc.xlsx'.
    xls_files_dict = [{'xls_w_path': xls_w_path, 'xls_stem': xls_w_path.stem, 'xls_name': xls_w_path.name}
                      for xls_w_path
                      in xls_files_w_path]

    csv_files_w_path = list(csv_dir.glob("*.csv"))

    # same as done with xls above
    csv_files_dict = [{'csv_w_path': csv_w_path, 'csv_stem': csv_w_path.stem, 'csv_name': csv_w_path.name}
                      for csv_w_path
                      in csv_files_w_path]

    # set of file names (without .xls), csv below
    # xls_stem = {file_info['xls_stem'] for file_info in xls_files_dict}
    csv_stem = {file_info['csv_stem'] for file_info in csv_files_dict}

    # compare list of xls files without csv
    # xls_stem_wo_csv = xls_stem - csv_stem
    # compare list of xls files with matching csv
    # xls_stem_w_csv = xls_stem & csv_stem

    # get lists of file_info for xls with and without csvs
    xls_wo_csv = [xls_info for xls_info in xls_files_dict if xls_info['xls_stem'] not in csv_stem]
    xls_w_csv = [xls_info for xls_info in xls_files_dict if xls_info['xls_stem'] in csv_stem]

    newline = '\n'  # cause can't use \n in f-string
    # prompt stating two lists - skip with_csv, process no_csv - ok to continue?
    exit_yes_no(f"The following files WILL BE converted to csvs: {newline}" +
                newline.join([file_info['xls_name'] for file_info in xls_wo_csv]) +
                f"{newline}{newline}The following files already HAVE CSVs and will be SKIPPED: {newline}" +
                newline.join([file_info['xls_name'] for file_info in xls_w_csv]),
                "Continue with convert?"
                )

    # read each xlsx into dataframe with options and write out with same name
    for file_info in xls_wo_csv:
        print(f"Reading {file_info['xls_name']}")
        df = pd.read_excel(file_info['xls_w_path'])
        print(f"Writing {(file_info['xls_stem'] + '.csv')}")
        df.to_csv(csv_dir / (file_info['xls_stem'] + '.csv'), index=False)
        print()


def main_program():
    """ processes multiple rawdata files using parameters set in a setup spreadsheet
    """
    choice = pymsgbox.confirm("What do you want to do?",
                              'Choose Action',
                              ['Format', 'Combine', 'Split', 'XLSXs to CSVs', 'Update Zip File', 'Exit'])

    pd.set_option('display.max_columns', None)  # shows all cols instead of truncating to first and last few

    if choice == 'XLSXs to CSVs':
        convert_xlsx_to_csvs()

    elif choice == 'Update Zip File':
        import textwrap
        msg = f'''\
          'Updating Zip File' will read data from zip files purchased from zip-codes.com.

          Two files are read:
               {MAIN_ZIP_FILE}
               {MULTI_COUNTY_ZIP_FILE}

          Download the csv files.  These files are updated periodically 
          and the newest versions should be downloaded into the same
          directory as ROVCleaver.py before this update is run.
          A dictionary will be created and saved to 'ZIP_TO_COUNTY_LIST_FILE'
          and used in future runs of ROVCleaver.

          Aside: the file 'Unique_County_List.xlsx' in the ROVCleaver
          directory is updated every time ROVCleaver is run
          and can be used to get the required county names
          for those which need remapping.'''

        msg = textwrap.dedent(msg)
        exit_yes_no(msg,
                    'Update Zip Files',
                    display_exiting=False)

        # create py dict file; file defined at top of program
        # os.path.join(ROV.root_path,
        create_zip_to_county_list_dict(os.path.join(ROV.exe_path,MAIN_ZIP_FILE),
                                       os.path.join(ROV.exe_path,MULTI_COUNTY_ZIP_FILE),
                                       os.path.join(ROV.exe_path,ZIP_TO_COUNTY_LIST_FILE))
        pymsgbox.alert("Ran Zip Dict file update", "Update zip files")

    elif choice == 'Exit':
        exit()

    else:

        get_setup_file_name()  # use TKInter to get the file/path of setup in campaign

        assign_rov_variables()  # read all the variables from the setup file and put the into ROV object

        create_field_lists()  # fill array with default field names of ' ' and add fields required by options selected

        create_dicts()  # create dicts and other file set up needed to run

        if choice == 'Format':

            # check dir structure for formatted file creation. Dont create combine or split to indicate if dir is for
            # Format files copied in.
            bad_path_create(os.path.expanduser(ROV.format_path))
            bad_path_create(os.path.join(os.path.expanduser(ROV.format_path), "Summary"))
            bad_path_create(os.path.join(os.path.expanduser(ROV.format_path), "Removed"))
            bad_path_create(os.path.join(os.path.expanduser(ROV.format_path), "Duplicates"))

            # allow to exit if desired, eg flag not correct, imported code not right
            exit_for_unwanted_setup_options()

            # Loop through all files in filelist to be formatted
            loop_through_format_files(ROV.filelist_sheet)

            pymsgbox.alert("Ran format section of main", "Alert")

        elif choice == 'Combine':
            bad_path_create(os.path.expanduser(ROV.combined_path))

            if ROV.xl_copy_other_format_files_flag:
                exit_yes_no("COPY FORMAT files from other directory in combine?  OK?\n\n",
                            'COPY FORMAT FILES?',
                            display_exiting=False)

                # COPY FILES via parm 'perform_copies=True' in addition to checking files/path existence
                formatfile_copy(ROV.copy_formatfile_filelist_sheet, perform_copies=True)

            if ROV.xl_run_last_code_flag:
                display_imported_code(ROV.xl_last_code_sheet, ROV.last_code)

            df = combine_files()  # V16.1 no longer writes out file

            if ROV.xl_id_dupes != 'none':
                print(f"{ROV.xl_dupe_key_formula=}")
                df['dupe_key'] = eval(ROV.xl_dupe_key_formula)

                # sort for dupe check
                # sort_fields, ascending_vals = split_tuples(ROV.xl_dupe_key_sort_tuples)
                # df.sort_values(by=['carol', 'dupe_key'], ascending=[False,True], inplace=True)
                df.sort_values(by=ROV.dupe_key_sort_fields, ascending=ROV.dupe_key_sort_orders, inplace=True)

                identify_duplicates(df, 'dupe_key', 'dupe_id_field')

                if ROV.xl_run_last_code_flag:
                    # *** Only run imported first_code and middle_code in format, not combine to keep things like remove
                    # assignment, random number from being overwritten.
                    # last_code can be run in combine since it's only setting remove code.
                    print('Ready to run last_code (remove code)')  # these prompts help if error in imported code

                    ROV.last_code_to_import_module.last_code_func(df, ROV.dict_concentrated_addresses,
                                                                  ROV.expectedstate)
                    # This is the function from the sheet with any parameters it needs

                    print('Ran last_code')  # these prompts help if error in imported code

                # dupfile = os.path.join(os.path.join(ROV.combined_path, 'DUPLICATES in ' + PurePath(
                #     ROV.xl_OPFile).stem) + ".xlsx")
                dupfile = os.path.join(os.path.join(ROV.combined_path, 'DUPLICATES in ' + PurePath(
                    ROV.xl_OPFile).stem) + ".csv")

                print("Ready to sort by ['dupe_id_field','dupe_key']")  # to speed up copy to excel
                df.sort_values(by=['dupe_id_field', 'dupe_key'], inplace=True)

                print('Ready to copy dupes to CSV')  # these prompts help if error in imported code
                # df[df['dupe_id_field'] != 'X'].to_excel(dupfile, index=False)
                df[df['dupe_id_field'] != 'X'].to_csv(dupfile, index=False)

            test_df_clean = df[df['remove'] == '']

            if ROV.sort_list:  # true if not empty
                df.sort_values(by=ROV.sort_list, inplace=True)

            # run pivot reports on combined
            print('Ready to run pivots')  # these prompts help if error in imported code
            file = os.path.join(os.path.join(ROV.combined_path, 'Summary of ' + PurePath(ROV.xl_OPFile).stem) + ".xlsx")
            pivot_reports(df,
                          file,
                          PurePath(ROV.xl_OPFile).stem + '.csv',
                          ROV.dict_concentrated_addresses)

            # V16.1 moved writeing of combined file to after dedupe
            # Write out combined file
            combine_file = os.path.join(ROV.combined_path, PurePath(ROV.xl_OPFile).stem + '.csv')

            df.to_csv(combine_file, index=False)

            pd.set_option('display.max_columns', None)
            print("\n", df.head(5), "\n\n")

            pymsgbox.alert("Ran combine section of main", "Alert")

        elif choice == 'Split':
            # V11.1 moved to check only where needed
            bad_path_create(os.path.expanduser(ROV.split_path))  # split files
            bad_path_create(os.path.expanduser(ROV.split_path_hold))
            bad_path_create(os.path.expanduser(ROV.split_path_done))

            split_files_for_sincere(ROV.xl_sub_split_limit)
            pymsgbox.alert("Ran split section of main", "Alert")

    # else:
    #     exit_yes("Menu item chosen with no code to run!!", "** Menu item error **")

# ======  end of program.  code below runs outside of functions so variables are global to all


if __name__ == '__main__':
    @dataclass
    class ROVCleaver_Setup_Class:
        """Class for holding global setup vars in ROVCleaver"""
        exe_path: str = None
        root_path: str = None
        root_path_one_level_up: str = None
        rawdata_path: str = None
        format_path: str = None
        split_path: str = None
        split_path_hold: str = None
        split_path_done: str = None
        combined_path: str = None
        op_path: str = None

        xl_OPFile: str = None
        xl_copy_other_format_files_flag: bool = False
        xl_inputfield_list: list = None
        xl_check_header_string: str = None
        xl_strcheck_header_col: str = None
        xl_run_middle_code_flag: bool = None
        xl_run_first_code_flag: bool = None
        xl_rows_to_read_limit: int = None
        expectedstate: str = None
        xl_splitfield: str = None
        xl_str_pivot_field1: str = None
        xl_str_pivot_field1_by_cnt: bool = None
        xl_str_pivot_field1_all: bool = None
        xl_str_pivot_field2: str = None
        xl_str_pivot_field2_by_cnt: bool = None
        xl_str_pivot_field2_all: bool = None
        xl_str_pivot_field3: str = None
        xl_str_pivot_field3_by_cnt: bool = None
        xl_str_pivot_field3_all: bool = None
        xl_str_pivot_field4: str = None
        xl_str_pivot_field4_by_cnt: bool = None
        xl_str_pivot_field4_all: bool = None
        xl_str_pivot_field5: str = None
        xl_str_pivot_field5_by_cnt: bool = None
        xl_str_pivot_field5_all: bool = None
        # xl_check_statecounty: str = None
        xl_inputfile_orig_list: list = None
        xl_inputfile_renamed_list: list = None
        xl_inputfile_type_list: list = None
        xl_sub_split_limit: int = None
        xl_splitfnbase: str = None
        xl_splitfile_field_list: list = None
        xl_concentrated_addresses_file: str = None
        xl_concentrated_address_browser_prompt_freq: int = None
        xl_add_filename_column_flag: bool = None
        xl_add_random_number_column_flag: bool = None
        xl_sortchoice: int = None

        xl_id_dupes: str = None
        xl_dupe_key_formula: str = None
        xl_dupe_key_sort_tuples: str = None
        dupe_key_sort_fields: str = None
        dupe_key_sort_orders: str = None
        # dupe_keys_to_keep: list = None

        xl_run_last_code_flag: bool = None
        xl_middle_code_sheet: str = None
        xl_first_code_sheet: str = None
        xl_last_code_sheet: str = None
        xl_run_county_check_code_flag: bool = None
        xl_skip_selected_zip_match_flag: bool = None
        xl_skip_selected_zip_sheet: str = None
        xl_combinefile_fields_to_delete_list: list = None
        xl_run_merge_data_flag: bool = None
        xl_orig_file_key_formula: str = None
        xl_update_file_key_formula: str = None
        xl_update_field_list: list = None

        # xl_formatfile_field_list_to_zip: list = None
        # xl_formatfile_field_list: list = None
        # xl_combinefile_field_list: list = None

        lst_pivot_field1: list = None
        lst_pivot_field2: list = None

        # fields_on_inputfile_list: list = None
        formatfile_field_list: list = None
        # type_fields_on_inputfile: list = None

        inputfile_delete_field_list: list = None
        inputfile_rename_fields_dict: dict = None

        setup_file_name: str = None
        setup_wb: any = None
        setup_sheet: any = None
        filelist_sheet: any = None
        copy_formatfile_filelist_sheet: any = None
        sort_list: int = None

        concentrated_addresses_sheet: any = None
        concentrated_addresses_wb: any = None
        dict_concentrated_addresses: dict = None
        dict_zip_to_countylist: dict = None
        zipskip_list: list = None
        # skipzip_list: dict = None
        dict_statecounty_to_alt_formats: dict = None
        first_code: any = None
        middle_code: any = None
        last_code: any = None

        first_code_to_import_module: any = None
        middle_code_to_import_module: any = None
        last_code_to_import_module: any = None


    ROV = ROVCleaver_Setup_Class()  # ROV is only instance ROVCleaver_Setup_Class used to hold 'global' setup variables

    main_program()
