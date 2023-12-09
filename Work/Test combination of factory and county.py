""" given lists of possible group elements, produce all grou combinations """

import itertools
import pathlib

import pandas as pd
from openpyxl import Workbook
from loguru import logger
import PySimpleGUI as sg
from pathlib import Path
from openpyxl import load_workbook

# a = [[1, 2, 3], [4, 5, 6], [7, 8, 9, 10]]
# groups = list(itertools.product(*a))

def get_file_name(box_title, title2, initial_dir):
    """ show an "Open" dialog box and return the selected file name. Replaced askopenfilename with pyeasygui
    :param title2: heading of the box
    :type title2: text next to input field
    """
    logger.debug('here')
    # "Select Sincere address export file 'all-parent-campaign-requests-yyyy-mm-dd.csv'"
    layout = [
        [sg.Text(title2, font=("Arial", 18))],
        [
         sg.Input(key="-IN-", expand_x=True),
         sg.FileBrowse(initial_folder=Path(initial_dir).expanduser())
         ],
        [sg.Button("Choose")],
    ]

    # event, values = sg.Window(heading_in_box, layout, size=(600, 100)).read(close=True)
    event, values = sg.Window(box_title, layout, titlebar_font=("Arial", 20), font=("Arial", 14),
                              size=(1000, 150), use_custom_titlebar=True).read(close=True)
    # sg.Window.close()

    file_name = values['-IN-']
    if file_name == "":
        exit("No file name chosen")

    return file_name

def read_sheet_w_header(ws):
    ws_info = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        ws_info.append(row)
    ws_header = ws_info.pop(0)
    return ws_header, ws_info


def read_sheet_cols_into_lists(ws):
    ws_info = []
    ws_header = []
    for col_tuple in ws.iter_cols(values_only=True):
        col = list(col_tuple)  # cause tuples are immutable
        col = [x for x in col if x is not None]
        ws_header.append(col.pop(0))
        ws_info.append(col)
    # ws_header = ws_info.pop(0)
    return ws_header, ws_info


# age = ['Old', 'Young']
# ethnicity = ['Black', 'Asian', 'Other']
# county = ['Kings', 'Queens', 'Nassau', 'Suffolk']

# county_header = ['{kings}','{queens}','{nassau}','{suffolk}',]
# county_header = ['{county}','{phone}','{url}',]
#
# county_info = [
#                 ['Kings', '111-111-1111', 'nyvoter.net\kings'],
#                 ['Queens', '222-222-2222', 'nyvoter.net\queens'],
#                 ['Nassau', '333-333-3333', r'nyvoter.net\nassau'],
#                 ['Suffolk', '444-444-4444', 'nyvoter.net\suffolk'],
#               ]

if False:
    wb_str = get_file_name("Select Input Workbook",
                           f"Select factory and county input workbook.",
                           "")
else:
    wb_str=  pathlib.Path("/Users/Denise/Library/CloudStorage/Dropbox/Postcard " \
            "Files/PythonProgs/ROVCleaver_on_Dropbox/Work/factory_county_input.xlsx")

wb_in = load_workbook(filename=wb_str)
factory_sheet = wb_in["factory"]
county_sheet = wb_in["county"]
factory_parts = wb_in["factory_parts"]

col_headers, col_lists = read_sheet_cols_into_lists(factory_parts)

factory_combos = list(itertools.product(*col_lists))

# factory_combos2 = [["-".join(tuple1)] + list(tuple1) for tuple1 in factory_combos]
# factory_combos2 = [["-".join(tuple1)], ["-".join(list(tuple1)[:-1])] + list(tuple1) for tuple1 in factory_combos]
factory_combos2 = [["-".join(tuple1)] + ["-".join(tuple1[:-1])] + list(tuple1) for tuple1 in factory_combos]
factory_combos2.insert(0, ["campaign", "factory"] + col_headers)

# clear
# write
# try:
#     del wb_in["factory_combined"]
# except:
#     pass

# if 'factory_combined' in wb_in.sheetnames:
#     print(f"Sheet 'factory_combined' exists in '{wb_str.name}'.  Rename, delete and start again.")
#     # prompt for overwrite and delete i Y
#     exit()

wb_in.create_sheet(title="factory_combined")
factory_combined = wb_in["factory_combined"]
for row in factory_combos2:
    factory_combined.append(row)
# wb_in.save(wb_str)
wb_in.save(wb_str)


factory_header, factory_info = read_sheet_w_header(factory_sheet)
county_header, county_info = read_sheet_w_header(county_sheet)

# factory_header = ['{factory}', '{salutation}', '{issue1}','{issue2}',]

# factory_info = [
#                 ['Black-Young', 'hello young black person', 'womens productive rights', 'school debt'],
#                 ['Black-Old', 'hello wise black person', 'social security', 'voting'],
#                 ['Asian-Young', 'hello young asian person', 'womens productive rights', 'school debt'],
#                 ['Asian-Old', 'hello wise asian person', 'social security', 'voting'],
#               ]


factory_county_list = list(itertools.product(factory_info, county_info))
# factory_county_dict = list(itertools.product(factory_dicts, county_dicts))

factory_county_list2 = [tuple1 + tuple2 for tuple1, tuple2 in factory_county_list]
header2 = factory_header + county_header
factory_county_list2.insert(0, header2)

wb_out = Workbook()  # creates a workbook object.
ws = wb_out.active  # creates a worksheet object.
ws.title = "FacCount"
for row in factory_county_list2:
    ws.append(row)  # adds values to cells, each list is a new row.

wb_out.save('factory_county_output.xlsx')

a=1

group_fields = ['ethnicity', 'age']
group_lists = [*group_fields]

# groups = list(itertools.product(ethnicity, age, county))
factories = list(itertools.product(*group_fields))

# all_data = [ [ethnicity, age, ethnicity + '-' + age] for ethnicity, age, county in factories]
all_data = [ [ethnicity, age, ethnicity + '-' + age] for ethnicity, age, county in factories]

# df = pd.DataFrame(all_data)
# writer = pd.ExcelWriter('all_groups.xlsx')  #, engine='xlsxwriter')
# df.to_excel(writer, sheet_name='welcome', index=False)
# writer.save()

wb_in = Workbook()  # creates a workbook object.
ws = wb_in.active  # creates a worksheet object.
ws.append(['ethnicity', 'age', 'county', 'factory', 'campaign'])
for row in all_data:
    ws.append(row)  # adds values to cells, each list is a new row.

wb_in.save('all_groups.xlsx')

a=1
