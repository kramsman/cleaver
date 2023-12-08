""" given lists of possible group elements, produce all grou combinations """

import itertools
import pandas as pd
from openpyxl import Workbook
from loguru import logger
import PySimpleGUI as sg
from pathlib import Path
from openpyxl import load_workbook

# a = [[1, 2, 3], [4, 5, 6], [7, 8, 9, 10]]
# groups = list(itertools.product(*a))

def get_file_name(box_title, title2, initial_dir):
    """ show an "Open" dialog box and return the selected file name. Replaced askopenfilename with pyeasygui
    :param title2: heading of the box
    :type title2: text next to input field
    """
    logger.debug('here')
    # "Select Sincere address export file 'all-parent-campaign-requests-yyyy-mm-dd.csv'"
    layout = [
        [sg.Text(title2, font=("Arial", 18))],
        [
         sg.Input(key="-IN-", expand_x=True),
         sg.FileBrowse(initial_folder=Path(initial_dir).expanduser())
         ],
        [sg.Button("Choose")],
    ]

    # event, values = sg.Window(heading_in_box, layout, size=(600, 100)).read(close=True)
    event, values = sg.Window(box_title, layout, titlebar_font=("Arial", 20), font=("Arial", 14),
                              size=(1000, 150), use_custom_titlebar=True).read(close=True)
    # sg.Window.close()

    file_name = values['-IN-']
    if file_name == "":
        exit("No file name chosen")

    return file_name



# age = ['Old', 'Young']
# ethnicity = ['Black', 'Asian', 'Other']
# county = ['Kings', 'Queens', 'Nassau', 'Suffolk']

# county_header = ['{kings}','{queens}','{nassau}','{suffolk}',]
# county_header = ['{county}','{phone}','{url}',]
#
# county_info = [
#                 ['Kings', '111-111-1111', 'nyvoter.net\kings'],
#                 ['Queens', '222-222-2222', 'nyvoter.net\queens'],
#                 ['Nassau', '333-333-3333', r'nyvoter.net\nassau'],
#                 ['Suffolk', '444-444-4444', 'nyvoter.net\suffolk'],
#               ]

wb_str = get_file_name("Select Input Workbook",
                       f"Select factory and county input workbook.",
                       "")
wb = load_workbook(filename=wb_str)
factory_sheet = wb["Factory"]
county_sheet = wb["County"]

# for fn, format_flag, combine_flag, update_fn, update_fields, pull_group, custom_field, notes, *_ \
#         in filelist_wks.iter_rows(min_row=2, values_only=True):
# for row in factory_sheet.iter_rows(min_row=1, values_only=True):
def read_sheet_w_header(ws):
    ws_info = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        ws_info.append(row)
    ws_header = ws_info.pop(0)
    return ws_header, ws_info

factory_header, factory_info = read_sheet_w_header(factory_sheet)
county_header, county_info = read_sheet_w_header(county_sheet)

# factory_header = ['{factory}', '{salutation}', '{issue1}','{issue2}',]

# factory_info = [
#                 ['Black-Young', 'hello young black person', 'womens productive rights', 'school debt'],
#                 ['Black-Old', 'hello wise black person', 'social security', 'voting'],
#                 ['Asian-Young', 'hello young asian person', 'womens productive rights', 'school debt'],
#                 ['Asian-Old', 'hello wise asian person', 'social security', 'voting'],
#               ]

# county_info2 = example = [[x.replace('\r\n','') for x in l] for l in county_info]  # what to do with '\n' in
# '\naussa'?

#     county_dicts = {}
#     for row in county_info:
#         temp_dict = dict(zip(county_header, row))
#         county_dicts[temp_dict['{county}']] = temp_dict

def dictify(l_of_l: list[list], l_header: list[str], key_field: str) -> dict[dict]:
    """ returns of dictionary of dictionaries given a list of lists, headers for the fields in the lists,
    and the value to be used as the key into the final dictionary

    Parameters
    ----------
    l_of_l : list of lists data, eg [['Kings:Young', 'nyvoter.net\kings'], ['Queens', 'nyvoter.net\queens']]
    l_header : field names for each, eg ['county','url',]
    key_field : field from headers to be used as the key into the sub lists, eg 'county'
    """

    dict_of_dicts = {}
    for row in l_of_l:
        temp_dict = dict(zip(l_header, row))
        dict_of_dicts[temp_dict[key_field]] = temp_dict

    return dict_of_dicts

# county_dicts = dictify(county_info, county_header,'{county}')
# factory_dicts = dictify(factory_info, factory_header,'{factory}')


factory_county_list = list(itertools.product(factory_info, county_info))
# factory_county_dict = list(itertools.product(factory_dicts, county_dicts))

factory_county_list2 = [tuple1 + tuple2 for tuple1, tuple2 in factory_county_list]
header2 = factory_header + county_header
factory_county_list2.insert(0,header2)

wb = Workbook()  # creates a workbook object.
ws = wb.active  # creates a worksheet object.
ws.title = "FacCount"
for row in factory_county_list2:
    ws.append(row)  # adds values to cells, each list is a new row.

wb.save('factory_county_output.xlsx')

a=1

# for row in county_info
#   for field in county_header
#       output field[x]:row[x]

# https://www.geeksforgeeks.org/python-convert-two-lists-into-a-dictionary/
# res = {test_keys[i]: test_values[i] for i in range(len(test_keys))}

# tuples = [(key, value) for i, (key, value) in enumerate(zip(test_keys, test_values
# res = dict(tuples)

# res = dict(zip(test_keys, test_values))

group_fields = ['ethnicity', 'age']
group_lists = [*group_fields]

# groups = list(itertools.product(ethnicity, age, county))
factories = list(itertools.product(*group_fields))

# all_data = [ [ethnicity, age, ethnicity + '-' + age] for ethnicity, age, county in factories]
all_data = [ [ethnicity, age, ethnicity + '-' + age] for ethnicity, age, county in factories]

# df = pd.DataFrame(all_data)
# writer = pd.ExcelWriter('all_groups.xlsx')  #, engine='xlsxwriter')
# df.to_excel(writer, sheet_name='welcome', index=False)
# writer.save()

wb = Workbook()  # creates a workbook object.
ws = wb.active  # creates a worksheet object.
ws.append(['ethnicity', 'age', 'county', 'factory', 'campaign'])
for row in all_data:
    ws.append(row)  # adds values to cells, each list is a new row.

wb.save('all_groups.xlsx')

a=1
