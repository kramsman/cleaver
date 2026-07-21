""" cleaver UniCode is a complete redo under the hood!  Taken from working V16.3
    Setup file drastically changed - variable definitions and documentation are contained in setup - no hardcoded rows!
    ROV. structure replace with dictionary for ease in adding and removing entries
    Allows moving and sorting rows.
    Some repetitious code - like pivot tables 1-5 - are looped.

If you get an error "Gitupdater not found:
  1. go into terminal
  2. copy and enter: source .venv/bin/activate
  3. then enter: uv pip install git+https://github.com/kramsman/gitupdater.git

test push
"""
# 12/6/23 Add ability to split by a group variable in addition to county for multi scripts.  Try git branch with
# group_split
#   changed splitfield to group_vars
#   added factory_var
# 12/9/23 Hardcode path to zip and concentration files so cleaver_prod and _test both find them

#  to remove unused functions used vulture.  In Pycharm terminal: vulture 'xxx.py'

# SKIP-put filetype in to select file and specify xlsx for Setup. CAN NOT WORK IN MAC VERSION OF TKINTER.
#   https://stackoverflow.com/questions/57443004/pysimplegui-file-browser-specific-file-type

# TODO add in "compare_csv_columns.py".  Needs cleaning up of import, logging, select directory.

# TODO redo formatcopy using pathlib / remove os.path.join.  Create paths from sheet using
#   /Users/Denise/Library/CloudStorage/Dropbox/Postcard Files/PythonPrograms/Development/cleaver/Work/formatcopy test paths.py

# TODO can we skip padding variable lists?  zip will work on shortest one
# FIXME make sure Format merge works - which fields missing are ok?
# TODO put .py code created from first, middle, last sheets in root dir (with setup) rather than exe dir (cleaver)
#   so different runs of Cleaver don't collide.  Problem is compiled object goes to exe so is not found when moved.
# TODO do we need to be able to change order of grouping variables or can campaign always be added to front? Or try
#    to add but skip if already in list?
# TODO remove concentrate property web browser and replace with prompt if props present with no description
# TODO XXXX add property concentration field automatically instead of adding to setup field list.  Is there a flag?

# TODO make pull_group mandatory. Can it be backward compatible on filelist page?
# TODO include pull_group on FORMAT_COPY where copied in
# TODO include in split file list name - maybe combine and format filenames?
# TODO Include pull_group number in remove reason, eg "Pull_group of 2 less than current of 3"
# TODO Add note 'THIS IS A MAIN" to help with formatcopy

# run gitupdater to make sure uvbekutils and bekgoogle utility libraries are updated  NOT NEEDED PER CLAUDE CHANGE
# import sys
# import os
# sys.path.append(os.path.expanduser("~/Dropbox/Postcard Files/"))
# if False:
#    import gitupdater

from cleaver import compare_csv_columns
import ast
import collections
import datetime
import importlib
import itertools
import math  # for ceil function
import os
import re
import shutil  # to copy file from other dirs
import webbrowser
from datetime import datetime
from itertools import islice  # to skip 1st row of iterated spreadsheet
from pathlib import Path
from typing import Any, Callable, Optional, TextIO, Union
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell

import numpy as np
import pandas as pd
# os.environ["APPDATA"] = ""
# from pandasgui import show
# from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font
# import PySimpleGUI as sg
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.utils.cell import column_index_from_string
import json
from loguru import logger

from uvbekutils import exe_path
from uvbekutils import select_file
from uvbekutils import setup_loguru
from uvbekutils import is_number
from uvbekutils import exit_yes
from uvbekutils import exit_yes_no
from uvbekutils import clean_field
from uvbekutils import autosize_xls_cols
from uvbekutils import bad_file_exit
from uvbekutils import bad_path_exit
from uvbekutils import bad_path_create
# from uvbekutils import calling_func
from uvbekutils import find_header_row_in_file
from uvbekutils import read_file_to_df
from uvbekutils import check_ws_headers
from uvbekutils import convert_bool
from uvbekutils import load_workbook_w_filepath, wb_name
from uvbekutils import confirm, alert

USE_HARDCODED_SETUP = False
HARDCODED_SETUP_FILE = Path("~/Dropbox/Postcard Files/InputFiles/Campaigns/VA General 7-2025 test 500 black voter group/ROVCleaver VA General 7-2025 UniversalSetup.xlsx"
                            "").expanduser()


setup_loguru("INFO", "DEBUG")
# log_level = "DEBUG"  # used for log file; screen set to INFO. TRACE, DEBUG, INFO, WARNING, ERROR

INITIAL_CAMPAIGN_DIR = Path("~/Dropbox/Postcard Files/InputFiles/Campaigns").expanduser()
# INITIAL_CAMPAIGN_DIR = Path("~/Dropbox/Postcard Files/TestInputFiles/").expanduser()

# MAIN_ZIP_FILE = 'zip-codes-database-DELUXE-BUSINESS.csv'
# MAIN_ZIP_FILE = Path("~/Dropbox/Postcard Files/"
#                              "PythonPrograms/ROVCleaver_Production/zip-codes-database-DELUXE-BUSINESS.csv").expanduser()
#'/Users/Denise/Library/CloudStorage/Dropbox/Postcard Files/ROVPrograms/ROVCleaver_Production/zip-codes-database-DELUXE-BUSINESS.csv'
# MAIN_ZIP_FILE = Path('~/Library/CloudStorage/Dropbox/Postcard '
#                      'Files/ROVPrograms/ROVCleaver_Production/zip-codes-database-DELUXE-BUSINESS.csv').expanduser()
MAIN_ZIP_FILE = Path('~/Dropbox/Postcard Files/ROVPrograms/ROVCleaver_Production/zip-codes-database-DELUXE-BUSINESS.csv').expanduser()
# MULTI_COUNTY_ZIP_FILE = Path("~/Dropbox/Postcard Files/PythonPrograms/ROVCleaver_Production/zip-codes-database-MULTI-COUNTY.csv").expanduser()
# MULTI_COUNTY_ZIP_FILE = Path("~/Library/CloudStorage/Dropbox/Postcard Files/ROVPrograms/ROVCleaver_Production/zip-codes-database-MULTI-COUNTY.csv").expanduser()
MULTI_COUNTY_ZIP_FILE = Path("~/Dropbox/Postcard Files/ROVPrograms/ROVCleaver_Production/zip-codes-database-MULTI-COUNTY.csv").expanduser()
# /Dropbox/Postcard Files/PythonPrograms/ROVCleaver_Production/zip-codes-database-MULTI-COUNTY.csv
# ZIP_TO_COUNTY_LIST_FILE = 'Zip_To_County_List_dict.py'  # file where the numeric zip to county list is stored (ie
# file where the numeric zip to county list is stored (ie 1011: ['hampden', 'hampshire'])
ZIP_TO_COUNTY_LIST_FILE = Path("~/Dropbox/Postcard Files/"
                                       "ROVPrograms/ROVCleaver_Production/Zip_To_County_List_dict.py").expanduser()

# Sentinel statecounty that all out-of-state counties are rolled up into so they don't
# clutter reports.  It is not a real key in the zip lookup, so it must be excluded from
# the missing-county check or it always shows up as a false mismatch.
ROLLUP_STATECOUNTY = "All_Counties"


def home_tilde(path: Union[str, os.PathLike]) -> str:
    """Return ``path`` with the user's home directory replaced by ``~``.

    Inverse of ``Path.expanduser()``: paths built with ``expanduser()`` contain the
    current login (e.g. ``/Users/Denise/...``), which is noise in user-facing messages.
    The returned string round-trips back through ``expanduser()``.  Paths that are not
    under the home directory are returned unchanged.

    Args:
        path (Union[str, os.PathLike]): The (usually absolute) path to shorten.

    Returns:
        str: The path with a leading ``~/`` in place of the home directory, or the
            original path string if it is not inside the home directory.
    """
    path = Path(path)
    try:
        return "~/" + path.relative_to(Path.home()).as_posix()
    except ValueError:
        return str(path)

# PROP_CONCENTRATION = 50
PROP_CONCENTRATION = 15
ZIP_CONCENTRATION = 10

ROV_SETUP = {}
# Following for reading variables from setup sheet
REMOVE_XL_FROM_SETUP = False  # remove or keep the temporary 'wor' variables prefixed with 'XL_'
# column in setup where fields are defined
FIELD_DEF_COL_ALPHA = 'E'  # field_keep is assumed one column left
FIELD_DEF_COL_NUMERIC = column_index_from_string(coordinate_from_string(FIELD_DEF_COL_ALPHA + '1')[0]) - 1  # -1 to
# zero index
# cols in setup where data starts (normally visible section)
DATA_STARTS_COL_ALPHA = 'G'  # field_keep is assumed one column left
DATA_STARTS_COL_NUMERIC = column_index_from_string(coordinate_from_string(DATA_STARTS_COL_ALPHA + '1')[0]) - 1  # -1 to


# def check_ws_headers(ws, vals):
#     """
#     Check list of (cell, val) tuples representing header labels in ws_to_chk and error if val not found in cell.
#     eg vals = [('A1', 'use'), ('B1', 'fromFilePath'), ('C1', 'fromfilename'), ....]
#     """
#
#     def chk_header_vals(ws_to_chk, cell, val):
#         """ error if val not found in wks cell. """
#         if str(ws_to_chk[cell].value).strip().lower() != str(val).lower():
#             exit_yes((f"Column heading '{cell}' on sheet '{ws_to_chk.title}' in workbook '{wb_name(ws.parent)}' "
#                       f"not equal to literal '{val}'."
#                       f"\n\nIt is '{str(ws_to_chk[cell].value)}'."),
#                      )
#
#     print(f"'{wb_name(ROV_SETUP['setup_sheet'].parent)=}'")
#
#     for pairs in vals:
#         chk_header_vals(ws, pairs[0], pairs[1])

def pad_list(my_list: list, to_len: int, pad_val: Any = "") -> list:
    """Pad a list with a value to reach a specified length.

    Args:
        my_list (list): The list to pad.
        to_len (int): The target length of the padded list.
        pad_val (Any): The value used for padding. Defaults to "".

    Returns:
        list: A new list of length ``to_len`` padded with ``pad_val``.
    """
    list_len = len(my_list)
    elem_needed = to_len - list_len
    padded_list = my_list + [pad_val] * elem_needed
    return padded_list


def row_to_list(row: tuple) -> list:
    """Return cell values from a worksheet row up to the last non-None cell.

    Args:
        row (tuple): A tuple of openpyxl Cell objects representing one worksheet row.

    Returns:
        list: Cell values from column 1 through the last column containing a non-None value.
    """
    for index, cell in enumerate(reversed(row)):
        if cell.value is not None:
            break
    row_list = [val.value for val in islice(row, 0, cell.col_idx)]
    return row_list


def range_to_list(ws: Worksheet, start_row: int, end_row: int,
                  start_col: int, end_col: int) -> Union[list[str], list[list[str]]]:
    """Convert a rectangular range of worksheet cells to a list of string values.

    Non-None cell values are lowercased and stripped; None values become "".
    When ``start_row == end_row``, returns a flat ``list[str]`` instead of
    a list of lists.

    Args:
        ws (Worksheet): The openpyxl worksheet to read from.
        start_row (int): First row number (1-indexed).
        end_row (int): Last row number (1-indexed, inclusive).
        start_col (int): First column number (1-indexed).
        end_col (int): Last column number (1-indexed, inclusive).

    Returns:
        Union[list[str], list[list[str]]]: Flat list when a single row is
        requested; list of row lists when multiple rows are requested.
    """

    # tried using openpyl's cells = setup['A18': 'E19'] but parsing letter cols would be ugly so iterate rows.
    # openpyxl does not offer cells((r,c),(r,c)) using numerics
    final_list = []
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col,
                            max_col=end_col):
        row_list = []
        for cell in row:
            if cell.value is not None:
                row_list.append(str(cell.value).lower().strip())
            else:
                row_list.append("")
        final_list.append(row_list)
    if start_row == end_row:  # produce a one dimensional list instead of a one element list in a list
        final_list = final_list[0]
    return final_list


def replace_boolean_column_vals(df: pd.DataFrame, field_list: list[str]) -> None:
    """Replace literal 'TRUE' or 'FALSE' string values in specified columns with an empty string.

    Prevents pandas from converting these strings to booleans, which would
    cause downstream errors.

    Args:
        df (pd.DataFrame): The DataFrame to modify in place.
        field_list (list[str]): Column names to check for boolean literals.
    """

    for field in field_list:
        if field in df.columns:
            df.loc[(df[field].str.upper() == "TRUE") | (df[field].str.upper() == "FALSE"), field] = ''


def identify_duplicates(df: pd.DataFrame, key: str, dupe_id_field: str) -> None:
    """Add a duplicate-identifier field to a DataFrame.

    Each row is labelled with one of:
        ``-`` not a duplicate, ``F`` first occurrence, ``L`` last occurrence,
        ``D`` intermediate duplicate, ``O`` other.

    Args:
        df (pd.DataFrame): The DataFrame to modify in place.
        key (str): Column name used to detect duplicates.
        dupe_id_field (str): Name of the new column to fill with identifiers.

    Raises:
        Exception: If the length of the identifier list does not match the DataFrame.
    """

    logger.info("Identifying duplicates")

    any_dupe_bool = df.duplicated([key], keep=False)  # id all dupes using true/false
    # 'first' returns true for non-dupes as well as 'real' firsts, so must 'and' with all dupes
    first_bool = ~df.duplicated([key], keep='first') & any_dupe_bool  # id firsts where first AND dupe as true
    last_bool = ~df.duplicated([key], keep='last') & any_dupe_bool  # id lasts where last AND dupe as true

    # set different numeric values for first and last so we can identify them as well as other dupes later
    first_numeric = np.where(first_bool, 2, 0)
    last_numeric = np.where(last_bool, 1, 0)
    any_dupe_numeric = np.where(any_dupe_bool, 1, 0)
    dupe_numeric = any_dupe_numeric + first_numeric + last_numeric  # sets each to different value
    dupe_alpha = [{0: '-', 1: 'D', 2: 'L', 3: 'F', 4: 'O'}[element] for element in dupe_numeric]  #
    # note that setting non-dupe, 0, to blank above removes them from list so can not be merged into df

    if len(dupe_alpha) != len(df):
        logger.info(f"*** assigning duplicate identifier error:  "
              f"Orig dataframe has {len(df)} rows but merged list has {len(dupe_alpha)}")
        raise Exception
    df[dupe_id_field] = dupe_alpha

    # show(df)  # pandasgui

    return


def merge_into_format_file(orig_df: pd.DataFrame, update_file: str,
                           cur_path: Union[str, Path]) -> None:
    """Merge fields from an update file into the original format DataFrame.

    Reads ``update_file`` from ``cur_path``, aligns on configured key fields,
    and left-joins updated columns into ``orig_df``.

    Args:
        orig_df (pd.DataFrame): The master DataFrame to update in place.
        update_file (str): File name of the update source (CSV or XLSX).
        cur_path (Union[str, Path]): Directory containing ``update_file``.
    """

    # TODO:  THIS NEEDS TO BE RE-WRITTEN AND CHECKED.  Copied and some refactoring but no logi or working checked.

    logger.info("Merging into format file")

    update_file_w_path = os.path.join(cur_path, update_file)
    # NOTE: update file is expected to have headers row 1 unlike raw data which may have title lines at top.

    update_df = read_file_to_df(update_file_w_path, **{'sheet_name': 0, 'nrows': ROV_SETUP['rows_to_read_limit'],
                                                       'keep_default_na': False})

    update_df.columns = [x.strip().lower() for x in list(update_df.columns)]
    # rename df fields as lower() because col matching is case-sensitive

    if set(ROV_SETUP['update_field_list']).issubset(set(update_df.columns)):  # update field list all contained in df fields
        update_df = update_df[update_df.columns[update_df.columns.isin(ROV_SETUP['update_field_list'])]]
    else:  # some specified update fields not on df
        extraelems = set(ROV_SETUP['update_field_list']) - set(update_df.columns)
        exit_yes((f"Following field(s) in update list not in input.\n\n"
                  f"{os.linesep.join(extraelems)}"
                  ))

    # add the key fields specified in setup to both files
    # update_df['update_key'] = address + city
    update_df['update_key'] = eval(
        ROV_SETUP['update_file_key_formula'])  # TODO: try sorting by keys to speed up dropping duplicates
    orig_df['master_key'] = eval(ROV_SETUP['orig_file_key_formula'])

    # BEK2/22 hardcode field input as variable updateField to 'updateField'
    # update_df.rename(columns={updateField : 'update_field'}, inplace=True)  # why specify field name to create in setup then rename to hardcoded 'update_field'?

    # update_df.columns += '_updt' # add suffix to all columns names to keep separate from master. Prefix is df.columns = 'prefix_' + df.columns
    update_df.columns = [str(col) + '_updt' if col not in ['update_key'] else 'update_key' for col in
                         update_df.columns]  # add suffix to all columns names except 'update_key' to keep separate from master.

    # create df after dropping  duplicates based on key
    updt_w_duplicates_removed = update_df.drop_duplicates(subset=['update_key'])
    update_count_before_dups_removed = len(update_df)
    update_count_after_dups_removed = len(updt_w_duplicates_removed)
    num_dups_in_updt = update_count_before_dups_removed - update_count_after_dups_removed

    # Output a df of duplicate key values
    logger.debug("list of counts of updt duplicate key values in ", update_file)
    # dups = update_df.duplicated(subset=['update_key'], keep='first')  # not sure what this shows.  needed?
    dups_in_update = update_df[update_df.duplicated(subset=['update_key'],
                                                    keep=False)]  # Keep = false will show all dups; only first is kept above
    dups_in_update = dups_in_update.sort_values("update_key")
    logger.debug("list of updt duplicate key values (showing the first) in ", update_file)
    logger.debug(dups_in_update)  # does this work for df or do we need str() or something to avoid object?
    # write all duplicates to a file so we can take a look if desired
    file_of_dups = ROV_SETUP['format_path'] / "Duplicates" / \
                   ("UPDATE DUPLICATES " + str(Path(update_file).stem) + ".xlsx")
    dups_in_update.to_excel(file_of_dups, index=False)

    # must check for dups because key might not be unique (eg using truncated name)
    if num_dups_in_updt > 0:  # calculate above by comparing before and after de-dup
        exit_yes_no("Continue?  Dups in update will be removed.\n\n\nUpdate file\n" +
                    orig_df + "\n contains " + str(num_dups_in_updt) + " duplicates.",
                    "CHECK FOR DUPLICATE KEYS IN UPDATE",
                    display_exiting=False)

        logger.debug('Count before de-dup in update file', len(update_df))
        logger.debug('Count after de-dup in update file', len(updt_w_duplicates_removed))
        update_df = updt_w_duplicates_removed
        # print('New update_df without dupes - should be old updt_w_duplicates_removed ',update_df.shape[0])
        logger.debug('New update_df without dupes - should be old updt_w_duplicates_removed ', len(update_df))

    # update_df = update_df[['update_key', 'update_field']]  # keep only the key to prevent dup fields from being renamed in df with suffix (_x)

    master_w_dupkey_removed = orig_df.drop_duplicates(subset=['master_key'])  # new df with all duplicate master_key

    num_of_master_dups = len(orig_df) - len(master_w_dupkey_removed)

    # prompt if wanting to update if dupe on master present
    if num_of_master_dups > 0:
        exit_yes_no("Continue with dups?\n\n\nMaster file\n" + orig_df + "\ncontains " +
                    str(num_of_master_dups) + " duplicates.",
                    "CHECK FOR DUPLICATE KEYS IN MASTER",
                    display_exiting=False)

        # write dupes in master to file for reviewing
        file_of_dups = ROV_SETUP['format_path'] / "Duplicates" / \
                       ("MASTER DUPLICATES " + str(Path(orig_df).stem) + ".xlsx")
        # writer = pd.ExcelWriter(file_of_dups)
        orig_dups = orig_df[orig_df.duplicated(subset=['master_key'], keep=False)]
        orig_dups = orig_dups.sort_values("master_key")
        orig_dups.to_excel(file_of_dups, index=False)

    #### Do the actual update!
    # In perfect world could use ", validate='1:1'" in merge to ensure integrity, but input data has dups,
    # so code was added to point out how many and allow continuance
    # Get counts before and after merge
    hold_orig_address_count = len(orig_df)
    orig_df = pd.merge(orig_df, update_df, how="left", left_on=['master_key'], right_on=['update_key'])
    after_address_count = len(orig_df)
    logger.debug('Orig address file had records: ', hold_orig_address_count, 'Result of merge had records: ',
          after_address_count)

    orig_number_created = after_address_count - hold_orig_address_count
    if orig_number_created > 0:  # the merge added records
        exit_yes_no(str(orig_number_created) + " addresses were duplicated by merge.  Continue?",
                    "EXTRA RECORDS WERE CREATED BY MERGE",
                    display_exiting=False)


def zip_file_to_county_dict(zip_csv_path: Union[str, os.PathLike],
                            xlsx_path: Union[str, os.PathLike]) -> dict:
    """Build a state-county lookup dictionary from purchased zip data and write a summary XLSX.

    Reads the zip-codes.com DELUXE CSV, removes military states, derives
    filename-safe county names, and writes a deduplicated state/county list to
    ``xlsx_path``.

    Args:
        zip_csv_path (Union[str, os.PathLike]): Path to the zip data CSV
            (e.g. ``zip-codes-database-DELUXE-BUSINESS.csv``), ~80 K rows.
        xlsx_path (Union[str, os.PathLike]): Destination path for the
            unique county XLSX summary.

    Returns:
        dict: Mapping of ``"STATE-COUNTY"`` -> ``[county_filename, countyToPrint, stateMixedCounty]``.
    """

    logger.debug("creating dictionary of zip to county")
    zip_rows_to_read = 999_999  # for testing
    # zip_rows_to_read = 999  # for testing

    # original county is uppercase
    main_zip_file = pd.read_csv(zip_csv_path, nrows=zip_rows_to_read, keep_default_na=False,
                                usecols=['State', 'County', 'CountyMixedCase'])

    main_zip_file.rename(
        columns={'ZipCode': 'zip', 'State': 'state', 'County': 'county', 'CountyMixedCase': 'county_mixedcase'},
        inplace=True)

    # remove military states like AA and AE because they have no county
    main_zip_file = main_zip_file.loc[main_zip_file['county'].str.strip() != ""]

    # 'county' already has city in it and assumes county so good for statecounty.  Add 'county' for printing.
    # remove characters like ,- space
    main_zip_file['countyclean'] = main_zip_file['county'].apply(clean_field, case_convert='keep')
    main_zip_file['statecounty'] = main_zip_file['state'] + "-" + main_zip_file['countyclean']
    main_zip_file['county_filename'] = main_zip_file['county_mixedcase'].apply(clean_field, case_convert='keep')
    # add 'County' if 'City' is not at end of name
    main_zip_file['countyToPrint'] = np.where(main_zip_file['county_mixedcase'].str[-4:] != "City",
                                              main_zip_file['county_mixedcase'] + " County",
                                              main_zip_file['county_mixedcase'])
    main_zip_file['statecounty_mixed'] = main_zip_file['state'] + "-" + main_zip_file['county_filename']

    unique_county = main_zip_file.drop_duplicates(subset=['statecounty'], keep='last')
    sorted_unique_county = unique_county.sort_values(['statecounty'], ascending=[True])

    sorted_unique_county.to_excel(xlsx_path, index=False,
                                  columns=['statecounty', 'county_filename', 'countyToPrint', 'statecounty_mixed'])

    # create dict using zip function
    county_dict = dict([(k, [a, b, c]) for k, a, b, c in zip(sorted_unique_county['statecounty'],
                                                   sorted_unique_county['county_filename'],
                                                   sorted_unique_county['countyToPrint'],
                                                   sorted_unique_county['statecounty_mixed'])])

    return county_dict


def create_zip_to_county_list_dict(unique_zips: Union[str, Path],
                                   split_zips: Union[str, Path],
                                   text_file_for_created_dict: Union[str, Path]) -> dict:
    """Build a zip-to-county-list dictionary from purchased zip data and persist it to a text file.

    Merges the one-to-one zip/county CSV with the multi-county CSV, deduplicates,
    groups by zip code, and writes the resulting dict as text.

    Args:
        unique_zips (Union[str, Path]): Path to the primary zip CSV
            (one-to-one zip/county, may contain multiple cities per zip).
        split_zips (Union[str, Path]): Path to the multi-county zip CSV.
        text_file_for_created_dict (Union[str, Path]): Output path where the
            dictionary is written as a Python literal string.

    Returns:
        dict: Mapping of zip code (int) -> list of ``"STATE-COUNTY"`` strings.
    """
    logger.info("creating zip to county lists")
    zip_rows_to_read = 999_999
    # zip_rows_to_read = 9_999  # for testing

    main_zip_temp = pd.read_csv(unique_zips, nrows=zip_rows_to_read, keep_default_na=False,
                                usecols=['State', 'County', 'ZipCode'])
    multi_county_temp = pd.read_csv(split_zips, nrows=zip_rows_to_read, keep_default_na=False,
                                    usecols=['State', 'County', 'ZipCode'])

    # combined_temp = main_zip_temp.append(multi_county_temp, ignore_index=True)
    combined_temp = pd.concat([main_zip_temp, multi_county_temp], ignore_index=True)

    combined_temp['countyclean'] = combined_temp['County'].apply(clean_field, case_convert='upper')

    combined_temp['statecounty'] = combined_temp['State'] + "-" + combined_temp['countyclean']

    # combined_temp2 = combined_temp[['ZipCode', 'countyclean']]  # keep only two cols
    combined_temp2 = combined_temp[['ZipCode', 'statecounty']]  # keep only two cols
    unique_zip_county = combined_temp2.drop_duplicates(subset=['ZipCode', 'statecounty'], keep='last')

    df_for_dict = unique_zip_county.groupby(["ZipCode"], as_index=False).agg({'statecounty': list})
    zip_to_county_list_dict = dict(df_for_dict.values.tolist())

    with open(text_file_for_created_dict, 'w') as f:
        print(zip_to_county_list_dict, file=f)

    return zip_to_county_list_dict


def single_pivot_report(df: pd.DataFrame,
                        index_fields: Union[str, list[str]],
                        value_fields: str,
                        sheet_name: str,
                        single_piv_writer: pd.ExcelWriter,
                        second_pivot_by_count: bool = False) -> None:
    """Run a single pivot-table report and write it to an Excel worksheet.

    Writes the pivot starting at row 6, leaving rows 1-5 for titles added later.
    Appends a numeric suffix to ``sheet_name`` if a sheet with that name already
    exists. Optionally writes a second sheet sorted by count descending.

    Args:
        df (pd.DataFrame): Source data to pivot.
        index_fields (Union[str, list[str]]): Field(s) used as pivot index rows.
            A bare string is automatically wrapped in a list.
        value_fields (str): Field used for the count aggregation.
        sheet_name (str): Target Excel sheet name (max ~30 chars before suffix).
        single_piv_writer (pd.ExcelWriter): Open ExcelWriter to write into.
        second_pivot_by_count (bool): If True, write a second sheet sorted by
            count descending. Defaults to False.
    """

    logger.debug(f"here '{sheet_name=}'")

    sheet_name2 = 'Cnt,' + sheet_name[:26]

    if sheet_name in single_piv_writer.book.sheetnames or sheet_name2 in single_piv_writer.book.sheetnames:
        file_counter_max = 0
        for sheet_ in single_piv_writer.book.sheetnames:
            if re.compile(r'-(\d+)$').search(sheet_):  # to avoid None if not found
                file_counter = re.compile(r'-(\d+)$').search(sheet_).group(1)
                if int(file_counter) > file_counter_max:
                    file_counter_max = file_counter
        file_counter_next = str(int(file_counter_max) + 1)

        sheet_name = sheet_name[:-(len(file_counter_next)+1)] + '-' + file_counter_next
        sheet_name2 = sheet_name2[:-(len(file_counter_next)+1)] + '-' + file_counter_next

    if isinstance(index_fields, str):
        index_fields = [index_fields]  # to generalize to list and not force user to enter []s
    fields_not_in_file = set(index_fields) - set(df.columns)  # some fields not in df
    if fields_not_in_file:
        alert(f"Specified index field on pivot not in dataframe:\n{','.join(fields_not_in_file)}"
              f"\n\nAvailable fields are:\n{', '.join(set(df.columns))}",
              'Warning:Tabulate field not on file, continuing')
    else:
        df_pt = pd.pivot_table(df,
                               index=index_fields,
                               values=value_fields,
                               aggfunc='count',
                               margins=True)
        # take the df created by pivot and add percent field by dividing count by total number of addresses
        df_pt['Pct_of_Total'] = round(df_pt[value_fields] / df.shape[0] * 100, 1)
        df_pt.to_excel(single_piv_writer, sheet_name=sheet_name, startrow=5)

        if second_pivot_by_count:
            # same as above but sorted by count
            df_pt = df_pt.sort_values(value_fields, ascending=False)
            df_pt.to_excel(single_piv_writer, sheet_name=sheet_name2, startrow=5)


def pivot_and_other_reports(df: pd.DataFrame,
                            output_wks: Union[str, Path],
                            input_fn: str,
                            dict_address_concentration: dict) -> None:
    """Generate all pivot-table and summary reports for a formatted data file.

    Creates an Excel workbook at ``output_wks`` containing county summaries,
    removal reasons, group/factory/campaign breakdowns, address-concentration
    analysis, and optional county-zip mismatch reports.

    Args:
        df (pd.DataFrame): Formatted (and optionally removed) voter data.
        output_wks (Union[str, Path]): Full path for the output XLSX report file.
        input_fn (str): Source file name, included in report title row 3.
        dict_address_concentration (dict): Mapping of (state, city, address) ->
            ``{'desc': str, 'remove': str}`` for concentrated-address lookup.
    """

    logger.info("creating reports")

    writer = pd.ExcelWriter(output_wks, engine='openpyxl')
    df_clean = df[df['remove'] == '']

    def report_by_pull_group(df: pd.DataFrame, rpt_field: str, sheet_name: str) -> None:
        """Write a deduplicated listing of a grouped field by pull_group to the report workbook.

        Args:
            df (pd.DataFrame): Source DataFrame; sorted internally by pull_group and rpt_field.
            rpt_field (str): Column name of the grouped variable to report on.
            sheet_name (str): Excel sheet name for this report (e.g. 'SUMMARY...').
        """

        sorted_df = df[[rpt_field, 'pull_group']] \
            .sort_values(['pull_group', rpt_field, ], ascending=[True, True])
        added_counties = sorted_df.drop_duplicates(subset=[rpt_field], keep='first')
        added_counties.to_excel(writer, sheet_name=sheet_name, startrow=5, index=False)

    # Create summary sheet of Rawdata, Formatted and Removed
    # for other states, roll all counties in to the single rollup sentinel
    df['countysummed'] = np.where(df['state'] == ROV_SETUP['expectedstate'], df['statecounty'], ROLLUP_STATECOUNTY)

    # State by county for all including removed
    single_pivot_report(df, index_fields=['state', 'countysummed'], value_fields=['address'],
                        sheet_name="RawData by State-County", single_piv_writer=writer, second_pivot_by_count=False)

    if ROV_SETUP['add_filename_column_flag']:
        single_pivot_report(df, index_fields=['filename'], value_fields=['address'],
                            sheet_name="RawData by Filename",
                            single_piv_writer=writer, second_pivot_by_count=False)

    single_pivot_report(df, index_fields=['remove'], value_fields=['address'],
                        sheet_name="Removed Reasons",
                        single_piv_writer=writer, second_pivot_by_count=True)

    if ROV_SETUP['group_vars']:
        single_pivot_report(df_clean, index_fields=ROV_SETUP['group_vars'], value_fields=['address'],
                            sheet_name=f"Clean by {','.join(ROV_SETUP['group_vars'])[:22]}",
                            single_piv_writer=writer, second_pivot_by_count=True)

    # add filtered sheet w factory and campaigns for filling sincere campaign table
    if ROV_SETUP['factory_vars']:
        df_pt = pd.pivot_table(df_clean,
                                index=["factory_vars_string", "group_vars_string"],
                                aggfunc='count',
                                values="address",
                                margins=False).reset_index()
        df_pt.to_excel(writer, sheet_name="Campaigns by Factories", startrow=5)
        writer.book['Campaigns by Factories'].auto_filter.ref = 'B6'
        writer.book['Campaigns by Factories'].auto_filter.add_filter_column(0, [])
        writer.book['Campaigns by Factories']["D3"] = "Sum of Shown"
        writer.book['Campaigns by Factories']["D4"] = "= AGGREGATE(9,1,D6:D999)"
        writer.book['Campaigns by Factories']["E3"] = "Sum of All"
        writer.book['Campaigns by Factories']["E4"] = "=SUM(D6:D999)"

    # create reports showing added county, factory and campaign_vars by pull_group so Sincere and BOE data can be
    # adjusted
    report_by_pull_group(df, 'statecounty', 'StateCounties by pull_group')

    # list added campaigns.  These need to be added in factory tables
    if ROV_SETUP['campaign_vars']:
        report_by_pull_group(df, 'campaign_vars_string', 'campaign_vars by pull_group')

    # list added factories.  These will need to be added as new factories in Sincere.
    if ROV_SETUP['factory_vars']:
        report_by_pull_group(df, 'factory_vars_string', 'Factory_vars by pull_group')

    # list added groups (factory + county).  These will be the split files produced.  May need to be added to
    # old factories.
    if ROV_SETUP['group_vars']:
        report_by_pull_group(df, 'group_vars_string', 'Group_vars by pull_group')

    df_pt = pd.pivot_table(df, index=['state', 'county', 'city', 'address', 'conc_addr_desc', 'remove'],
                           values=['lastname'],
                           aggfunc='count')
    df_from_query = df_pt.query(f"lastname >= {PROP_CONCENTRATION}")
    if len(df_from_query) > 0:
        df_from_query.reset_index(inplace=True)
        df_from_query.rename(columns={'lastname': 'address_count', 'conc_addr_desc': 'addr_desc'}, inplace=True)

        # tried to eliminate error, "A value is trying to be set on a copy of a slice from a DataFrame", but couldn't so
        # isolated to with .copy() to show all is well
        dfq = df_from_query.copy()

        # open browser windows for concentrated properties
        address_concentration_open_browser(df_from_query)

        dfq.to_excel(writer, sheet_name=f"Address GT {PROP_CONCENTRATION}",
                     columns=['state', 'city', 'address', 'addr_desc', 'remove', 'address_count', 'county'],
                     startrow=5, index=False)

    # fields specific to county check - zip, county, zip/county match
    if ROV_SETUP['run_county_check_code_flag']:
        logger.debug('running countyCheck pivots')
        # recode opposite of mismatch field value so we can sum
        df['match_county'] = np.where(df["mismatch_county"] == 1, 0, 1)

        # % county mismatched
        df_pt = pd.pivot_table(df, index=['state', 'countysummed'],
                               values=['address', 'mismatch_county', 'match_county'],
                               aggfunc={'address': 'count', 'mismatch_county': 'sum', 'match_county': 'sum'},
                               margins=True)
        df_pt['Pct_Mis_Matched'] = round(df_pt['mismatch_county'] / df_pt['address'] * 100, 1)
        df_pt['Pct_Matched'] = round(df_pt['match_county'] / df_pt['address'] * 100, 1)
        df_pt.to_excel(writer, sheet_name="County Match Summary", startrow=5)

        # zips w/ # occurrences in ZIP_CONCENTRATION, mismatch first
        df_pt = pd.pivot_table(df, index=['state', 'statecounty', 'zip_county_list', 'zip', 'mismatch_county'],
                               values=['address'],
                               aggfunc={'address': 'count'})
        df_pt['Pct_of_Total'] = round(df_pt['address'] / df.shape[0] * 100, 1)
        df_from_query = df_pt.query("address >= " + str(ZIP_CONCENTRATION))
        df_from_query = df_from_query.sort_values(["mismatch_county", "address"], ascending=(False, False))
        df_from_query.rename(columns={'address': 'address_count'}, inplace=True)
        df_from_query.reset_index().to_excel(writer, sheet_name=f"Zips Over {ZIP_CONCENTRATION}", startrow=5,
                                             index=False)

    # pivots for ad hoc fields
    for specs in ROV_SETUP['pivot_specs']:
        if specs['pivot_fields']:
            dfx = (df if specs['pivot_for_all'] else df_clean)
            universe = ('All-' if specs['pivot_for_all'] else 'Cln-')
            single_pivot_report(dfx, index_fields=specs['pivot_fields'], value_fields=['address'],
                                sheet_name=f"{universe} {','.join(specs['pivot_fields'])[:26]}",
                                single_piv_writer=writer,
                                second_pivot_by_count=(True if specs['pivot_by_cnt'] else False))

    piv_wb = writer.book
    for ws in piv_wb.worksheets:
        autosize_xls_cols(ws)  # widen columns using BEK routine before wide titles

        ws["A1"] = "Address Summary Report"
        ws['A1'].font = Font(b=True, size=16)
        ws["A2"] = ws.title
        ws['A2'].font = Font(b=True, size=12)
        ws["A3"] = "Input File: " + input_fn
        ws['A3'].font = Font(b=True, size=12)
        ws["A4"] = datetime.now().strftime('%m/%d/%Y')

    logger.debug('out of pivots and other reports')

    writer.close()


def create_import_code_from_sheet(sheet_with_python_code: Worksheet,
                                  output_file: Union[str, Path]) -> TextIO:
    """Write Python code stored in a worksheet to a ``.py`` text file.

    Iterates rows in ``sheet_with_python_code``; for each row, the first
    non-None, non-comment cell determines indentation (4 spaces x column index).

    Args:
        sheet_with_python_code (Worksheet): Worksheet whose cells contain
            indented Python source lines.
        output_file (Union[str, Path]): File name (relative to ``ROV_SETUP['exe_path']``)
            of the output ``.py`` file.

    Returns:
        TextIO: The (closed) file object of the written ``.py`` file.
    """

    logger.debug("here")

    # sheet_with_python_code = setup_wb[sheet_with_python_code]
    with open(ROV_SETUP['exe_path'] / output_file, "w") as new_code_text:

        # Loop through sheet positions in excel file and indent as needed in python code (4 characters per col hardcoded)
        for rowidx, row_cells in enumerate(sheet_with_python_code.iter_rows()):
            # col and rows indexes needs +1 cause python 0 indexed, worksheet cells start at 1
            for colidx, cell in enumerate(row_cells):
                if sheet_with_python_code.cell(rowidx + 1, colidx + 1).value is not None and \
                        not sheet_with_python_code.cell(rowidx + 1, colidx + 1).value.strip().startswith('#'):
                    new_code_text.write(" " * colidx * 4 + sheet_with_python_code.cell(rowidx + 1, colidx + 1).value + '\n')
                    break  # dont go to subsequent columns once data is found to avoid extraneous info
    new_code_text.close()
    return new_code_text


def address_concentration_open_browser(df: pd.DataFrame) -> None:
    """Open browser search tabs for addresses flagged as highly concentrated.

    Behaviour is controlled by ``ROV_SETUP['concentrated_address_browser_prompt_freq']``:
    ``1`` opens without prompting, ``2`` prompts for confirmation first.
    Only opens tabs for rows where ``addr_desc`` and ``remove`` are both empty.

    Args:
        df (pd.DataFrame): Concentrated-address DataFrame with columns
            ``state``, ``city``, ``address``, ``addr_desc``, and ``remove``.
    """

    logger.debug("here")

    openbrowser = True
    if ROV_SETUP['concentrated_address_browser_prompt_freq'] in [1, 2]:
        if ROV_SETUP['concentrated_address_browser_prompt_freq'] == 2:
            choice = confirm("Open " + str(len(df)) + " tabs in browser?.  OK?\n\n",
                             'Open browser windows?', ['Yes', 'No'])
            if choice == "no":
                openbrowser = False

        if openbrowser:
            df = df.reset_index()  # converts multi-index to columns
            for index, row in df.iterrows():
                if row['addr_desc'] == '' and row['remove'] == '':
                    webbrowser.open(f"https://www.google.com/search?q={row['state']}+{row['city']}+{row['address']}", new=2)
        else:
            pass
    else:
        pass


def add_fields_to_list(base_list: list[str], new_fields: str) -> None:
    """Append comma-separated field names to a list, exiting if any already exist.

    Args:
        base_list (list[str]): The existing list of field names to extend in place.
        new_fields (str): Comma-separated string of field names to add.

    Raises:
        SystemExit: Via ``exit_yes`` if any field in ``new_fields`` is already
            present in ``base_list``.
    """

    new_fields_lst = [field.strip().lower() for field in new_fields.split(",")]
    same_fields = [field.strip().lower() for field in new_fields_lst if field in base_list]

    if not same_fields:
        base_list.extend(new_fields_lst)
    else:
        exit_yes(f"Field(s) '{str(same_fields)}' already exists on list.  Can not add it.")


def split_files_for_sincere(combined_csv: Union[str, Path],
                            lim: int,
                            op_wks: Union[str, Path]) -> None:
    """Split the combined CSV into per-group_var CSV files for Sincere/VoterLetters import.

    Large groups are further chunked into sub-files if they exceed ``lim`` rows.
    A summary XLSX listing each output file and its address count is written to ``op_wks``.

    Args:
        combined_csv (Union[str, Path]): Path to the combined CSV produced by the Combine step.
        lim (int): Maximum addresses per Sincere import file; 0 means no limit.
        op_wks (Union[str, Path]): Path for the output split-file-list XLSX.
    """

    logger.info("split files for Sincere")

    try:
        os.remove(op_wks)
    except Exception:
        pass

    wb = Workbook()
    ws = wb.active
    ws['A1'] = f"{ROV_SETUP['expectedstate']} {ROV_SETUP['splitfnbase'].stem}"
    ws['A2'] = 'Split Files to be Loaded'
    ws['A4'] = 'File'
    ws['B4'] = 'Address Count'

    def chunk_split_file(df: pd.DataFrame, limit: int,
                         split_path_hold: Path, split_filename: str) -> None:
        """Write a DataFrame to one or more chunked CSV files under ``split_path_hold``.

        If ``len(df) <= limit``, writes a single file. Otherwise splits into
        numbered sub-files (``split_filename file-1.csv``, ``file-2.csv``, ...).

        Args:
            df (pd.DataFrame): Address records for one split group.
            limit (int): Maximum rows per output file.
            split_path_hold (Path): Directory where split CSV files are written.
            split_filename (str): Root file name (without extension); used as-is
                for a single file or with a numeric suffix for chunks.
        """

        logger.trace("here")

        addresses_to_write = len(df)  # don't use df reference to save time

        if addresses_to_write <= limit:  # write one file
            split_file = split_path_hold / f"{split_filename}.csv"
            df.to_csv(split_file, index=False, columns=ROV_SETUP['splitfile_field_list'])

            ws.cell(column=1, row=ws.max_row + 1, value=f"{split_filename}.csv")
            ws.cell(column=2, row=ws.max_row, value=len(df))

            logger.info(f"'{split_filename}' written, {len(df)} addresses.")
        else:  # need to create split files by looping
            for file_counter in range(1, math.ceil(addresses_to_write / limit) + 1):
                low_record = ((file_counter - 1) * limit)
                hi_record = (file_counter * limit) - 1

                split_file = split_path_hold / f"{split_filename} file-{file_counter}.csv"
                df_chunk = df[low_record: hi_record + 1]
                df_chunk.to_csv(split_file, index=False, columns=ROV_SETUP['splitfile_field_list'])

                ws.cell(column=1, row=ws.max_row + 1, value=f"{split_filename} file-{file_counter}.csv")
                ws.cell(column=2, row=ws.max_row, value=len(df_chunk))

                logger.info(f" -'{split_filename} file- {file_counter}' written, {len(df_chunk)} addresses.")

    if lim == 0:
        lim = 99999999
    op_stem = ROV_SETUP['splitfnbase'].stem

    # Ask if sorting by zip ok.  Assumes current sort order is how data was sorted under Combine.
    if ROV_SETUP['sortchoice'] in [1, 2]:
        exit_yes_no("Output is being sorted by zip. OK?",
                    'SORT BY ZIP?',
                    display_exiting=False)

    try:
        df_combo_w_no_remove = pd.read_csv(combined_csv, header=0, keep_default_na=False)
    except Exception as e:
        logger.exception(e)
        exit_yes("Unable to read  the Combinedfile in the split step.  Was 'Combine' run?"
                 f"\n\nMissing file:\n\n{combined_csv}"
                 )

    df_combo_w_no_remove = df_combo_w_no_remove[df_combo_w_no_remove["remove"] == ""]
    logger.info(f"Split will process {len(df_combo_w_no_remove)} clean addresses.")

    exit_yes_no(f"Split will process {len(df_combo_w_no_remove)} clean addresses.",
                'SPLIT RECORDS',
                display_exiting=False)

    if ROV_SETUP['sort_list']:
        df_combo_w_no_remove.sort_values(by=ROV_SETUP['sort_list'], inplace=True)

    fields_missing_from_combinefile = set(ROV_SETUP['splitfile_field_list']) - set(df_combo_w_no_remove.columns)
    if fields_missing_from_combinefile:
        exit_yes("Combinefile is missing the following fields to write to Splitfiles:"
                 f"\n\n{', '.join(fields_missing_from_combinefile)}"
                 )

    if not ROV_SETUP['group_vars']:  # no split field specified so write out one file with name "Combined"
        split_filename = ROV_SETUP['expectedstate'] + '-' + "Combined " + op_stem
        chunk_split_file(df_combo_w_no_remove, lim, ROV_SETUP['split_path_hold'], split_filename)

    else:
        # Loop and create a split file for each group_var
        #
        # if len(ROV_SETUP['campaign_vars']) == 1 and ROV_SETUP['campaign_vars'][0].lower() == 'county':
        #     splitfield = 'statecounty'
        # else:
        #     splitfield = ROV_SETUP['campaign_vars']
        unique_split_values = sorted(df_combo_w_no_remove['group_vars_string'].unique())

        # for each - write out a csv file.
        for unique_split_value in unique_split_values:
            # print("split " + splitfield_value)
            df_one_splifield = df_combo_w_no_remove[df_combo_w_no_remove['group_vars_string'] == unique_split_value]

            split_filename = unique_split_value + "-" + op_stem

            # write out file, broken into chinks if needed
            chunk_split_file(df_one_splifield, lim, ROV_SETUP['split_path_hold'], split_filename)

    wb.save(op_wks)

    alert(f"The split-file names and address counts produced by 'split' have been put into: \n{op_wks}"
          f"\n\n\nCopy this information into the 'Sincere load checklist' tab of the setup sheet: "
          f"\n{ROV_SETUP['setup_file_name']} \n\nto get ready to load Sincere.",
          f"SPLIT COMPLETE")



def check_county_to_zips(df: pd.DataFrame,
                         zipskip_list: list[tuple],
                         dict_statecounty: dict) -> None:
    """Validate voter county values against the zip-code lookup and flag mismatches.

    Adds/updates in-place: ``orig_county``, ``clean_county``, ``statecounty``,
    ``numzip``, ``zip_county_list``, and ``mismatch_county`` columns.

    Args:
        df (pd.DataFrame): Formatted voter DataFrame to annotate in place.
        zipskip_list (list[tuple]): List of ``(county_lower, zip_int)`` tuples
            that should not be flagged as mismatches.
        dict_statecounty (dict): State-county lookup mapping ``"STATE-COUNTY"`` ->
            ``[county_filename, countyToPrint, stateMixedCounty]``.
    """

    logger.info("check county to zip file")

    df['orig_county'] = df['county']
    df['clean_county'] = df['county'].apply(clean_field)

    df['statecounty'] = df[['state', 'clean_county']] \
        .apply(lambda row: (row['state'] + "-" + row['clean_county']).upper(),
               axis="columns")

    # collapse all statecounty into the rollup sentinel for bad states
    df.loc[(df['state'] != ROV_SETUP['expectedstate']), 'statecounty'] = ROLLUP_STATECOUNTY

    df['numzip'] = df['zip'].map(lambda x: (int(float(x)) if is_number(x) else 0))

    # TODO: below mixed state counties with counties and makes it impossible to produce simple summaries of data
    # replace county with standard version so it matches filename, etc (value [0] is clean, mixed-case)
    df['county'] = df['statecounty'].map(lambda x: dict_statecounty.get(x, [x+'-statecounty not found'])[0])

    # set 'zip_county_list' to string of statecounties based on zip
    df['zip_county_list'] = df['numzip'].map(lambda x: ","
                                             .join(ROV_SETUP['dict_zip_to_countylist'].get(x, ['zip not found'])))

    # check for county in list to account for split zips
    # df['mismatch_county'] = df.apply(lambda dfx: (1 if dfx['clean_county'] not in dfx['zip_county_list']
    #                                               else 0), axis=1)
    df['mismatch_county'] = df.apply(lambda row: (1 if row['statecounty'] not in row['zip_county_list']
                                                  else 0), axis=1)

    # if flag is set, use zip skip list to reset county/zip mismatches to 0
    if ROV_SETUP['skip_selected_zip_match_flag']:
        df['mismatch_county'] = df.apply(lambda row:
                                         (0 if (row['county'].lower(), row['zip'])
                                               in zipskip_list else row['mismatch_county']), axis='columns')

    logger.debug("Done filling zip and county info")


def get_setup_file_name(initial_campaign_dir: Union[str, Path]) -> Path:
    """Prompt the user to select an cleaver setup XLSX file and return its path.

    In test mode (``USE_HARDCODED_SETUP = True``), skips the dialog and returns
    ``HARDCODED_SETUP_FILE`` after a confirmation prompt. Exits if the selected
    file name does not contain ``'UniversalSetup'``.

    Args:
        initial_campaign_dir (Union[str, Path]): Starting directory for the file
            picker dialog.

    Returns:
        Path: Resolved path to the selected setup XLSX file.
    """

    logger.debug('picking setup file')

    if USE_HARDCODED_SETUP:
        # Hardcode in TEST INPUT FILE directory for repetitive testing
        setup_file_name = Path(HARDCODED_SETUP_FILE).expanduser()

        exit_yes_no("Running hardcoded Setup file.  OK?\n\n" + str(setup_file_name),
                    'RUN IN TEST?',
                    display_exiting=False)
    else:
        # show an "Open" dialog box and return the path to the selected file
        xl_setup_file_name = select_file("Select Setup File",
                                         start_dir=str(initial_campaign_dir),
                                         files_like="*UniversalSetup*.xlsx",
                                         mode="file",
                                         title2="Select cleaver setup file xlsx. Must have 'UniversalSetup' in name")
        if xl_setup_file_name is None:
            exit_yes("No setup file selected.")
        setup_file_name = Path(xl_setup_file_name).expanduser()

    if not 'universalsetup' in str(setup_file_name).lower():
        exit_yes(("The chosen setup file does not contain 'UniversalSetup' in it's name.\n"
                  "\n\nYou need to pick a different setup file or "
                  ))

    return setup_file_name


def check_for_unwanted_setup_options() -> None:
    """Validate setup options and prompt the user to confirm or abort before processing.

    Checks for potentially unintended combinations such as running county/zip
    validation without a county split field, unusual sort orders, format-file
    copies, and imported code blocks. Reads options from ``ROV_SETUP``.
    """

    logger.info('Checking setup options')

    # flag if checking county/zip and county or statecounty not in group_vars
    if ROV_SETUP['run_county_check_code_flag'] and \
            set(ROV_SETUP['group_vars']).intersection(set(['county', 'statecounty'])) == set():
        exit_yes_no(("You are checking for zip/county mismatches"
                     "\nbut you are not splitting by county."
                     "\n\n\nIs this what you meant?"))

    if ROV_SETUP['skip_selected_zip_match_flag']:
        str_zipskip = str()
        for county, zip in ROV_SETUP['zipskip_list']:
            str_zipskip += f"{county:<20} {zip}\n"

        exit_yes_no("County and zip combinations to ignore when flagging mismatches "
                    f"(listed in 'ZipSkip' setup sheet): \n{str_zipskip}",
                    "Zip Codes to Ignore in County Matching")

    if ROV_SETUP['sortchoice'] in [1, 2]:
        exit_yes_no("Sorting by other than county / random number can cause ugly clumps of addresses "
                    "(eg all PO boxes).  OK?",
                    "Sort Order")

    if ROV_SETUP['copy_other_format_files_flag']:
        exit_yes_no("Copy FORMAT files listed in setup sheet 'FormatCopies' during combine?\n\n",
                    'COPY FORMAT FILES?')

        formatfile_copy(ROV_SETUP['copy_formatfile_filelist_sheet'], perform_copies=True)

    if ROV_SETUP['run_first_code_flag']:
        # display code and prompt if it should be run
        display_imported_code(ROV_SETUP['first_code_sheet'], ROV_SETUP['first_code'])

    if ROV_SETUP['run_middle_code_flag']:
        display_imported_code(ROV_SETUP['middle_code_sheet'], ROV_SETUP['middle_code'])

    if ROV_SETUP['run_last_code_flag']:
        display_imported_code(ROV_SETUP['last_code_sheet'], ROV_SETUP['last_code'])


def create_field_lists() -> None:
    """Build all derived field-name lists in ``ROV_SETUP`` based on setup options.

    Populates ``formatfile_field_list``, ``formatfile_field_list_to_zip``,
    ``inputfile_delete_field_list``, ``inputfile_rename_fields_dict``,
    ``combinefile_field_list``, and related lists. Exits on duplicate field
    names, missing group_var fields, or factory/campaign misconfigurations.
    """

    logger.info('Creating field lists')

    # create 'formatfile' field list by replacing blanks with inputfile names and replacing those to be renamed,
    # keep 'x' so we can zip with input fields
    ROV_SETUP['formatfile_field_list_to_zip'] = [input_field if rename_field == '' else rename_field
                                        for (input_field, rename_field)
                                        in itertools.zip_longest(ROV_SETUP['inputfile_orig_list'],
                                                                 ROV_SETUP['inputfile_renamed_list'])]

    ROV_SETUP['inputfile_delete_field_list'] = ['_' + ofield
                                       for ofield, new_field
                                       in zip(ROV_SETUP['inputfile_orig_list'], ROV_SETUP['formatfile_field_list_to_zip'])
                                       if new_field.strip().lower() == 'x']

    # create dict to rename fields that need it, otherwise leave with prefix
    ROV_SETUP['inputfile_rename_fields_dict'] = {('_' + orig_field): new_field
                                        for orig_field, new_field
                                        in zip(ROV_SETUP['inputfile_orig_list'], ROV_SETUP['formatfile_field_list_to_zip'])
                                        if new_field.strip().lower() != 'x'}

    # list of actual fields to be included on file ('x's removed)
    ROV_SETUP['formatfile_field_list'] = [field for field in ROV_SETUP['formatfile_field_list_to_zip'] if field != 'x']

    # check for duplicate field values
    duplicate_fields = [field for field, count in collections.Counter(ROV_SETUP['formatfile_field_list']).items() if count > 1]
    # duplicate_fields = [field for field, count in collections.Counter(ROV_SETUP['formatfile_field_list']).items() if count > 1]
    if duplicate_fields:
        exit_yes((f"The following fields are duplicated on the Format file.  Remove one occurance."
                  f"\n\n{', '.join(duplicate_fields)}"
                  ))

    ROV_SETUP['combinefile_field_list'] = [field
                                  for field in ROV_SETUP['formatfile_field_list']
                                  if field not in ROV_SETUP['combinefile_fields_to_delete_list']]  # dont use set to keep order

    if ROV_SETUP['run_county_check_code_flag']:  # V15.0
        add_fields_to_list(ROV_SETUP['formatfile_field_list'], "orig_county,clean_county,zip_county_list,statecounty,"
                                                  "mismatch_county,"
                                                  "numzip")

    # add 'filename' field if requested
    if ROV_SETUP['add_filename_column_flag']:  # V15.0
        add_fields_to_list(ROV_SETUP['formatfile_field_list'], "filename")

    # Always add 'remove' field - makes reporting easier and will almost always be used
    add_fields_to_list(ROV_SETUP['formatfile_field_list'], 'remove')
    add_fields_to_list(ROV_SETUP['formatfile_field_list'], 'rownum')  # V15.0
    add_fields_to_list(ROV_SETUP['formatfile_field_list'], 'dupe_key')  # V16.0
    add_fields_to_list(ROV_SETUP['formatfile_field_list'], 'dupe_id_field')  # V16.0
    add_fields_to_list(ROV_SETUP['formatfile_field_list'], 'pull_group')
    add_fields_to_list(ROV_SETUP['formatfile_field_list'], 'custom_field')
    add_fields_to_list(ROV_SETUP['formatfile_field_list'], 'campaign_vars_string')
    add_fields_to_list(ROV_SETUP['formatfile_field_list'], 'factory_vars_string')
    add_fields_to_list(ROV_SETUP['formatfile_field_list'], 'group_vars_string')

    if ROV_SETUP['add_random_number_column_flag']:
        # add field randnum so we can sort FORMAT files by county and randnum
        add_fields_to_list(ROV_SETUP['formatfile_field_list'], "randnum")  # V15.0

    # Check that no fields are specified on output that are not on input  # V15.0 commented out -
    fields_missing_from_input = set(ROV_SETUP['combinefile_field_list']) - \
                                set(ROV_SETUP['formatfile_field_list'])
    # this doesn't need to check input to format which is what it was
    if fields_missing_from_input:
        exit_yes((f"Field(s) are specified to output on Combine file but are not present on Format.  "
                  f"\n\nMissing field(s) specified are:\n\n"
                  f"'{', '.join(fields_missing_from_input)}'"
                  f"\n\nFields on input are:\n\n"
                  f"{', '.join(ROV_SETUP['formatfile_field_list'])}"
                  ))

    # Flag if factory and no campaign
    if ROV_SETUP['factory_vars'] and not ROV_SETUP['campaign_vars']:
        exit_yes((f"factory_vars specified '{ROV_SETUP['factory_vars']}' but no campaign_vars.  Switch and rerun.\n\n"
                  ))

    # Check if group_vars is on field list, error if not
    if ROV_SETUP['group_vars'] and \
            not set(ROV_SETUP['group_vars']).issubset(set(ROV_SETUP['formatfile_field_list'])):
        exit_yes((f"some of group_vars '{ROV_SETUP['group_vars']}' are missing from Format file field list.\n\n"
                  f"Available fields are:\n{', '.join(ROV_SETUP['formatfile_field_list'])}"
                  ))


def create_dicts() -> None:
    """Build runtime lookup dictionaries and load supporting data into ``ROV_SETUP``.

    Creates: ``dict_statecounty_to_alt_formats`` (state-county -> format variants),
    optionally ``dict_zip_to_countylist`` (zip -> county list), and
    ``dict_concentrated_addresses`` (address tuple -> description/remove reason).
    Also populates ``zipskip_list``.
    """

    logger.info('Creating dictionaries')

    # create county dict returning various formats with
    # GA-CHATHAM as key: [0] is filename format; [1] print; [2] state-mixed
    ROV_SETUP['dict_statecounty_to_alt_formats'] = zip_file_to_county_dict(
        MAIN_ZIP_FILE,
        MULTI_COUNTY_ZIP_FILE)

    logger.debug('Ran Counties_to_xls')

    if ROV_SETUP['run_county_check_code_flag']:
        with open(ZIP_TO_COUNTY_LIST_FILE, "r") as dict_file:
            ROV_SETUP['dict_zip_to_countylist'] = ast.literal_eval(dict_file.read())
            logger.debug("Imported " + ZIP_TO_COUNTY_LIST_FILE.name)

    # check headers
    check_ws_headers(ROV_SETUP['concentrated_addresses_sheet'],
                       [('A1', 'State'),
                        ('B1', 'City'),
                        ('C1', 'Address'),
                        ('D1', 'Desc'),
                        ('E1', 'Remove'),
                        ('F1', 'cnt'),
                        ('G1', 'County'),
                        ])

    conc_addr_df = pd.read_excel(ROV_SETUP['concentrated_addresses_file'], sheet_name='Addresses', header=0, dtype=str)
    conc_addr_df.columns = [str(col).lower() for col in conc_addr_df.columns]
    conc_addr_df.fillna("", inplace=True)

    ROV_SETUP['dict_concentrated_addresses'] = dict([
        ((state, city, address), {'desc': desc, 'remove': remove})
        for state, city, address, desc, remove
        in zip(conc_addr_df['state'].apply(lambda x: clean_field(x)),
               conc_addr_df['city'].apply(lambda x: clean_field(x)),
               conc_addr_df['address'].apply(lambda x: clean_field(x)),
               conc_addr_df['desc'].str.strip(),
               conc_addr_df['remove'].str.strip()
               )
    ])

    # ROV_SETUP['dict_concentrated_addresses'] = dict(address_desc_list)
    # this dict can be used in remove code sheet; returns a two-tuple [0] is address description, [1] remove reason
    # code like this: if ROV_SETUP['dict_concentrated_addresses'].get(('NC','new hanover','wilmington', '811 martin st'),
    # 'Other') != 'Other' => set code
    # or df['remove'] = ROV_SETUP['dict_concentrated_addresses'].get(('NC','new hanover','wilmington', '811 martin st'), 'Other')[1]

    # Create list of county/zips to not flag as mismatched for a particular county.  Key is tuple of county and zip.
    if ROV_SETUP['skip_selected_zip_match_flag']:
        zip_skip_range = range_to_list(ROV_SETUP['skip_selected_zip_sheet'], 2, len(ROV_SETUP['skip_selected_zip_sheet']['A']), 1, 2)
        ROV_SETUP['zipskip_list'] = [(cnty.lower(), int(zipcode)) for cnty, zipcode in zip_skip_range]
    else:
        ROV_SETUP['zipskip_list'] = []  # passed to subs so need a value

    # FIXME: Can we fill group_vars here and rename function to create_dicts_fill_groups?

def display_imported_code(sheet_name: Worksheet, py_file_name: str) -> None:
    """Extract Python code from a setup worksheet, compile it, display it, and prompt to run.

    Writes the code to a ``.py`` file, compiles it for syntax checking,
    displays up to 40 lines in a message box, and loads the module into
    ``ROV_SETUP`` for later invocation.

    Args:
        sheet_name (Worksheet): The worksheet object containing Python source lines.
        py_file_name (str): Name of the output ``.py`` file
            (e.g. ``'first_code_to_import.py'``).

    Raises:
        Exception: If the Python code fails to compile (syntax error).
    """

    logger.debug('here')

    # put the text code in py_file_name
    create_import_code_from_sheet(sheet_name, py_file_name)

    with open(ROV_SETUP['exe_path'] / py_file_name, "r") as myfile:  # this copies in the code but does not execute it
        logger.debug(f"compiling {py_file_name} - taken from sheet: {sheet_name}")
        source_text = myfile.read()
        try:
            codeobj = compile(source_text, py_file_name, 'exec')
        except Exception as e:
            msg = (f"There was an error compiling '{py_file_name}', line number {e.lineno}, from sheet: "
                        f"'{sheet_name.title}'.  "
                        f"The text of the line is: '{e.text.strip()}'.  "
                        f"Fix and rerun.")
            logger.info(msg)
            logger.exception(e)
            bek_text_box(msg, 'Import Code Error', '')
            raise Exception

        # list the lines to be executed for debugging, with line nums and comments/blanks
        source_lines = source_text.splitlines(keepends=True)
        code_lines = [f"{index + 1:<4}  {line}"
                      for index, line in enumerate(source_lines)]
        if len(code_lines) > 40:  # too many to display on screen, disables the 'ok' box
            skipped = len(code_lines) - 40
            code_lines = code_lines[:40]
            code_lines.append(f"<{skipped} lines not shown>")
        code_lines = ' '.join(code_lines)
        logger.info(f"\n{py_file_name} to be run - taken from sheet: {sheet_name}")
        [print(f"{index + 1:<4}  {line}", end=' ') for index, line in enumerate(source_lines)]

        # Below names module based on py code name, eg 'first_code_to_be_run_module'; puts pointer to module in
        # ROV_SETUP
        # A function from this module (can even have passed parameters) will be called later in the program.
        ROV_SETUP[Path(py_file_name).stem + '_module'] = \
            importlib.import_module(Path(py_file_name).stem)

    exit_yes_no(f"'{py_file_name}' to be run - taken from setup sheet: '{sheet_name}'"
                f"\n\nCode and line numbers printed in python console log, too.\n\n"
                f"{code_lines}",
                f"Check '{py_file_name}' Code",
                display_exiting=False)


def process_format_files(filelist_wks: Worksheet) -> None:
    """Iterate the FileList sheet and format each row flagged with 'x'.

    Calls ``process_format_file`` for each flagged row and optionally
    ``merge_into_format_file`` when an update file is specified. Collects
    missing-county names across all files and alerts the user at the end.

    Args:
        filelist_wks (Worksheet): The 'FileList' sheet from the setup workbook.
    """
    logger.info('process format files')

    cumulative_missing_counties_list = []

    for fn, format_flag, combine_flag, update_fn, update_fields, pull_group, custom_field, notes, *_ \
            in filelist_wks.iter_rows(min_row=4, values_only=True):

        if pull_group is None:  # FIXME 0 should work if not set, not being used.  Unintended consequences?
            pull_group = 0

        if str(format_flag).strip().lower() == "x":
            ip = process_format_file(fn, pull_group, custom_field, ROV_SETUP['rawdata_path'],
                                     ROV_SETUP['format_path'],
                                     cumulative_missing_counties_list)  # file appended to in function

            if str(update_fn).strip().lower() != 'none':
                merge_into_format_file(ip, str(update_fn), ROV_SETUP['format_path'])

    missing_counties_file = ROV_SETUP['root_path'] / 'missing_counties.csv'
    if os.path.isfile(missing_counties_file):
        os.remove(missing_counties_file)

    if cumulative_missing_counties_list:
        alert("The following counties are not in the lookup file, written to 'missing_counties.csv':\n\n" +
              ",".join(cumulative_missing_counties_list),
              "Alert")
        with open(ROV_SETUP['exe_path'] / missing_counties_file, mode='wt', encoding='utf-8') as myfile:
            myfile.write('\n'.join(cumulative_missing_counties_list))


def process_format_file(fn: str,
                        pull_group: int,
                        custom_field: str,
                        input_path: Path,
                        op_path: Path,
                        missing_counties_list_all_formatfiles: list) -> pd.DataFrame:
    """Read a raw data file, apply all transformations, and write a FORMATTED CSV.

    Reads ``fn`` from ``input_path``, renames/prefixes fields per setup, runs
    first/middle/last code blocks, optionally checks county-zip consistency,
    writes a FORMATTED CSV and a REMOVED CSV to ``op_path``, and writes pivot
    summary reports.

    Args:
        fn (str): File name of the raw data file (CSV or XLSX).
        pull_group (int): Pull-group identifier assigned to all rows in this file.
        custom_field (str): Custom field value assigned to all rows in this file.
        input_path (Path): Directory containing the raw data file.
        op_path (Path): Directory where FORMATTED and REMOVED output files are written.
        missing_counties_list_all_formatfiles (list): Accumulator list; counties
            not found in the state-county dictionary are appended here.

    Returns:
        pd.DataFrame: The formatted DataFrame (including removed rows).
    """
    logger.info(f"Creating formatted file for: '{fn}'")

    start_time = datetime.now()

    ip_file_w_path = input_path / fn

    # Make sure data file exists
    bad_file_exit(input_path / fn, "Data file does not exist. Change filelist or "
                                  "copy into Rawdata directory.\n\n" + str(input_path / fn))
    # find the header row in the input file using an expected string and column.  error if > 30.  Defaults to search sheet0.
    header_row = find_header_row_in_file(ip_file_w_path, ROV_SETUP['check_header_string'], ROV_SETUP['strcheck_header_col'])

    ip = read_file_to_df(ip_file_w_path, **{'header': header_row, 'sheet_name': 0, 'nrows': ROV_SETUP['rows_to_read_limit'], 'keep_default_na': False, 'dtype': str})

    # convert column names to lower case
    ip.columns = [field.strip().lower() for field in ip.columns.values]

    # do all fields on file being read match expected?
    bad_header_fields = set(ip.columns) ^ set([field for field in ROV_SETUP['inputfile_orig_list'] if field != ''])
    if bad_header_fields:
        exit_yes((f"The fields in header for '{fn}' do not match the ones in orig format files."
                  f"\nDifferences:\n\n{os.linesep.join(bad_header_fields)}"
                  # had to use os.linesep cause \n not allowed in fstrings
                  ))

    # convert columns identified as numeric in setup field list
    numeric_cols = [column_index for column_index, col_type in enumerate(ROV_SETUP['inputfile_type_list']) if
                    col_type == 'int']
    # non-blank are set to integer; code options here if move to other types
    for col in numeric_cols:
        ip[ip.columns[col]] = pd.to_numeric(ip[ip.columns[col]], errors='coerce', downcast='integer')
        # try to force specified numeric columns to integer
    logger.debug("Done reading in one file in process_format_file, csv or xlsx")

    ## rename fields with prefix of '_' so not conflicted with renamed field of same name
    logger.debug("Ready to prefix fields to with '_'")
    ip = ip.add_prefix('_')

    ip.rename(ROV_SETUP['inputfile_rename_fields_dict'], axis='columns', inplace=True)

    add_fields_list = [field for field in ROV_SETUP['formatfile_field_list'] if field not in ip.columns]
    # add new fields. check for duplicate field names on output is already done above
    ip = ip.reindex(columns=ip.columns.tolist() + add_fields_list, fill_value='')

    # fill rownum field
    ip['rownum'] = ip.index + 2  # +1 for 0index, +1 for header
    logger.debug("Done fill rownum")

    # Check for literal of TRUE or FALSE in fields because it gets converted to boolean and kills the program
    logger.debug("Ready to fill firstname true/false")
    replace_boolean_column_vals(ip, ['firstname', 'lastname', 'lastnametemp', 'address',
                                     'addresstemp', 'address2', 'city'])
    logger.debug("Done filling specified fields true/false with blank")

    # create filename filed here so we can use in recode below
    if ROV_SETUP['add_filename_column_flag']:
        ip['filename'] = fn

    # V16.1 add 'custom_field' to df to use in recoding sections if desired
    ip['pull_group'] = pull_group
    ip['custom_field'] = custom_field

    # create random number so we can sort
    if ROV_SETUP['add_random_number_column_flag']:
        np.random.seed(0)
        ip['randnum'] = np.random.random(ip.shape[0])

    if ROV_SETUP['run_first_code_flag']:
        # *** Only run imported first_code and middle_code in format, not combine to keep things like remove
        # assignment, random number from being overwritten.
        # last_code can be run in combine since it's only setting remove code.
        logger.info('Ready to run first_code')

        ROV_SETUP['first_code_to_import_module'].first_code_func(ip, ROV_SETUP=ROV_SETUP)

        # This is the function from the sheet with any parameters it needs
        logger.debug('Ran first_code')  # these prompts help if error in imported code

    if ROV_SETUP['run_county_check_code_flag']:
        check_county_to_zips(ip, ROV_SETUP['zipskip_list'], ROV_SETUP['dict_statecounty_to_alt_formats'])
        # FIXME pymsgbox list of mismatched counties not showing

    if ROV_SETUP['run_middle_code_flag']:
        # *** Only run imported first_code and middle_code in format, not combine to keep things like remove
        # assignment, random number from being overwritten.
        # last_code can be run in combine since it's only setting remove code.
        logger.debug('Ready to run middle_code')  # these prompts help if error in imported code

        ROV_SETUP['middle_code_to_import_module'].middle_code_func(ip, ROV_SETUP=ROV_SETUP)

        # This is the function from the sheet with any parameters it needs
        logger.debug('Ran middle_code')  # these prompts help if error in imported code

    # fill campaign, factory and group var strings with concat of var values separated by '-'; after middle so created
    # vars are set
    ip['campaign_vars_string'] = ip[ROV_SETUP['campaign_vars']].agg('-'.join, axis=1)  # FIXME: df field should be
    # filled here
    ip['factory_vars_string'] = ip[ROV_SETUP['factory_vars']].agg('-'.join, axis=1)
    ip['group_vars_string'] = ip[ROV_SETUP['group_vars']].agg('-'.join, axis=1)

    # if requested, RUN code to remove unwanted records
    if ROV_SETUP['run_last_code_flag']:
        # *** Only run imported first_code and middle_code in format, not combine to keep things like remove
        # assignment, random number from being overwritten.
        # last_code can be run in combine since it's only setting remove code.
        logger.debug('Ready to run last_code (remove code)')  # these prompts help if error in imported code

        # ROV_SETUP['last_code_to_import_module'].last_code_func(ip, ROV_SETUP['dict_concentrated_addresses'],
        #                                                     ROV_SETUP['expectedstate'])
        ROV_SETUP['last_code_to_import_module'].last_code_func(ip, ROV_SETUP=ROV_SETUP)
        # This is the function from the sheet with any parameters it needs

        # set remove based on max_pull_group if use_max_pull_group is set
        if ROV_SETUP['use_max_pull_group']:
            ip.loc[(ip['remove'] == '') & (ip['pull_group'] < ROV_SETUP['max_pull_group']), 'remove'] = \
                f"Clean but pull group less than {ROV_SETUP['max_pull_group']}"

        logger.debug('Ran last_code')  # these prompts help if error in imported code

    ip.drop(ROV_SETUP['inputfile_delete_field_list'], axis=1, inplace=True)

    # Sort file by randnum if flag is set
    if ROV_SETUP['sort_list']:  # TODO why are we even if sorting if not set?  to make remove code faster?
        ip.sort_values(by=['remove', 'zip', 'address', 'randnum'], inplace=True)

    # Make sure county is on file, find mismatched counties and return for accumulating
    if ROV_SETUP['run_county_check_code_flag']:
        if 'county' not in ip.columns:
            exit_yes("'CHECK COUNTY INFO FLAG' option is specified as 'TRUE' but 'county' field is not present.")

        unique_statecounties = ip.loc[ip['remove'] == '', 'statecounty'].unique()
        unique_statecounties = sorted(ip['statecounty'].unique())
        # The rollup sentinel for out-of-state rows is never a real zip-lookup key, so ignore
        # it here.  Otherwise it always appears as a missing county and fires the alert popup
        # even when the only "mismatch" is the rollup.
        missing_counties_this_formatfile = [chkfield
                                            for chkfield in unique_statecounties
                                            if chkfield.upper() != ROLLUP_STATECOUNTY.upper()
                                            and chkfield.upper() not in ROV_SETUP['dict_statecounty_to_alt_formats']]

        missing_counties_list_all_formatfiles.extend(missing_counties_this_formatfile)

    op_file = op_path / ("FORMATTED " + str(Path(fn).stem) + ".csv")
    # must check if remove flag is set otherwise field remove field is not on df

    ip.to_csv(op_file, index=False, columns=ROV_SETUP['formatfile_field_list'])

    if ROV_SETUP['run_last_code_flag']:
        # write out the 'removed' file
        op_file = ROV_SETUP['format_path'] / "Removed" / ("REMOVED " + str(Path(fn).stem) + ".csv")
        ip[ip["remove"] != ""].to_csv(op_file, index=False, columns=ROV_SETUP['formatfile_field_list'])

    logger.info("\nFormatted file: " + fn)
    logger.info("   Input records: " + str(len(ip)))
    if ROV_SETUP['run_last_code_flag']:
        logger.info("   Kept: " + str(len(ip[ip["remove"] == ""])))
        logger.info("   Removed: " + str(len(ip[ip["remove"] != ""])))

    logger.debug("\n", ip.head(5), "\n\n")  # TODO why does this not print??

    logger.debug("End time is ", datetime.now().strftime("%H:%M:%S"), "  Elapsed time is", str(datetime.now() - start_time),
          " (H:M:S.s)")

    # Create summary sheet of Rawdata, Formatted and Removed
    pivot_file = ROV_SETUP['format_path'] / "Summary" / f"SUMMARY {Path(fn).stem}.xlsx"
    pivot_and_other_reports(ip, pivot_file, fn, ROV_SETUP['dict_concentrated_addresses'])

    logger.debug("Leaving process_format_file")
    return ip


def formatfile_copy(ws_copy_formatfile_filelist: Worksheet,
                    perform_copies: bool = True) -> None:
    """Copy FORMATTED files listed in the FormatCopies sheet to their destination directories.

    Validates paths and file existence for all rows flagged with 'x'. When
    ``perform_copies`` is False, performs validation only without copying.

    Args:
        ws_copy_formatfile_filelist (Worksheet): The 'FormatCopies' worksheet
            with columns: use, fromFilePath, fromfilename, tofilepath, tofilename.
        perform_copies (bool): If True, execute the file copies. If False,
            only validate paths and files. Defaults to True.
    """

    logger.info('In formatfile_copy')

    def copy_file(source: Union[str, Path], destination: Union[str, Path]) -> None:
        """Copy a single file to a destination, exiting with a message on failure.

        Args:
            source (Union[str, Path]): Full path to the source file.
            destination (Union[str, Path]): Destination directory or file path.
        """
        try:
            shutil.copy(os.path.expanduser(source), destination)
            logger.info(f"File '{source}' copied successfully to\n'{destination}'.")
        except Exception as e:
            logger.exception(e)
            exit_yes((f"File not copied:\n\n {source} \n\nto\n\n {destination}"
                      ))

    # Make sure column heading/locations are as expected and nothing was moved
    check_ws_headers(ws_copy_formatfile_filelist,
                     [('A1', 'use'),
                      ('B1', 'fromFilePath'),
                      ('C1', 'fromfilename'),
                      ('D1', 'tofilepath'),
                      ('E1', 'tofilename'),
                     ])

    # make sure all files and directories exist for row in ws_filelist:
    for copy_formatfile_flag, from_path, from_fn, to_path, to_fn, *_ in islice(ws_copy_formatfile_filelist, 1, None):
        # islice starts at row index 1 not 0 to skip header; *_ discards unused cols (notes, etc)
        if str(copy_formatfile_flag.value).strip().lower() == "x":
            src_path, src_file_w_path, dest_path, dest_file_w_path = \
                assign_formatcopy_vars(from_path, from_fn, to_path, to_fn)

            bad_path_exit(src_path)
            bad_file_exit(file=Path(src_file_w_path))
            bad_path_exit(dest_path)
            # bad_file_exit(file=dest_file_w_path)  # cant check cause doesn't exist

            # if we got to here, paths and files are ok so copy
            if perform_copies:
                copy_file(src_file_w_path, dest_path)


def assign_formatcopy_vars(from_path: Cell, from_fn: Cell,
                           to_path: Cell, to_fn: Cell) -> list:
    """Resolve openpyxl cell values for source/destination paths into usable filesystem paths.

    Source path may be a Python expression (evaluated if it contains ``','``,
    ``'('``, or ``' / '``), a literal path string, or a ``~``-expandable path.
    A blank or ``'none'`` destination path defaults to ``ROV_SETUP['format_path']``.

    Args:
        from_path (Cell): Cell whose value is the source directory expression or path.
        from_fn (Cell): Cell whose value is the source file name.
        to_path (Cell): Cell whose value is the destination directory (or blank/none).
        to_fn (Cell): Cell whose value is the destination file name (or blank/none).

    Returns:
        list: ``[src_path, src_file_w_path, dest_path, dest_file_w_path]``
            where each element is a resolved path string or Path object.
    """

    logger.debug('here')
    # from_path, from_fn, to_path, to_fn
    src_path_expr = from_path.value
    src_file = str(from_fn.value).strip()
    xl_dest_path = str(to_path.value).strip()
    xl_dest_file = str(to_fn.value).strip()

    if "," in src_path_expr or "(" in src_path_expr or " / " in src_path_expr:
        src_path = eval(src_path_expr)
    elif "/" in src_path_expr or "\\" in src_path_expr:
        src_path = src_path_expr
    else:
        src_path = eval(os.path.expanduser(src_path_expr))

    src_file_w_path = os.path.join(src_path, src_file)  # FIXME test formatcopy, remove os.path.join

    if xl_dest_path == "" or xl_dest_path.lower() == 'none':  # no "to path" specified so use default, format dir
        dest_path = ROV_SETUP['format_path']
    else:
        dest_path = os.path.join(xl_dest_path, "")

    if xl_dest_file == "" or xl_dest_file.lower() == 'none':  # no "to path" specified so use default, format dir
        dest_file_w_path = os.path.join(dest_path, src_file)
    else:
        dest_file_w_path = os.path.join(dest_path, xl_dest_file)

    return [src_path, src_file_w_path, dest_path, dest_file_w_path]


def combine_formatfiles(orig_df: pd.DataFrame,
                        copy_formatfile_list_sheet: Worksheet) -> pd.DataFrame:
    """Append records from externally copied FORMAT files into the combined DataFrame.

    Reads each file flagged with 'x' in ``copy_formatfile_list_sheet``, validates
    that required split fields are present, and concatenates the records.

    Args:
        orig_df (pd.DataFrame): The current combined DataFrame to extend.
        copy_formatfile_list_sheet (Worksheet): The 'FormatCopies' worksheet with
            columns: use, fromFilePath, fromfilename, tofilepath, tofilename.

    Returns:
        pd.DataFrame: Extended DataFrame with copied format-file records appended.
    """
    logger.info('combining format files')

    # Make sure column heading/locations are as expected and nothing was moved in Formatfile copy sheet
    check_ws_headers(copy_formatfile_list_sheet,
                     [('A1', 'use'),
                        ('B1', 'fromFilePath'),
                        ('C1', 'fromfilename'),
                        ('D1', 'tofilepath')])

    # append records from FORMAT file into dataframe
    for copy_formatfile_flag, from_path, from_fn, to_path, to_fn, *_ in islice(copy_formatfile_list_sheet, 1, None):
        # Filelist is a sheet in setup opened using openpyxl
        if str(copy_formatfile_flag.value).strip().lower() == "x":
            src_path, src_file_w_path, dest_path, \
                dest_file_w_path = assign_formatcopy_vars(from_path, from_fn, to_path, to_fn)

            # make sure file and directories exist
            bad_path_exit(src_path)
            bad_file_exit(file=Path(src_file_w_path))
            bad_path_exit(dest_path)
            bad_file_exit(file=Path(dest_file_w_path))

            logger.debug("Ready to combine copied format file in combine_formatfiles: ", src_file_w_path)

            # TODO somewhere in here need to set pull_group in file if we want to bring it from filelist
            df_one_file = pd.read_csv(src_file_w_path, nrows=ROV_SETUP['rows_to_read_limit'],
                                      keep_default_na=False)

            missing_split_fields = set(ROV_SETUP['splitfile_field_list']) - set(df_one_file.columns.tolist())
            # missing_split_fields = set(df_one_file.columns.tolist()) ^ set(orig_df.columns.tolist())
            if missing_split_fields:
                # should we exit if field lists dont match exactly?  just show mismatches? what if one has a ton?
                exit_yes((f"The input fields of the copied Formatfile: \n\n'{src_file_w_path}' "
                          f"\n\nare missing the following fields to Split:\n\n "
                          f"{', '.join(missing_split_fields)}"
                          ))
            # orig_df = orig_df.append(df_one_file, ignore_index=True)
            orig_df = pd.concat([orig_df, df_one_file], ignore_index=True)

    return orig_df


def process_format_files_into_combine() -> pd.DataFrame:
    """Concatenate all FORMATTED files marked for combine into a single DataFrame.

    Iterates the FileList sheet for rows flagged with 'x' in the combine column,
    reads each corresponding FORMATTED CSV, validates field order, and concatenates.
    Optionally appends externally copied format files via ``combine_formatfiles``.
    Removes fields listed in ``ROV_SETUP['combinefile_fields_to_delete_list']``.

    Returns:
        pd.DataFrame: Combined DataFrame of all flagged formatted files.
    """
    logger.info('starting process_format_files_into_combine')

    df_combined = pd.DataFrame()  # empty dataframe

    for fn, format_flag, combine_flag, update_fn, update_fields_cell, *_ \
            in islice(ROV_SETUP['filelist_sheet'], 3, None):
        # islice starts in row 3 to skip header and 2 variable defs

        if str(combine_flag.value).strip().lower() == "x":
            logger.debug(f"Ready to combine file '{fn.value}'")
            fnstem = Path(fn.value).stem
            formatfile_w_path = ROV_SETUP['format_path'] / ('FORMATTED ' + fnstem + '.csv')

            bad_file_exit(formatfile_w_path)

            df_one_file = pd.read_csv(formatfile_w_path, nrows=ROV_SETUP['rows_to_read_limit'], keep_default_na=False)

            if df_one_file.columns.tolist() != ROV_SETUP['formatfile_field_list']:
                exit_yes((f"The input fields of '{fn.value}' do not match specified OP fields:\n\n"
                          ",".join(df_one_file.columns.tolist())
                          ))
            # df_combined = df_combined.append(df_one_file, ignore_index=True)
            df_combined = pd.concat([df_combined, df_one_file], ignore_index=True)

    # if flag set to copy external Format files, copy them after to match fields
    if ROV_SETUP['copy_other_format_files_flag']:
        df_combined = combine_formatfiles(df_combined, ROV_SETUP['copy_formatfile_filelist_sheet'])

    # delete fields specified in setup
    try:
        df_combined.drop(ROV_SETUP['combinefile_fields_to_delete_list'], axis=1, inplace=True)
    except KeyError as e:
        logger.exception(e)
        fields_missing_from_formatfile = set(ROV_SETUP['combinefile_fields_to_delete_list']) - set(df_combined.columns)
        exit_yes("Formatfile is missing the following fields to delete when creating Combinefile:"
                 f"\n\n{', '.join(fields_missing_from_formatfile)}"
                 )

    return df_combined


def bek_text_box(txt: str, title: str = '', box_title: str = "",
                 buttons: Optional[list[str]] = None) -> Optional[str]:
    """Display a scrollable text block with selectable buttons using PySimpleGUI.

    Window width and height are auto-scaled to the text content within
    configured min/max limits. Scrolling is enabled when lines exceed 80.

    Args:
        txt (str): The body text to display, with lines separated by ``\\n``.
        title (str): Large heading displayed above the text body. Defaults to "".
        box_title (str): Title bar caption for the window. Defaults to "".
        buttons (Optional[list[str]]): Button labels to show at the bottom.
            Defaults to ``["OK", "Exit"]``.

    Returns:
        Optional[str]: The lowercased label of the button clicked, or None if
            the window was closed without a selection.
    """

    if buttons is None:
        buttons = ["OK", "Exit"]
    return confirm(txt, title=title or box_title, buttons=buttons)


def convert_xlsx_to_csvs() -> None:
    """Interactively convert XLSX files in a selected directory to CSV files.

    Prompts the user to select a source directory of XLSX files and a
    destination directory for CSVs. Skips any XLSX that already has a
    matching CSV. Prompts for confirmation before converting.
    """
    logger.info('starting convert_xlsx_to_csv')

    str_xls_dir = select_file("Select a DIRECTORY containing XLSX files to convert to CSVs",
                              start_dir=str(INITIAL_CAMPAIGN_DIR),
                              files_like="*",
                              mode="dir",
                              title2="XLSX Directory")
    if str_xls_dir is None:
        exit_yes("No directory selected.")
    xls_dir = Path(str_xls_dir)

    # str_xls_dir = "/Users/Denise/Library/CloudStorage/Dropbox/Postcard Files/TestInputFiles/TestCampaigns/TestXlsToCsv/Rawdata"

    str_csv_dir = select_file("Select a DIRECTORY where converted CSVs will be placed",
                              start_dir=str(xls_dir.parents[1]),
                              files_like="*",
                              mode="dir",
                              title2="CSV Directory (usually 'Rawdata')")
    if str_csv_dir is None:
        exit_yes("No directory selected.")
    # str_csv_dir = "/Users/Denise/Library/CloudStorage/Dropbox/Postcard Files/TestInputFiles/TestCampaigns/TestXlsToCsv/csv"

    xls_dir = Path(str_xls_dir)
    csv_dir = Path(str_csv_dir)

    # get list of files with xls or xlsx
    xls_files_w_path = list(xls_dir.glob("*.xls?"))

    # create list of dicts, each key a variable name corresponding to a file value,
    # eg xls_files_dict[0]['xls_name'] = 'abc.xlsx'.
    xls_files_dict = [{'xls_w_path': xls_w_path, 'xls_stem': xls_w_path.stem, 'xls_name': xls_w_path.name}
                      for xls_w_path
                      in xls_files_w_path]

    csv_files_w_path = list(csv_dir.glob("*.csv"))

    # same as done with xls above
    csv_files_dict = [{'csv_w_path': csv_w_path, 'csv_stem': csv_w_path.stem, 'csv_name': csv_w_path.name}
                      for csv_w_path
                      in csv_files_w_path]

    # set of file names (without .xls), csv below
    # xls_stem = {file_info['xls_stem'] for file_info in xls_files_dict}
    csv_stem = {file_info['csv_stem'] for file_info in csv_files_dict}

    # get lists of file_info for xls with and without csvs
    xls_wo_csv = [xls_info for xls_info in xls_files_dict if xls_info['xls_stem'] not in csv_stem]
    xls_w_csv = [xls_info for xls_info in xls_files_dict if xls_info['xls_stem'] in csv_stem]

    newline = '\n'  # cause can't use \n in f-string
    # prompt stating two lists - skip with_csv, process no_csv - ok to continue?
    exit_yes_no(f"The following files WILL BE converted to csvs: {newline}" +
                newline.join([file_info['xls_name'] for file_info in xls_wo_csv]) +
                f"{newline}{newline}The following files already HAVE CSVs and will be SKIPPED: {newline}" +
                newline.join([file_info['xls_name'] for file_info in xls_w_csv]),
                "Continue with convert?"
                )

    # read each xlsx into dataframe with options and write out with same name
    for file_info in xls_wo_csv:
        logger.info(f"Reading '{file_info['xls_name']}'")
        df = pd.read_excel(file_info['xls_w_path'], dtype=str)
        logger.info(f"Writing '{(file_info['xls_stem'] + '.csv')}'")
        df.to_csv(csv_dir / (file_info['xls_stem'] + '.csv'), index=False)
        logger.debug('')


def main_format() -> None:
    """Run the Format phase: create required directories, validate setup options, and format all flagged raw files.

    Reads setup from ``ROV_SETUP``, creates output directories under
    ``format_path``, confirms setup options, and calls ``process_format_files``.
    """

    logger.debug("in main_format")

    # check dir structure for formatted file creation. Don't create combine or split to indicate if dir is for
    # Format files copied in.
    bad_path_create(ROV_SETUP['format_path'])
    bad_path_create(ROV_SETUP['format_path'] / "Summary")
    bad_path_create(ROV_SETUP['format_path'] / "Removed")
    bad_path_create(ROV_SETUP['format_path'] / "Duplicates")

    # allow to exit if desired, eg flag not correct, imported code not right
    check_for_unwanted_setup_options()

    # Loop through all files in filelist to be formatted
    process_format_files(ROV_SETUP['filelist_sheet'])

    alert("Ran format section of main", "Alert")


def main_combine() -> None:
    """Run the Combine phase: merge formatted files, identify duplicates, run last code, and write the combined CSV.

    Reads setup from ``ROV_SETUP``, optionally copies external format files,
    runs the last-code block, identifies duplicates, sorts, runs pivot reports,
    and writes the final combined CSV.
    """

    logger.info("in main combine")

    bad_path_create(os.path.expanduser(ROV_SETUP['combined_path']))
    logger.debug("picked 'Combine'")

    if ROV_SETUP['copy_other_format_files_flag']:
        exit_yes_no("COPY FORMAT files from other directory in combine?  OK?\n\n",
                    'COPY FORMAT FILES?')

        # COPY FILES via parm 'perform_copies=True' in addition to checking files/path existence
        formatfile_copy(ROV_SETUP['copy_formatfile_filelist_sheet'], perform_copies=True)

    if ROV_SETUP['run_last_code_flag']:
        display_imported_code(ROV_SETUP['last_code_sheet'], ROV_SETUP['last_code'])

    # combine format files into combine if marked with 'x'
    df = process_format_files_into_combine()

    if ROV_SETUP['id_dupes']:
        logger.info(f"{ROV_SETUP['dupe_key_formula']=}")
        df['dupe_key'] = eval(ROV_SETUP['dupe_key_formula'])

        df.sort_values(by=[k for k, v in ROV_SETUP['dupe_key_sort_tuples']],
                       ascending=[v for k, v in ROV_SETUP['dupe_key_sort_tuples']], inplace=True)

        identify_duplicates(df, 'dupe_key', 'dupe_id_field')

        dupfile = ROV_SETUP['combined_path'] / f"DUPLICATES in {ROV_SETUP['OPFile'].stem}.csv"

        # set remove for dupes
        df.loc[~(df['dupe_id_field'].isin(['F','-','X'])), 'remove'] = \
            f"Not unique and not a first duplicate; not in 'F-X'"


    if ROV_SETUP['run_last_code_flag']:
        # *** Only run imported first_code and middle_code in format, not combine to keep things like remove
        # assignment, random number from being overwritten.
        # last_code can be run in combine since it's only setting remove code.
        logger.debug('Ready to run last_code (remove code)')  # these prompts help if error in imported code

        ROV_SETUP['last_code_to_import_module'].last_code_func(df, ROV_SETUP=ROV_SETUP)
        # This is the function from the sheet with any parameters it needs

        logger.debug('Ran last_code')

    # set remove based on max_pull_group if use_max_pull_group is set
    if ROV_SETUP['use_max_pull_group']:
        df.loc[(df['remove'] == '') & (df['pull_group'] < ROV_SETUP['max_pull_group']), 'remove'] = \
            f"Clean but pull group less than {ROV_SETUP['max_pull_group']}"

    if ROV_SETUP['id_dupes']:
        logger.debug("Going to write out duplicaes to file")
        logger.debug("Ready to sort by ['dupe_id_field','dupe_key']")  # to speed up copy to excel
        df.sort_values(by=['dupe_id_field', 'dupe_key'], inplace=True)

        logger.debug('Ready to copy dupes to CSV')  # these prompts help if error in imported code
        df[df['dupe_id_field'] != '-'].to_csv(dupfile, index=False)

    if ROV_SETUP['sort_list']:
        df.sort_values(by=ROV_SETUP['sort_list'], inplace=True)

    # run pivot reports on combined
    logger.debug('Ready to run pivots')  # these prompts help if error in imported code
    file = ROV_SETUP['combined_path'] / f"Summary of {ROV_SETUP['OPFile'].stem}.xlsx"
    pivot_and_other_reports(df,
                            file,
                            ROV_SETUP['OPFile'].stem + '.csv',
                            ROV_SETUP['dict_concentrated_addresses'])

    # Write out combined file
    combine_file = ROV_SETUP['combined_path'] / f"{ROV_SETUP['OPFile'].stem}.csv"
    df.to_csv(combine_file, index=False)

    pd.set_option('display.max_columns', None)
    logger.info("\n", df.head(5), "\n\n")

    alert("Ran combine section of main", "Alert")


def main() -> None:
    """Entry point: display the action menu and dispatch to the selected workflow.

    Presents a PySimpleGUI menu with choices: Format, Combine, Split,
    XLSXs to CSVs, Update Zip File, Compare CSV cols, Exit. For Format/Combine/Split,
    loads the setup file, initialises all setup dictionaries, and calls the
    appropriate workflow function.
    """

    logger.info("starting main")

    # shows all cols for dataframe head instead of truncating to first and last few
    pd.set_option('display.max_columns', None)

    ROV_SETUP['exe_path'] = exe_path()

    choice = bek_text_box("What do you want to do?", "Choose an Action",
                              '',
                              ['Format', 'Combine', 'Split', 'XLSXs to CSVs', 'Update Zip File', 'Compare CSV cols',
                               'Exit'])
    # PICK CHOICES
    if choice == 'xlsxs to csvs':
        logger.debug("chose 'XLSXs to CSVs'")

        convert_xlsx_to_csvs()
        alert("Ran convert_xlsx to csv", "Convert Xlsxs to CSVs")

    elif choice == 'update zip file':
        logger.debug("chose 'Update Zip File'")

        # Remind the user what must be downloaded from the vendor and where it must be
        # placed BEFORE the update runs.  'Continue' proceeds; 'Exit' quits so the files
        # can be put in place first.
        zip_dir = home_tilde(MAIN_ZIP_FILE.parent)
        dict_file = home_tilde(ZIP_TO_COUNTY_LIST_FILE)
        msg = (
            f"\n'Update Zip File' rebuilds the zip-to-county lookup from data purchased "
            f"from the vendor (Zip-Codes.com).\n\n"
            f"You will be updating two files:\n"
            f"  1. The main zip/county database:  '{MAIN_ZIP_FILE.name}'\n"
            f"  2. The multi-county (split-zip) database:  '{MULTI_COUNTY_ZIP_FILE.name}'\n\n"
            f"BEFORE continuing, log in to Zip-Codes.com and click 'dowload' under Subscriptions, then type 'csv'. "
            f"This will download a zip file containing a bunch of files, though you will only be concerned with the "
            f"two above.\n\n"
            f"Extract the zip (ususally to your 'downloads' directory).\n\n"
            f"In the destination directory\n\n"
            f"   '{zip_dir}'\n\n"
            f"move the old files '{MAIN_ZIP_FILE.name}' and '{MULTI_COUNTY_ZIP_FILE.name}' to '/old_zip_files' "
            f"for temporary holding in case things go amuck.\n\n"
            f"Move the two new files of the same name to '{zip_dir}'.\n\n"
            f"When you continue, these two files are read and a dictionary is written to:  '{dict_file}'  "
            f"which is used in future runs of cleaver.\n\n"
            f"Aside: 'Unique_County_List.xlsx' in the cleaver directory is refreshed every "
            f"time cleaver runs and lists the county names available for remapping.\n\n"
            f"Have both files been downloaded and placed in the folder above?\n\n"
        )

        exit_yes_no(msg, 'Update Zip Files')

        create_zip_to_county_list_dict(MAIN_ZIP_FILE, MULTI_COUNTY_ZIP_FILE, ZIP_TO_COUNTY_LIST_FILE)
        logger.debug("done 'Update Zip File'")
        alert("Ran Zip Dict file update", "Update zip files")

    elif choice == 'compare csv cols':
        logger.debug("chose 'Compare CSV cols'")
        compare_csv_columns.main_compare_csv_columns(INITIAL_CAMPAIGN_DIR)
        logger.debug("ran 'Compare CSV cols'")

    elif choice == 'exit':
        logger.debug("chose 'exit'")
        exit()

    else:
        # run code common to format and combine
        ROV_SETUP['setup_file_name'] = get_setup_file_name(INITIAL_CAMPAIGN_DIR)  # use TKInter to get the file/path of setup in campaign
        # ROV_SETUP['setup_file_name'] = PosixPath('/Users/Denise/Library/CloudStorage/Dropbox/Postcard '
        #                                           'Files/InputFiles/Campaigns/WI 01-25 State General/cleaver WI 01-25 State General UniversalSetup.xlsx')
        logger.debug("in 'else' to pick format, combine, split")

        # FIXME: somewhere in here set campaign, factory nd group_vars?
        init_setup_dict()
        # setup_backward_compatability()
        read_setup_vars(FIELD_DEF_COL_NUMERIC)
        format_setup_vars()
        create_field_lists()  # fill array with default field names of ' ' and add fields required by options selected
        create_dicts()  # create dicts and other file set up needed to run

        if choice == 'format':
            main_format()

        elif choice == 'combine':
            main_combine()

        elif choice == 'split':
            logger.debug("picked 'Split'")

            bad_path_create(ROV_SETUP['split_path'])
            bad_path_create(ROV_SETUP['split_path_hold'])
            bad_path_create(ROV_SETUP['split_path_done'])

            split_files_for_sincere(ROV_SETUP['combined_path'] / (ROV_SETUP['OPFile'].stem + '.csv'),
                                    ROV_SETUP['sub_split_limit'],
                                    ROV_SETUP['split_path'] / f"Split file list-{ROV_SETUP['expectedstate']}"
                                                             f" {ROV_SETUP['splitfnbase'].stem}.xlsx")
            logger.info('ran split')
            alert("Ran split section of main", "Alert")


def init_setup_dict() -> None:
    """Initialise ``ROV_SETUP`` with workbook objects, fixed paths, and file names from the setup XLSX.

    Opens the setup workbook, assigns sheet objects (Setup, FileList, FormatCopies),
    derives all standard directory paths (rawdata, formatted, split, combined),
    reads pull-group settings, and validates FileList column headers.
    """

    logger.info('starting init_setup_dict')

    ROV_SETUP['setup_wb'] = load_workbook_w_filepath(ROV_SETUP['setup_file_name'])
    ROV_SETUP['setup_sheet'] = ROV_SETUP['setup_wb']["Setup"]

    ROV_SETUP['filelist_sheet'] = ROV_SETUP['setup_wb']["FileList"]
    ROV_SETUP['use_max_pull_group'] = convert_bool(ROV_SETUP['setup_wb']["FileList"]['B1'].value)
    if ROV_SETUP['use_max_pull_group']:  # only read max_pull_group if using it for check
        ROV_SETUP['max_pull_group'] = ROV_SETUP['setup_wb']["FileList"]['B2'].value
        if not is_number(ROV_SETUP['max_pull_group']):
            msg = f"max_pull_group must be number but it is '{ROV_SETUP['max_pull_group']}'"
            logger.error(msg)
            exit_yes(msg)

    ROV_SETUP['copy_formatfile_filelist_sheet'] = ROV_SETUP['setup_wb']["FormatCopies"]

    ROV_SETUP['exe_path'] = exe_path()
    ROV_SETUP['root_path'] = ROV_SETUP['setup_file_name'].parent  # dir containing the setup file

    # use new function formatcopy test paths.py and path.parents[0]
    # ROV_SETUP['root_path_one_level_up'] = ROV_SETUP['root_path'].parent

    ROV_SETUP['rawdata_path'] = ROV_SETUP['root_path'] / 'Rawdata'
    ROV_SETUP['format_path'] = ROV_SETUP['root_path'] / 'Formatted'
    ROV_SETUP['split_path'] = ROV_SETUP['root_path'] / 'Split'
    ROV_SETUP['split_path_hold'] = ROV_SETUP['root_path'] / 'Split' / 'Hold'
    ROV_SETUP['split_path_done'] = ROV_SETUP['root_path'] / 'Split' / 'Done'
    ROV_SETUP['combined_path'] = ROV_SETUP['root_path'] / 'Combined'
    ROV_SETUP['op_path'] = ROV_SETUP['root_path'] / 'UpdateFiles'

    # exit if no path exists for rawdata
    bad_path_exit(ROV_SETUP['rawdata_path'])

    # hardcoded file names created from first_code, middle_code, and last_code sheets
    ROV_SETUP['first_code'] = "first_code_to_import.py"
    ROV_SETUP['middle_code'] = "middle_code_to_import.py"
    ROV_SETUP['last_code'] = "last_code_to_import.py"

    # Check heading fields in filelist sheet of setup file to make sure it didn't move/change
    check_ws_headers(ROV_SETUP['filelist_sheet'],
                     [('A3', 'file'),
                        ('B3', 'formatfile'),
                        ('C3', 'concatfile'),
                        ('D3', 'updatefile'),
                        ('E3', 'updatefilenames'),
                        ('F3', 'pull_group'),
                        ('G3', 'custom_field'),
                     ])


def setup_backward_compatability() -> None:
    """Verify that the setup file meets minimum version requirements.

    Checks that the FileList sheet contains expected header text indicating a
    compatible setup file version. Exits with an informative message if not.
    """

    # if use_pull_flag and max_pull_group not in file we can not set variables and continue because first row of
    # filelist sheet will be off.

    # Below code does not work because would have to be run before init_setup_dict() (where paths are set),
    # but headers in filelist are also checked there.
    if ROV_SETUP['setup_wb']["FileList"]['A1'].value != "remove using max_pull_group":
        msg = f"Cell 'A1' in filelist must be 'remove using max_pull_group' but it " \
              f"is '{ROV_SETUP['setup_wb']['FileList']['A1'].value}'.\n\n"\
                "You must a more recent setup file"
        logger.error(msg)
        exit_yes(msg)

def read_setup_vars(field_col: int) -> None:
    """Iterate all setup-sheet rows and populate ``ROV_SETUP`` with parsed variable values.

    Skips the header row (row 1). For each subsequent row, calls ``read_setup_var``
    with the row converted to a list via ``row_to_list``.

    Args:
        field_col (int): Zero-indexed column number of the field-definition column
            (e.g. ``FIELD_DEF_COL_NUMERIC``).
    """
    logger.info('starting read_setup_vars')

    # for testing read one row
    # row = list(ROV_SETUP['setup_sheet'].rows)[37]  # 4, 7, 12, 18, 37, 97, 140, 41 pythobj
    # logger.debug(f"{row[5].value=}")
    # row_list = row_to_list(row)
    # read_setup_var(ROV_SETUP, row_list)
    # logger.info(f"{ROV_SETUP=}")
    # exit()

    # returns a tuple of 0-indexed cell references.  row[0] = 1, first_row[0] is 'A'. use .value to get value in cell.
    for row in islice(list(ROV_SETUP['setup_sheet'].rows), 1, None):
        # print(f"{row[1].value=}")
        row_list = row_to_list(row)
        read_setup_var(row_list)

def read_setup_var(row_data: list) -> None:
    """Parse one setup-sheet row and store its variable(s) in ``ROV_SETUP``.

    Reads the field-definition string from the column at ``FIELD_DEF_COL_NUMERIC``,
    validates its format, determines whether the row defines a list or scalar
    variable(s), converts each value using the declared type, and assigns the
    result(s) to ``ROV_SETUP``.

    Args:
        row_data (list): Cell values for one setup-sheet row, zero-indexed.
    """

    logger.debug(f"starting read_setup_var {row_data[ min(len(row_data)-1,FIELD_DEF_COL_NUMERIC) ]}")

    def len_tuple(tuple: Any) -> int:
        """Return the length of an object, or -99 if ``len()`` raises a TypeError.

        Args:
            tuple (Any): Any object to measure.

        Returns:
            int: ``len(tuple)``, or ``-99`` if the object has no length.
        """
        try:
            len(tuple)
        except:
            return -99
        return len(tuple)

    def return_func(var_type: str, str_case: str = 'l',
                    str_strip: str = 'b', **kwargs) -> Callable:
        """Return a conversion function corresponding to a setup variable type string.

        Args:
            var_type (str): Type identifier; one of ``'str'``, ``'int'``,
                ``'bool'``, ``'float'``, ``'file'``, ``'pythobj'``.
            str_case (str): Case conversion for ``'str'`` type - ``'l'`` lower,
                ``'u'`` upper, ``'k'`` keep. Defaults to ``'l'``.
            str_strip (str): Strip mode for ``'str'`` type - ``'b'`` both,
                ``'l'`` left, ``'r'`` right. Defaults to ``'b'``.
            **kwargs: Additional keyword arguments forwarded to the returned function.

        Returns:
            Callable: A function that converts a raw cell string to the target type.

        Raises:
            ValueError: If ``var_type`` is not one of the recognised type strings.
        """

        def my_expanduser(file_str: str) -> Path:
            """Convert a file path string to a Path and expand ``~`` to the home directory.

            Args:
                file_str (str): File path string, possibly starting with ``~``.

            Returns:
                Path: Resolved Path with home directory expanded.
            """
            p = Path(file_str)
            return p.expanduser()

        def my_python_obj(code_str: str) -> Any:
            """Parse a JSON-formatted string into a Python object.

            Limitations: double quotes are not supported; tuples must be expressed as
            lists; boolean literals must be lowercase (``true``/``false``).

            Args:
                code_str (str): A JSON-compatible string representing a Python object.

            Returns:
                Any: The parsed Python object (list, dict, int, bool, etc.).
            """
            python_object = json.loads(code_str)
            return python_object

        def my_lower_strip(valx: Optional[str], *extra_args,
                           str_case: str = 'l', str_strip: str = 'b',
                           **extra_kwargs) -> str:
            """Apply strip and case conversion to a string value.

            None input is treated as an empty string.

            Args:
                valx (Optional[str]): The string to process, or None.
                *extra_args: Ignored positional arguments (for signature compatibility).
                str_case (str): Case conversion - ``'l'`` lower, ``'u'`` upper,
                    ``'k'`` keep original. Defaults to ``'l'``.
                str_strip (str): Strip mode - ``'b'`` both ends, ``'l'`` left,
                    ``'r'`` right. Defaults to ``'b'``.
                **extra_kwargs: Ignored keyword arguments (for signature compatibility).

            Returns:
                str: Processed string value.

            Raises:
                Exception: If ``str_strip`` or ``str_case`` is not a recognised option.
            """
            val = valx
            if val is None:
                val = ''  # TODO is this better None?
            elif str_strip == 'b':
                val = val.strip()
            elif str_strip == 'l':
                val = val.lstrip()
            elif str_strip == 'r':
                val = val.rstrip()
            else:
                raise Exception  # 'Invalid option for str_strip parameter'

            if val is None:
                val = ''  # TODO is this better None?
            elif str_case == 'l':
                val = val.lower()
            elif str_case == 'u':
                val = val.upper()
            elif str_case == 'k':  # (k)eep the same
                pass
            else:
                raise Exception  # 'Invalid option for str_lower parameter'

            return val

        if var_type == 'str':
            return_fnc = my_lower_strip
        elif var_type == 'int':
            return_fnc = int
        elif var_type == 'bool':
            return_fnc = convert_bool
        elif var_type == 'float':
            return_fnc = float
        elif var_type == 'file':
            return_fnc = my_expanduser
        elif var_type == 'pythobj':
            return_fnc = my_python_obj
        else:
            raise ValueError
        return return_fnc

    def check_vars_string_for_errors(vars_string: str) -> list:
        """Validate and parse the field-definition string from a setup-sheet cell.

        The string must evaluate to a list whose first element is a bool (True for
        a list-valued variable, False for one or more scalars), followed by
        ``(var_name, var_type[, options_dict])`` tuples. Recognised type strings:
        ``'str'``, ``'int'``, ``'float'``, ``'bool'``, ``'file'``, ``'pythobj'``.

        Args:
            vars_string (str): Raw cell text from the field-definition column.

        Returns:
            list: Validated variable specification list with the bool flag as
                element 0 and ``[var_name, var_type, ...]`` sublists for each variable.

        Raises:
            Exception: If the string fails to eval, is not a list, lacks a bool
                first element, has fewer than two elements, has conflicting list/multi
                specification, or contains unrecognised type strings.
        """
        try:
            vars_evaled = eval(vars_string)
        except Exception as e:
            logger.exception(e)
            logger.info(f"Field_vals '{vars_string}' is not a valid format.  check missing quotes.")
            raise Exception  # FIXME print this f"Field_vals '{vars_string}' is not a valid format.  check missing
            # quotes."

        # TODO optional/missing dicts caused problems in list comp so used loop.  way to use list comp?
        vars_list = [vars_evaled[0]]  # [0] is boolean which indicates if list of variables across sheet row
        for var, var_type, *other in islice(vars_evaled, 1, None):  #
            if not other:
                vars_list.append([var.strip(), var_type.strip()])
            else:
                vars_list.append([var.strip(), var_type.strip()] + other)

        # must be tuple or list
        if not isinstance(vars_list, list):
            logger.info(
                f"Field_vals '{vars_string}' is not an expected tuple 'True,('xxx','yyy'),('aaa','bbb')...' format")
            raise Exception  # catch one explicit tuple like (a,int) as opposed to

        # first element must be bool (whether var is list of not)
        if not isinstance(vars_list[0], bool):
            logger.info(
                f"First part of Field_vals '{vars_string}' must be True=field with list of values or False=one or more "
                f"single variables "
                f"'('var', 'type')'")
            raise Exception  # catch one explicit tuple like (a,int) as opposed to

        # must have at least two element, bool for list and one variable specification
        if len(vars_list) < 2:
            logger.info(
                f"Field_vals '{vars_string}' must have at least two elements, the first True/False 'True,('xxx','yyy')...)")
            raise Exception  # catch one explicit tuple like (a,int) as opposed to

        # list (bool = True) but more than one var and type specified
        if vars_list[0] and len(vars_list) != 2:  # can only have one variable definition if type is a list (True)
            logger.info(
                f"If field_type is a list (first element is True) '{vars_string}' can only have one other element "
                f"('var_name','var_type') and all values must be of that type")
            raise Exception

        # check that var file_types are in accepted list
        bad_tuples = [(var, my_type) for var, my_type, *other
                      in islice(vars_list, 1, None)
                      if not isinstance(var, str) or my_type not in ['str', 'int', 'float', 'bool', 'file', 'pythobj']]
        # started adding 'pyth' for python code or objects but code needed to execute (eg have df defined)
        if bad_tuples:
            logger.info(
                f"Field_vals var/type in '{vars_string}' must have two elements, the first (the variable name) a string, "
                f"the second, the type, either str, int, float, file, or bool")
            raise Exception  # catch one explicit tuple like (a,int) as opposed to

        return vars_list

    if len(row_data) <= FIELD_DEF_COL_NUMERIC:  # no field data in field_vals column
        pass
    else:
        vars_string = row_data[FIELD_DEF_COL_NUMERIC]
        if vars_string is None:  # skip row - no variable info specified - blank or info row
            pass
        # only keep if field_keep is true
        elif str(row_data[FIELD_DEF_COL_NUMERIC - 1]).lower().strip() != 'true':
            pass
        else:
            # check structure and var tuples for input errors
            vars_list = check_vars_string_for_errors(vars_string)  # raises exceptions it error found

            if vars_list[0]:  # if True process list (first item is whether vars a list of values)
                my_dict = ({} if len(vars_list[1]) < 3 else vars_list[1][2])  # len()<3 means no dict if dict not
                # supplied
                func = return_func(vars_list[1][1], **my_dict)  # elements 1 (type) and 2(dict) for first and only var
                row_list = []
                for cell_val in islice(row_data, DATA_STARTS_COL_NUMERIC, None):  # TODO replace loop with list comprehension
                    if cell_val is not None:
                        row_list.append(func(cell_val, **my_dict))
                    else:
                        row_list.append("")
                ROV_SETUP[vars_list[1][0]] = row_list
            else:
                for index, (field_name, field_type, *my_dict) in enumerate(list(islice(vars_list, 1, None))):
                    my_dict = ({} if my_dict == [] else my_dict[0])
                    func = return_func(field_type, **my_dict)
                    ROV_SETUP[field_name] = func(row_data[DATA_STARTS_COL_NUMERIC + index], **my_dict)



def format_setup_vars() -> None:
    """Post-process raw ``xl_*`` setup values into typed, derived ``ROV_SETUP`` entries.

    Assigns worksheet objects, pads field lists to equal length, parses
    campaign/factory/group variable lists, validates for overlap, loads the
    concentrated-addresses workbook, determines the sort list from ``sortchoice``,
    compiles pivot specifications, and optionally removes ``xl_`` prefixed keys.
    """
    logger.debug(f"in format_setup_vars top {id(ROV_SETUP)=}")
    logger.info(f"starting format_setup_vars")

    # put worksheet object of string in ROV_DICT
    if ROV_SETUP['skip_selected_zip_match_flag']:
        ROV_SETUP['skip_selected_zip_sheet'] = ROV_SETUP['setup_wb'][ROV_SETUP['xl_skip_selected_zip_sheet']]
    ROV_SETUP['first_code_sheet'] = ROV_SETUP['setup_wb'][ROV_SETUP['xl_first_code_sheet']]
    ROV_SETUP['middle_code_sheet'] = ROV_SETUP['setup_wb'][ROV_SETUP['xl_middle_code_sheet']]
    ROV_SETUP['last_code_sheet'] = ROV_SETUP['setup_wb'][ROV_SETUP['xl_last_code_sheet']]

    max_len = max(len(ROV_SETUP['xl_inputfile_orig_list']),
                  len(ROV_SETUP['xl_inputfile_renamed_list']),
                  len(ROV_SETUP['xl_inputfile_type_list']),
                  )

    ROV_SETUP['inputfile_orig_list'] = pad_list(ROV_SETUP['xl_inputfile_orig_list'], max_len, pad_val="")
    ROV_SETUP['inputfile_renamed_list'] = pad_list(ROV_SETUP['xl_inputfile_renamed_list'], max_len, pad_val="")
    ROV_SETUP['inputfile_type_list'] = pad_list(ROV_SETUP['xl_inputfile_type_list'], max_len, pad_val="")

    ROV_SETUP['campaign_vars'] = [fld.strip() for fld
                                 in ROV_SETUP['xl_campaign_vars'].split(',') if fld != '']
    ROV_SETUP['factory_vars'] = [fld.strip() for fld
                         in ROV_SETUP['xl_factory_vars'].split(',') if fld != '']

    # check if overlapping variables
    overlap_fields = set(ROV_SETUP['campaign_vars']).intersection(set(ROV_SETUP['factory_vars']))
    if overlap_fields:
        msg = (f"The following variable(s) overlap between the factory and campaign splits:"
               f"\n\n{', '.join(overlap_fields)}"
              )
        logger.error(msg)
        exit_yes(msg)

    ROV_SETUP['group_vars'] = ROV_SETUP['campaign_vars'] + ROV_SETUP['factory_vars']
    # TODO can we check if group_vars are in fieldlist?  What if some are created in middle_code - are they
    #  available? when are they added to list?

    ROV_SETUP['concentrated_addresses_wb'] = load_workbook_w_filepath(ROV_SETUP['concentrated_addresses_file'])
    ROV_SETUP['concentrated_addresses_sheet'] = ROV_SETUP['concentrated_addresses_wb']["Addresses"]  # sheet "Addresses" hardcoded


    if ROV_SETUP['run_merge_data_flag']:
        bad_path_exit(ROV_SETUP['rawdata_path'])
    if ROV_SETUP['sortchoice'] == 1:
        ROV_SETUP['sort_list'] = ROV_SETUP['group_vars'] + ['remove', 'zip', 'address', 'randnum']
    elif ROV_SETUP['sortchoice'] == 2:
        ROV_SETUP['sort_list'] = ROV_SETUP['group_vars'] + ['remove', 'zip', 'randnum']
    elif ROV_SETUP['sortchoice'] == 3:
        ROV_SETUP['sort_list'] = ROV_SETUP['group_vars'] + ['remove', 'randnum']
    else:
        ROV_SETUP['sort_list'] = []

    # combine pivot info into an object we can loop through
    ROV_SETUP['pivot_specs'] = []
    for piv_num in islice(range(6), 1, None):
        d = dict()
        d['pivot_fields'] = [fld.strip() for fld
                             in ROV_SETUP['xl_str_pivot_field' + str(piv_num)].split(',') if fld != '']
        d['pivot_by_cnt'] = ROV_SETUP['xl_str_pivot_field' + str(piv_num) + '_by_cnt']
        d['pivot_for_all'] = ROV_SETUP['xl_str_pivot_field' + str(piv_num) + '_all']
        ROV_SETUP['pivot_specs'].append(d)

    # sort dictionary DO NOT DO THIS - creates a new object even though reference is the same! https://stackoverflow.com/questions/61645769/sort-dict-in-place
    # local_dict = dict(sorted(local_dict.items()))

    # delete entries for master_dict that are temporary (begin with xl_)
    logger.info(f"{id(ROV_SETUP)=}")
    if REMOVE_XL_FROM_SETUP:
        # noinspection PyUnreachableCode
        for k, v in list(ROV_SETUP.items()):
            if k.startswith('xl_'):
                del ROV_SETUP[k]

    logger.debug(f"in format_setup_vars - finished")


if __name__ == '__main__':

    main()
    a = 1
